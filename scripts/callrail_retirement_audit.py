"""
CallRail retirement audit: find CallRail-linked clients who haven't booked in 6+ months
(or 12+ months) so we can retire their tracking numbers and stop paying for them.

Three sheets:
  1. Over 6 months expired — last order 180-365 days ago
  2. Over 1 year expired   — last order 365+ days ago (or never)
  3. Summary stats

Each row includes: client name, market, status, CallRail name/ID, last order date,
days since last order, last call date, calls in last 12 months.
"""

import os
import sys
from datetime import date
from pathlib import Path
from dotenv import load_dotenv

load_dotenv(override=True)
from supabase import create_client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PROJECT_ROOT = Path(__file__).resolve().parent.parent
OUTPUT = PROJECT_ROOT / "output" / f"[C] CallRail Retirement Audit {date.today().isoformat()}.xlsx"

sb = create_client(os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_KEY"))


def fetch_all(table, sel, **filt):
    rows, off = [], 0
    while True:
        q = sb.table(table).select(sel).range(off, off + 999)
        for k, v in filt.items():
            q = q.eq(k, v)
        b = q.execute().data
        rows.extend(b)
        if len(b) < 1000:
            break
        off += 1000
    return rows


print("Fetching CallRail platform IDs...")
cr_ids = fetch_all("client_platform_ids", "client_id,external_id,external_name", platform="callrail")
print(f"  {len(cr_ids)} CallRail-linked clients")

print("Fetching clients...")
clients = fetch_all("clients", "id,name,status,primary_market_id")
client_map = {c["id"]: c for c in clients}

print("Fetching markets...")
markets = {m["id"]: m["code"] for m in sb.table("markets").select("id,code").execute().data}

print("Fetching orders (max date per client)...")
last_order = {}
offset = 0
while True:
    batch = sb.table("orders").select("client_id,issue_date_parsed").range(offset, offset + 999).execute().data
    for r in batch:
        cid, d = r.get("client_id"), r.get("issue_date_parsed")
        if not cid or not d:
            continue
        if cid not in last_order or d > last_order[cid]:
            last_order[cid] = d
    if len(batch) < 1000:
        break
    offset += 1000

print("Fetching recent calls (last 12 months)...")
from datetime import timedelta
cutoff = (date.today() - timedelta(days=365)).isoformat()
calls_12mo = {}
last_call_date = {}
offset = 0
while True:
    batch = (sb.table("calls").select("client_id,call_time")
             .gte("call_time", cutoff).range(offset, offset + 999).execute().data)
    for r in batch:
        cid = r.get("client_id")
        d = r.get("call_time")
        if not cid:
            continue
        calls_12mo[cid] = calls_12mo.get(cid, 0) + 1
        if d and (cid not in last_call_date or d > last_call_date[cid]):
            last_call_date[cid] = d
    if len(batch) < 1000:
        break
    offset += 1000

print(f"  {sum(calls_12mo.values())} calls for {len(calls_12mo)} CallRail clients")

TODAY = date.today()


def make_row(cr):
    c = client_map.get(cr["client_id"])
    if not c:
        return None
    last = last_order.get(cr["client_id"])
    days_since = (TODAY - date.fromisoformat(last)).days if last else None
    lc = last_call_date.get(cr["client_id"], "")
    lc_short = lc[:10] if lc else ""
    return {
        "client_name": c["name"],
        "market": markets.get(c.get("primary_market_id"), ""),
        "status": c.get("status", ""),
        "callrail_name": cr.get("external_name", ""),
        "callrail_id": cr.get("external_id", ""),
        "last_order_date": last or "",
        "days_since_last_order": days_since if days_since is not None else "",
        "last_call_date": lc_short,
        "calls_last_12mo": calls_12mo.get(cr["client_id"], 0),
    }


# Classify
bucket_6mo = []        # 180-365 days since last order
bucket_12mo = []       # 365+ days OR no orders ever
active_or_recent = []  # < 180 days (keep)

for cr in cr_ids:
    row = make_row(cr)
    if not row:
        continue
    days = row["days_since_last_order"]
    if days == "":
        bucket_12mo.append({**row, "category": "Never ordered"})
    elif days >= 365:
        bucket_12mo.append({**row, "category": "Dormant (1+ yr)"})
    elif days >= 180:
        bucket_6mo.append(row)
    else:
        active_or_recent.append(row)

bucket_6mo.sort(key=lambda r: r["days_since_last_order"], reverse=True)
bucket_12mo.sort(key=lambda r: (r["days_since_last_order"] if isinstance(r["days_since_last_order"], int) else 99999), reverse=True)

print(f"\nBucket 1 (6-12 months expired): {len(bucket_6mo)}")
print(f"Bucket 2 (1+ year expired or never ordered): {len(bucket_12mo)}")
print(f"Still active/recent (<6mo): {len(active_or_recent)}")


# ---- Build Excel ----
wb = Workbook()

NAVY = "1A3A5C"
RED = "C0392B"
ORANGE = "D95D39"
GRAY = "F2F4F7"

header_font = Font(bold=True, color="FFFFFF")
header_fill_navy = PatternFill("solid", fgColor=NAVY)
header_fill_orange = PatternFill("solid", fgColor=ORANGE)
header_fill_red = PatternFill("solid", fgColor=RED)
zebra_fill = PatternFill("solid", fgColor=GRAY)
thin = Side(border_style="thin", color="D0D5DD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = [
    ("Client Name", "client_name", 38),
    ("Market", "market", 8),
    ("Status", "status", 11),
    ("CallRail Account Name", "callrail_name", 38),
    ("CallRail Company ID", "callrail_id", 38),
    ("Last Order Date", "last_order_date", 14),
    ("Days Since Last Order", "days_since_last_order", 14),
    ("Last Call", "last_call_date", 12),
    ("Calls (last 12mo)", "calls_last_12mo", 14),
]


def write_sheet(ws, rows, title, header_fill, extra_col=None):
    headers = [c[0] for c in COLS]
    if extra_col:
        headers.append(extra_col[0])
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

    for i, r in enumerate(rows):
        row_vals = [r.get(c[1], "") for c in COLS]
        if extra_col:
            row_vals.append(r.get(extra_col[1], ""))
        ws.append(row_vals)
        if i % 2 == 1:
            for cell in ws[ws.max_row]:
                cell.fill = zebra_fill
        for cell in ws[ws.max_row]:
            cell.border = border

    for idx, (_, _, width) in enumerate(COLS):
        ws.column_dimensions[chr(65 + idx)].width = width
    if extra_col:
        ws.column_dimensions[chr(65 + len(COLS))].width = extra_col[2]
    ws.freeze_panes = "A2"
    ws.title = title


# Summary first
ws0 = wb.active
ws0.title = "Summary"
ws0.append(["CallRail Retirement Audit", ""])
ws0["A1"].font = Font(bold=True, size=14, color=NAVY)
ws0.append(["Generated", date.today().isoformat()])
ws0.append(["Total CallRail-linked clients", len(cr_ids)])
ws0.append([])
ws0.append(["Bucket", "Count", "Recommendation"])
for cell in ws0[5]:
    cell.font = header_font
    cell.fill = header_fill_navy
ws0.append(["Over 6 months expired (180-365 days since last order)", len(bucket_6mo), "Review for retirement — likely cancelled"])
ws0.append(["Over 1 year expired or never ordered", len(bucket_12mo), "Strong candidates to retire immediately"])
ws0.append(["Still active/recent (< 6 months)", len(active_or_recent), "Keep"])
ws0.column_dimensions["A"].width = 55
ws0.column_dimensions["B"].width = 10
ws0.column_dimensions["C"].width = 45

# 6-12 month bucket
ws1 = wb.create_sheet()
write_sheet(ws1, bucket_6mo, "6-12 Months Expired", header_fill_orange)

# 1+ year bucket (with category column showing "Dormant" vs "Never ordered")
ws2 = wb.create_sheet()
write_sheet(ws2, bucket_12mo, "1+ Year Expired or Never", header_fill_red,
            extra_col=("Reason", "category", 20))

OUTPUT.parent.mkdir(exist_ok=True)
wb.save(OUTPUT)
print(f"\nSaved: {OUTPUT}")
