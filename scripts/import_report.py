"""
Post-import report: client status sync, change log, and email notification.

Can be called standalone (generates report from current DB state) or wired
into import_orders.py with explicit change data.

Usage:
  python scripts/import_report.py                    # Generate report + email from DB state
  python scripts/import_report.py --no-email         # Report only, no email
  python scripts/import_report.py --test-email       # Send a short test email
"""

import sys
import os
import argparse
import smtplib
from pathlib import Path
from datetime import date, datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
load_dotenv(override=True)

from supabase import create_client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
GMAIL_USER = os.getenv("GMAIL_USER")
GMAIL_APP_PASSWORD = os.getenv("GMAIL_APP_PASSWORD")
NOTIFY_EMAILS = [e.strip() for e in (os.getenv("NOTIFY_EMAILS") or GMAIL_USER or "").split(",") if e.strip()]

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)


def fetch_all(sb, table, select, batch=1000, **filters):
    rows, offset = [], 0
    while True:
        q = sb.table(table).select(select).range(offset, offset + batch - 1)
        for k, v in filters.items():
            q = q.eq(k, v)
        data = q.execute().data
        rows.extend(data)
        if len(data) < batch:
            break
        offset += batch
    return rows


# ---------- Status sync ----------
#
# Status rules (all derived from orders.issue_date_parsed):
#   active    - max(issue_date) >= today  (current or future order)
#   cancelled - last order within past 90 days
#   expired   - last order 90-365 days ago
#   dormant   - last order over a year ago
#   prospect  - zero orders ever
#
# Legacy 'inactive' is still allowed by the check constraint but no longer set.

from datetime import timedelta
from collections import defaultdict


def _compute_status(last_date_str, today_str, c90_str, c365_str):
    if not last_date_str:
        return "prospect"
    if last_date_str >= today_str:
        return "active"
    if last_date_str >= c90_str:
        return "cancelled"
    if last_date_str >= c365_str:
        return "expired"
    return "dormant"


def sync_client_statuses(sb, dry_run=False):
    """Recompute every client's status from order data.
    Returns dict keyed by transition: {('old','new'): [client_dicts]}."""
    print("Syncing client statuses from order data...")

    today = date.today()
    today_s = today.isoformat()
    c90 = (today - timedelta(days=90)).isoformat()
    c365 = (today - timedelta(days=365)).isoformat()

    clients = fetch_all(sb, "clients", "id,name,status,primary_market_id")

    # Max issue date per client (single scan of orders)
    max_date = {}
    offset = 0
    while True:
        batch = (
            sb.table("orders")
            .select("client_id,issue_date_parsed")
            .range(offset, offset + 999)
            .execute()
            .data
        )
        for r in batch:
            cid = r.get("client_id")
            d = r.get("issue_date_parsed")
            if not cid or not d:
                continue
            cur = max_date.get(cid)
            if cur is None or d > cur:
                max_date[cid] = d
        if len(batch) < 1000:
            break
        offset += 1000

    # Group client_ids needing update by new_status
    to_update = defaultdict(list)
    transitions = defaultdict(list)
    for c in clients:
        cid = c["id"]
        old = c.get("status")
        new = _compute_status(max_date.get(cid), today_s, c90, c365)
        if old != new:
            to_update[new].append(cid)
            transitions[(old or "null", new)].append(c)

    total_changed = sum(len(v) for v in to_update.values())
    print(f"  Clients changing status: {total_changed} / {len(clients)}")
    for (old, new), rows in sorted(transitions.items(), key=lambda x: -len(x[1])):
        print(f"    {old:10s} -> {new:10s} {len(rows):5d}")

    if not dry_run and to_update:
        for new_status, ids in to_update.items():
            for i in range(0, len(ids), 200):
                chunk = ids[i:i + 200]
                sb.table("clients").update({"status": new_status}).in_("id", chunk).execute()

    return transitions


# ---------- Expiring orders ----------

def find_expiring_clients(sb, days=60):
    """Find clients whose last future order is within N days (expiring soon)."""
    cutoff = date.today().isoformat()
    from datetime import timedelta
    horizon = (date.today() + timedelta(days=days)).isoformat()

    # Get clients with their max future issue_date
    clients_with_orders = fetch_all(sb, "clients", "id,name,primary_market_id,status")
    client_map = {c["id"]: c for c in clients_with_orders}

    markets = {m["id"]: m["code"] for m in sb.table("markets").select("id,code").execute().data}

    # Get max issue_date per client for future orders
    expiring = []
    offset = 0
    last_order_by_client = {}
    while True:
        batch = (
            sb.table("orders")
            .select("client_id,issue_date_parsed")
            .gte("issue_date_parsed", cutoff)
            .range(offset, offset + 999)
            .execute()
            .data
        )
        for r in batch:
            cid = r["client_id"]
            d = r["issue_date_parsed"]
            if cid and d:
                if cid not in last_order_by_client or d > last_order_by_client[cid]:
                    last_order_by_client[cid] = d
        if len(batch) < 1000:
            break
        offset += 1000

    for cid, last_date in last_order_by_client.items():
        if last_date <= horizon:
            c = client_map.get(cid)
            if c:
                mkt = markets.get(c.get("primary_market_id"), "?")
                expiring.append({
                    "name": c["name"],
                    "market": mkt,
                    "last_order": last_date,
                })

    expiring.sort(key=lambda x: x["last_order"])
    return expiring


# ---------- Change log xlsx ----------

def write_change_log(new_clients, activated, cancelled_clients, expiring, removed_count, total_orders, path):
    """cancelled_clients: list of dicts {name, market, orders_removed, gross_lost}"""
    wb = Workbook()
    hfont = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="4472C4")

    def style_header(ws):
        for cell in ws[1]:
            cell.font = hfont
            cell.fill = hfill

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Count"])
    style_header(ws)
    ws.append(["Import date", date.today().isoformat()])
    ws.append(["Total orders in DB", total_orders])
    ws.append(["New clients created", len(new_clients)])
    ws.append(["Clients activated (new current/future orders)", len(activated)])
    ws.append(["Clients with orders cancelled this week", len(cancelled_clients)])
    ws.append(["Total orders cancelled this week", removed_count])
    ws.append(["Orders expiring within 60 days", len(expiring)])
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 14

    # New clients
    ws2 = wb.create_sheet("New Clients")
    ws2.append(["Client", "Market", "Category"])
    style_header(ws2)
    for c in new_clients:
        ws2.append([c.get("name", ""), c.get("market", ""), c.get("category", "")])
    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 35

    # Cancelled this week (orders removed during import)
    ws3 = wb.create_sheet("Cancelled This Week")
    ws3.append(["Client", "Market", "Orders Removed", "Gross Lost"])
    style_header(ws3)
    for c in cancelled_clients:
        ws3.append([c["name"], c.get("market", ""), c.get("orders_removed", 0), c.get("gross_lost", 0)])
    ws3.column_dimensions["A"].width = 45
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 16
    ws3.column_dimensions["D"].width = 14

    # Expiring
    ws4 = wb.create_sheet("Expiring (60 days)")
    ws4.append(["Client", "Market", "Last Order Date"])
    style_header(ws4)
    red = PatternFill("solid", fgColor="FFC7CE")
    for e in expiring:
        ws4.append([e["name"], e["market"], e["last_order"]])
        if e["last_order"] <= (date.today().isoformat()):
            for cell in ws4[ws4.max_row]:
                cell.fill = red
    ws4.column_dimensions["A"].width = 45
    ws4.column_dimensions["B"].width = 10
    ws4.column_dimensions["C"].width = 16

    wb.save(path)
    return path


# ---------- Email ----------

def build_email_html(new_clients, activated, cancelled_clients, expiring, removed_count, total_orders):
    """cancelled_clients: list of dicts {name, market, orders_removed, gross_lost}"""
    today = date.today().strftime("%B %d, %Y")
    sections = []

    sections.append(f"<h2 style='color:#1A3A5C;margin-bottom:4px;'>Weekly Import Report — {today}</h2>")
    sections.append(
        f"<p style='color:#5C6370;'>Total orders in database: <b>{total_orders:,}</b> | "
        f"Orders cancelled this week: <b>{removed_count}</b> across <b>{len(cancelled_clients)}</b> clients</p>"
    )

    # New clients
    if new_clients:
        rows = "".join(
            f"<tr><td style='padding:4px 10px;'>{c.get('name','')}</td>"
            f"<td style='padding:4px 10px;'>{c.get('market','')}</td>"
            f"<td style='padding:4px 10px;'>{c.get('category','')}</td></tr>"
            for c in new_clients
        )
        sections.append(f"""
        <h3 style='color:#1A3A5C;'>New Clients Added ({len(new_clients)})</h3>
        <table border='1' cellspacing='0' style='border-collapse:collapse;border-color:#D0D5DD;font-size:14px;'>
        <tr style='background:#1A3A5C;color:white;'><th style='padding:6px 10px;text-align:left;'>Client</th><th style='padding:6px 10px;text-align:left;'>Market</th><th style='padding:6px 10px;text-align:left;'>Category</th></tr>
        {rows}</table>""")
    else:
        sections.append("<h3 style='color:#1A3A5C;'>New Clients Added: 0</h3>")

    # Cancelled this week — clients who had orders removed during this import
    if cancelled_clients:
        # Sort by gross_lost desc so biggest losses appear first
        sorted_c = sorted(cancelled_clients, key=lambda x: -x.get("gross_lost", 0))
        rows = "".join(
            f"<tr><td style='padding:4px 10px;'>{c['name']}</td>"
            f"<td style='padding:4px 10px;'>{c.get('market', '')}</td>"
            f"<td style='padding:4px 10px;text-align:right;'>{c.get('orders_removed', 0)}</td>"
            f"<td style='padding:4px 10px;text-align:right;'>${c.get('gross_lost', 0):,.0f}</td></tr>"
            for c in sorted_c[:50]
        )
        more = f"<p><i>...and {len(sorted_c)-50} more (see attached report)</i></p>" if len(sorted_c) > 50 else ""
        sections.append(f"""
        <h3 style='color:#D95D39;'>Cancelled This Week — Orders Removed ({len(cancelled_clients)} clients, {removed_count} orders)</h3>
        <p style='color:#5C6370;font-size:13px;margin:0 0 6px 0;'>Clients who had one or more orders dropped from the Waterfall spreadsheet this week.</p>
        <table border='1' cellspacing='0' style='border-collapse:collapse;border-color:#D0D5DD;font-size:14px;'>
        <tr style='background:#D95D39;color:white;'><th style='padding:6px 10px;text-align:left;'>Client</th><th style='padding:6px 10px;text-align:left;'>Market</th><th style='padding:6px 10px;text-align:right;'>Orders Removed</th><th style='padding:6px 10px;text-align:right;'>Gross Lost</th></tr>
        {rows}</table>{more}""")

    # Expiring
    if expiring:
        rows = "".join(
            f"<tr><td style='padding:4px 10px;'>{e['name']}</td>"
            f"<td style='padding:4px 10px;'>{e['market']}</td>"
            f"<td style='padding:4px 10px;'>{e['last_order']}</td></tr>"
            for e in expiring[:30]
        )
        more = f"<p><i>...and {len(expiring)-30} more (see attached report)</i></p>" if len(expiring) > 30 else ""
        sections.append(f"""
        <h3 style='color:#1A3A5C;'>Orders Expiring Within 60 Days ({len(expiring)} clients)</h3>
        <table border='1' cellspacing='0' style='border-collapse:collapse;border-color:#D0D5DD;font-size:14px;'>
        <tr style='background:#1A3A5C;color:white;'><th style='padding:6px 10px;text-align:left;'>Client</th><th style='padding:6px 10px;text-align:left;'>Market</th><th style='padding:6px 10px;text-align:left;'>Last Order</th></tr>
        {rows}</table>{more}""")

    sections.append("<br><p style='color:#5C6370;font-size:12px;'>Full change log attached. Generated by THMedia Data Hub.</p>")

    return "\n".join(sections)


def send_email(subject, html_body, attachment_path=None, recipients=None):
    if not GMAIL_USER or not GMAIL_APP_PASSWORD:
        print("  WARNING: Gmail credentials not set. Skipping email.")
        return False

    recipients = recipients or NOTIFY_EMAILS
    if not recipients:
        print("  WARNING: No recipients configured. Skipping email.")
        return False

    msg = MIMEMultipart()
    msg["From"] = f"THMedia Data Hub <{GMAIL_USER}>"
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = subject

    msg.attach(MIMEText(html_body, "html"))

    if attachment_path and Path(attachment_path).exists():
        with open(attachment_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename={Path(attachment_path).name}")
            msg.attach(part)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
            server.sendmail(GMAIL_USER, recipients, msg.as_string())
        print(f"  Email sent to: {', '.join(recipients)}")
        return True
    except Exception as e:
        print(f"  ERROR sending email: {e}")
        return False


# ---------- Main ----------

def generate_report(new_clients=None, removed_count=0, removed_by_client=None, send=True):
    """Generate the full post-import report.
    removed_by_client: dict {client_id: {orders_removed, gross_lost, market}}
      from import_orders.py — the actual orders dropped from the spreadsheet this week.
    """
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Status sync (still runs for DB hygiene, but no longer drives the email section)
    transitions = sync_client_statuses(sb)
    activated = [c for (old, new), rows in transitions.items() if new == "active" and old != "active" for c in rows]

    # Resolve removed_by_client → list of dicts with client names/markets
    cancelled_clients = []
    if removed_by_client:
        client_ids = list(removed_by_client.keys())
        client_rows = []
        for i in range(0, len(client_ids), 100):
            chunk = client_ids[i:i+100]
            res = sb.table("clients").select("id,name,primary_market_id").in_("id", chunk).execute().data
            client_rows.extend(res)
        market_codes = {m["id"]: m["code"] for m in sb.table("markets").select("id,code").execute().data}
        client_map = {c["id"]: c for c in client_rows}
        for cid, info in removed_by_client.items():
            c = client_map.get(cid)
            if not c:
                continue
            cancelled_clients.append({
                "name": c["name"],
                "market": info.get("market") or market_codes.get(c.get("primary_market_id"), ""),
                "orders_removed": info.get("orders_removed", 0),
                "gross_lost": info.get("gross_lost", 0),
            })

    # Expiring orders
    print("Finding clients with expiring orders...")
    expiring = find_expiring_clients(sb, days=60)
    print(f"  {len(expiring)} clients expiring within 60 days")

    # Total orders
    total_result = sb.table("orders").select("id", count="exact").execute()
    total_orders = total_result.count if hasattr(total_result, 'count') and total_result.count else 0
    if not total_orders:
        total_orders = len(fetch_all(sb, "orders", "id"))

    new_clients = new_clients or []

    # Write change log
    log_path = OUTPUT_DIR / f"[C] Import Change Log {date.today().isoformat()}.xlsx"
    write_change_log(new_clients, activated, cancelled_clients, expiring, removed_count, total_orders, log_path)
    print(f"  Change log saved: {log_path.name}")

    # Email
    if send:
        html = build_email_html(new_clients, activated, cancelled_clients, expiring, removed_count, total_orders)
        subject = f"THMedia Data Hub — Weekly Import Report ({date.today().strftime('%b %d, %Y')})"
        send_email(subject, html, attachment_path=log_path)

    return log_path


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--no-email", action="store_true")
    parser.add_argument("--test-email", action="store_true", help="Send a short test email and exit")
    args = parser.parse_args()

    if args.test_email:
        print("Sending test email...")
        ok = send_email(
            subject="THMedia Data Hub — Test Email",
            html_body="<h2>Test successful!</h2><p>Your Data Hub email notifications are working.</p>",
        )
        sys.exit(0 if ok else 1)

    generate_report(send=not args.no_email)


if __name__ == "__main__":
    main()