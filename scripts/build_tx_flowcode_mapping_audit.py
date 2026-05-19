"""Build a TX Flowcode -> Supabase client mapping audit xlsx.

Pulls TX Suites from Flowcode (Austin + San Antonio folders) and tries to auto-
match each Suite to a Supabase client. Candidate scope is restricted to the TX
tenant (`mm_global_id LIKE '764-%'`) -- CO and UT clients are never offered as
matches, even on a fuzzy name hit.

Match cascade:
  1. MM ID prefix parsed from Suite name (e.g. "764-12345 | Client Name AU")
  2. Fuzzy substring match against TX clients only, ranked by:
       has_orders -> status priority (active > cancelled > expired > dormant > prospect)
       -> order_ct desc

The Override columns are PRE-FILLED with the top-ranked TX candidate when one
exists with orders > 0. You confirm by leaving as-is, change by typing a
different MM ID, or clear by deleting both cells (which we treat as "archive
this Suite, do not map").

Output: output/[C] TX Flowcode Mapping Audit {date}.xlsx
"""
from __future__ import annotations

import re
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from config import supabase as sb
from etl.flowcode_client import FlowcodeClient

# Only the TX tenant produces candidate matches. CO=2400, UT=1169 are excluded.
TX_TENANT_PREFIX = "764-"

# Lower is better. Drives candidate ranking inside fuzzy_match().
STATUS_PRIORITY = {
    "active": 0,
    "cancelled": 1,
    "expired": 2,
    "dormant": 3,
    "prospect": 4,
    None: 5,
    "": 5,
}

# Styles
HEADER_FILL = PatternFill("solid", fgColor="2D2A4A")
HEADER_FONT = Font(name="Plus Jakarta Sans", color="FFFFFF", bold=True, size=11)
BODY_FONT = Font(name="Plus Jakarta Sans", size=10)
OK_FILL = PatternFill("solid", fgColor="E6F4EA")
REVIEW_FILL = PatternFill("solid", fgColor="FCE7C9")
ERR_FILL = PatternFill("solid", fgColor="FCE7E7")
THIN = Side(border_style="thin", color="C9C5DC")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, headers: list[str], widths: list[int]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = BOX
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def write_cell(c, value, fill=None):
    c.value = value
    c.font = BODY_FONT
    c.border = BOX
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    if fill:
        c.fill = fill


# Regex: "764-12345 | Some Client Name AU" or "764-12345 - Something"
MM_ID_PREFIX_RE = re.compile(r"^(\d{3,4}-\d+)\s*[|\-]\s*(.+?)$")
TRAILING_MARKET_RE = re.compile(r"\s+(AU|SA|TX|CO|UT)\s*$", re.IGNORECASE)

# Words to strip from a parsed brand name before fuzzy match. These are
# placement/state hints (Ad, POPOUT, Flyer), revision markers, or date stamps
# left over from designers naming their drafts.
NOISE_TOKENS = {
    "ad", "ads", "popout", "popouts", "po", "bc", "fc", "fb", "flyer", "flyers",
    "old", "new", "1st", "2nd", "3rd", "try", "revision", "rev", "draft",
    "dup", "duplicate", "test", "copy", "v1", "v2", "v3",
}
# Stop tokens always dropped before token-set comparison (corporate noise).
STOP_TOKENS = {
    "the", "llc", "inc", "co", "company", "&", "and", "of", "a", "an",
    "for", "to", "by", "ltd", "corp", "corporation",
    # Generic pronouns/short words that aren't brand-distinctive
    "my", "your", "our", "his", "her", "this", "that",
}
# Geographic tokens — they appear in many TX clients (e.g. every "San Antonio
# X") and would inflate scores between unrelated brands. We keep them for
# tiebreaking but require at least one non-geo shared token to confirm a match.
GEO_TOKENS = {
    "texas", "tx", "austin", "san", "antonio", "houston", "dallas",
    "fort", "ft", "worth", "thm", "aus",  # 'aus' = austin abbrev fallback
}
# Abbreviation expansions applied to the parsed brand before tokenising.
# "TX Pools & Patios" -> "Texas Pools & Patios" so it matches the real client.
ABBREVIATION_EXPANSIONS = {
    "tx": "texas",
    "sa": "san antonio",
    "au": "austin",
}
# Strip leftover date stamps in Suite names: "12-2021", "4_23", "2021-12"
DATE_STAMP_RE = re.compile(r"\b\d{1,2}[-_/]\d{2,4}\b|\b\d{4}[-_/]\d{1,2}\b")
TRAILING_DASH_NOISE_RE = re.compile(r"\s*-\s*(old|new|1st try|2nd try|revision|rev|dup|duplicate|test|copy|v\d)\s*$", re.IGNORECASE)


def parse_mm_id_from_name(name: str) -> tuple[str | None, str, str | None]:
    """Returns (mm_global_id, cleaned_brand_name, market_hint).

    market_hint is the trailing 'AU' / 'SA' / 'TX' / 'CO' / 'UT' suffix on the
    Suite name. Designers consistently tag market on Suites (e.g. "Houk AU",
    "Diamond Decks SA"), so this is a high-signal disambiguator when several
    TX clients share a brand token across markets.
    """
    if not name:
        return None, "", None
    s = name.strip()
    market_m = TRAILING_MARKET_RE.search(s)
    market = market_m.group(1).upper() if market_m else None
    m = MM_ID_PREFIX_RE.match(s)
    if m:
        mm_id = m.group(1)
        brand = TRAILING_MARKET_RE.sub("", m.group(2).strip()).strip()
        return mm_id, brand, market
    brand = TRAILING_MARKET_RE.sub("", s).strip()
    return None, brand, market


def brand_tokens(raw: str) -> set[str]:
    """Lower, expand abbreviations, strip dates/noise/stop words, return token set.

    Used for token-set similarity scoring against client names. The goal is to
    let "Total Concrete Ad" match "Total Concrete Solutions" via shared
    {"total", "concrete"} regardless of the trailing placement-hint word.
    """
    if not raw:
        return set()
    s = raw.lower()
    s = TRAILING_DASH_NOISE_RE.sub("", s)
    s = DATE_STAMP_RE.sub(" ", s)
    # Token split on any non-alphanumeric
    toks = re.split(r"[^a-z0-9]+", s)
    out: set[str] = set()
    for t in toks:
        if not t:
            continue
        if len(t) == 1:  # standalone single letter/digit (e.g. "J.E." → "j","e")
            continue
        if t in NOISE_TOKENS or t in STOP_TOKENS:
            continue
        if t in ABBREVIATION_EXPANSIONS:
            for et in ABBREVIATION_EXPANSIONS[t].split():
                out.add(et)
            continue
        # Stem trailing "s" on tokens longer than 4 chars so
        # "carlsons" matches "carlson's" / "carlson" after splitting on the
        # apostrophe. Length guard avoids destroying short words like "doors".
        if len(t) > 4 and t.endswith("s") and not t.endswith("ss"):
            out.add(t[:-1])
        else:
            out.add(t)
    return out


def fetch_tx_clients() -> list[dict]:
    """Return all clients in the TX tenant (mm_global_id LIKE '764-%').

    Includes prospect-status rows because some TX brands haven't booked yet,
    but excludes mapping-stub ghosts.
    """
    out: list[dict] = []
    offset = 0
    while True:
        batch = (
            sb.table("clients")
            .select("id, name, mm_global_id, status, is_mapping_stub")
            .like("mm_global_id", f"{TX_TENANT_PREFIX}%")
            .eq("is_mapping_stub", False)
            .range(offset, offset + 999)
            .execute()
            .data
        )
        out.extend(batch)
        if len(batch) < 1000:
            break
        offset += 1000
    return out


def fetch_order_counts(client_ids: list[str]) -> dict[str, int]:
    """Bulk-fetch order counts for a candidate set. Returns {client_id: count}.

    PostgREST doesn't expose GROUP BY, so we pull (client_id) rows filtered to
    the candidate set and tally client-side. Chunked to keep URLs short.
    """
    counts: dict[str, int] = {cid: 0 for cid in client_ids}
    if not client_ids:
        return counts
    CHUNK = 200
    for i in range(0, len(client_ids), CHUNK):
        chunk = client_ids[i : i + CHUNK]
        offset = 0
        while True:
            batch = (
                sb.table("orders")
                .select("client_id")
                .in_("client_id", chunk)
                .range(offset, offset + 999)
                .execute()
                .data
            )
            for r in batch:
                cid = r["client_id"]
                if cid in counts:
                    counts[cid] += 1
            if len(batch) < 1000:
                break
            offset += 1000
    return counts


def main() -> None:
    fc = FlowcodeClient()
    print("Listing TX Suites from Flowcode (including drafts + archived) ...")
    suites = fc.list_suites()  # default: paginated, includes ACTIVE+DRAFT+ARCHIVED
    tx_suites = [
        s
        for s in suites
        if any(k in (s.get("folderPath", "") or "").upper() for k in ["AUSTIN", "SAN ANTONIO", "THMTX"])
    ]
    print(f"  {len(tx_suites)} TX Suites found")

    # Pull lifetime scan volume per Suite via GetConversionRateSummary
    print("Pulling lifetime scan totals per Suite ...")
    scan_base = {
        "interval": "INTERVAL_CUSTOM",
        "timezone": "America/Denver",
        "orgId": fc.org_id,
        "workspaceId": fc.workspace_id,
        "period": "PERIOD_DAY",
        "timeRange": {"startTime": "2020-01-01T00:00:00Z", "endTime": "2026-12-31T23:59:59Z"},
    }
    suite_scans: dict[str, int] = {}
    for i, s in enumerate(tx_suites, 1):
        body = {**scan_base, "filter": {"suiteId": s["id"]}}
        try:
            res = fc.post("/abacus.v2.AbacusService/GetConversionRateSummary", body)
            suite_scans[s["id"]] = int(res.get("summary", {}).get("totalScans", 0) or 0)
        except Exception:
            suite_scans[s["id"]] = -1
        if i % 25 == 0:
            print(f"  {i}/{len(tx_suites)} ...", flush=True)

    # Codes per Suite — best effort. Many drafts have no listable batch.
    print("Fetching Codes inside each Suite (best effort) ...")
    suite_codes: dict[str, list[dict]] = {}
    for s in tx_suites:
        suite_id = s["id"]
        try:
            suite_obj = fc.get_suite(suite_id).get("suite", {})
            batch_id = None
            for _ak, asset in (suite_obj.get("assets") or {}).items():
                if asset.get("type") == "ASSET_TYPE_ENTRYPOINT_BATCH":
                    batch_id = asset.get("assetId")
                    break
            if batch_id:
                codes = fc.list_codes(batch_id, page_size=100)
                suite_codes[suite_id] = [{**c, "_batch_id": batch_id} for c in codes]
            else:
                suite_codes[suite_id] = []
        except Exception:
            suite_codes[suite_id] = []
    total_codes = sum(len(v) for v in suite_codes.values())
    print(f"  {total_codes} total Codes across {len(tx_suites)} Suites")

    # Load TX-tenant Supabase clients only (no CO/UT pollution in candidates)
    print("Loading TX Supabase clients (mm_global_id LIKE '764-%') ...")
    tx_clients = fetch_tx_clients()
    print(f"  {len(tx_clients)} TX clients available")

    by_mm: dict[str, dict] = {c["mm_global_id"]: c for c in tx_clients if c.get("mm_global_id")}
    by_id: dict[str, dict] = {c["id"]: c for c in tx_clients}

    # Junction: client_mm_identities can map a TX-tenant mm_global_id to a multi-
    # tenant client whose canonical mm_global_id lives in CO or UT. Pull only the
    # 764- rows so we stay in scope.
    print("Loading client_mm_identities (TX rows) ...")
    jct = (
        sb.table("client_mm_identities")
        .select("client_id, mm_global_id")
        .like("mm_global_id", f"{TX_TENANT_PREFIX}%")
        .execute()
        .data
    )
    extra_client_ids = [r["client_id"] for r in jct if r["client_id"] not in by_id]
    if extra_client_ids:
        # Fetch the parent clients for those junction rows (they may be CO/UT-anchored)
        for i in range(0, len(extra_client_ids), 200):
            chunk = extra_client_ids[i : i + 200]
            for c in (
                sb.table("clients")
                .select("id, name, mm_global_id, status, is_mapping_stub")
                .in_("id", chunk)
                .execute()
                .data
            ):
                if c.get("is_mapping_stub"):
                    continue
                by_id[c["id"]] = c
    for r in jct:
        mm = r.get("mm_global_id")
        if mm and mm not in by_mm and r["client_id"] in by_id:
            by_mm[mm] = by_id[r["client_id"]]
    candidate_pool = list(by_id.values())
    print(f"  matching pool: {len(candidate_pool)} clients (TX + junction-linked)")

    # Order counts (used to rank fuzzy candidates by real activity)
    print("Fetching order counts for the matching pool ...")
    order_counts = fetch_order_counts([c["id"] for c in candidate_pool])
    for c in candidate_pool:
        c["_order_ct"] = order_counts.get(c["id"], 0)
    print(f"  {sum(1 for c in candidate_pool if c['_order_ct'] > 0)} clients with at least 1 order")

    # Pre-compute name -> token set for each client so we can score quickly.
    for c in candidate_pool:
        c["_tokens"] = brand_tokens(c.get("name") or "")

    def score_bucket(score: float) -> int:
        # Loose buckets: 0=near-exact, 1=strong, 2=weak.
        # Cutoff at 0.85 keeps a 0.89 perfect-name+ordered match (e.g.
        # "Showroom Windows & Doors") in the same bucket as 1.0, so an
        # active 0.5 weak-match ("Window World Austin") can't outrank it
        # just on status alone.
        if score >= 0.85:
            return 0
        if score >= 0.65:
            return 1
        return 2

    def rank_key(c: dict, score: float, market_aligned: bool) -> tuple:
        """Best candidate first.

        Sort by: score_bucket -> has_orders -> market_aligned -> status ->
        -order_count -> -score.

        market_aligned is true when the client's name ends with the same
        market suffix as the Suite (e.g. probe "TX Pools & Patios SA" ->
        prefer client "Texas Pools & Patios - SA" over "...of Austin").
        """
        bucket = score_bucket(score)
        has_orders = 0 if c.get("_order_ct", 0) > 0 else 1
        market_rank = 0 if market_aligned else 1
        status_rank = STATUS_PRIORITY.get(c.get("status"), 5)
        return (bucket, has_orders, market_rank, status_rank, -c.get("_order_ct", 0), -score)

    def token_score(probe_tokens: set[str], c_tokens: set[str]) -> float:
        """Token-set overlap weighted toward probe coverage, GEO-gated.

        Gate: at least one shared non-geo token is required. Without this,
        "TX Pools & Patios" would match every TX client with "texas" in its
        name, and "My SA Handyman" would match every "...San Antonio" client.

        Score: probe-coverage weighted 2:1 over client-coverage, computed on
        the non-geo intersection so geographic tokens don't dilute or inflate.
        """
        if not probe_tokens or not c_tokens:
            return 0.0
        non_geo_probe = probe_tokens - GEO_TOKENS
        non_geo_cli = c_tokens - GEO_TOKENS
        non_geo_inter = non_geo_probe & non_geo_cli
        if not non_geo_inter:
            return 0.0
        # Use non-geo coverage for the headline score; falls back to probe
        # tokens (full set) when the probe has no non-geo tokens of its own.
        probe_basis = non_geo_probe or probe_tokens
        cli_basis = non_geo_cli or c_tokens
        probe_cov = len(non_geo_inter) / len(probe_basis)
        cli_cov = len(non_geo_inter) / len(cli_basis)
        return (probe_cov * 2 + cli_cov) / 3

    # Suffix variants we accept as "this client is the AU / SA flavour".
    # The right tail of the client name is the dominant signal — "Bath Tune-Up
    # Round Rock AU" definitively belongs to Austin even though many SA clients
    # also have "AU" appear somewhere internally.
    MARKET_SUFFIX_HINTS = {
        "AU": (" AU", " AUSTIN", " - AUSTIN", " OF AUSTIN"),
        "SA": (" SA", " SAN ANTONIO", " - SAN ANTONIO", " OF SAN ANTONIO"),
    }

    def is_market_aligned(client_name: str, market_hint: str | None) -> bool:
        if not market_hint:
            return False
        n = (client_name or "").upper().rstrip(" .,-")
        for suffix in MARKET_SUFFIX_HINTS.get(market_hint, ()):
            if n.endswith(suffix):
                return True
        return False

    def fuzzy_match(brand: str, market_hint: str | None = None) -> list[dict]:
        """Top 5 TX candidates ranked by similarity, activity, market alignment."""
        if not brand:
            return []
        probe_tokens = brand_tokens(brand)
        if not probe_tokens:
            return []

        # 0.5 threshold: at least half the probe's distinctive non-geo tokens
        # are in the client name. Below that we don't surface as a candidate.
        THRESHOLD = 0.5
        scored: list[tuple[tuple, dict, float]] = []
        for c in candidate_pool:
            s = token_score(probe_tokens, c["_tokens"])
            if s >= THRESHOLD:
                aligned = is_market_aligned(c.get("name", ""), market_hint)
                scored.append((rank_key(c, s, aligned), c, s))

        scored.sort(key=lambda x: x[0])
        return [c for _, c, _ in scored[:5]]

    # Match each Suite
    suite_rows = []
    for s in tx_suites:
        name = s.get("name") or ""
        mm_id, brand, market_hint = parse_mm_id_from_name(name)
        match_method = "—"
        matched_client = None
        candidates: list[dict] = []

        if mm_id and mm_id in by_mm:
            matched_client = by_mm[mm_id]
            match_method = "MM ID prefix"
        else:
            cands = fuzzy_match(brand, market_hint=market_hint)
            if not cands:
                match_method = "no match"
            else:
                matched_client = cands[0]
                candidates = cands[1:]  # alternates only
                if len(cands) == 1:
                    match_method = "fuzzy (only TX candidate)"
                elif matched_client["_order_ct"] > 0 and all(
                    c["_order_ct"] == 0 for c in cands[1:]
                ):
                    match_method = "fuzzy (only one with orders)"
                else:
                    match_method = f"fuzzy ({len(cands)} candidates — review)"

        suite_rows.append(
            {
                "folder": s.get("folderPath") or "(root)",
                "name": name,
                "suite_id": s["id"],
                "state": (s.get("state") or "").replace("ASSET_STATE_", ""),
                "scans": suite_scans.get(s["id"], 0),
                "parsed_mm_id": mm_id or "",
                "parsed_brand": brand,
                "match_method": match_method,
                "matched_client_name": matched_client["name"] if matched_client else "",
                "matched_mm_id": matched_client["mm_global_id"] if matched_client else "",
                "matched_status": matched_client.get("status") if matched_client else "",
                "matched_orders": matched_client.get("_order_ct", 0) if matched_client else 0,
                "candidates": candidates,
                "num_codes": len(suite_codes.get(s["id"], [])),
            }
        )

    # Sort: rows that need user input float to the top, then scans desc.
    # "MM ID prefix" matches are most trustworthy and sink to the bottom.
    def sort_key(r):
        m = r["match_method"]
        if m == "MM ID prefix":
            method_rank = 4
        elif m == "fuzzy (only one with orders)" or m == "fuzzy (only TX candidate)":
            method_rank = 3
        elif "candidates" in m:
            method_rank = 1
        elif m == "no match":
            method_rank = 0
        else:
            method_rank = 2
        s = r["scans"] if isinstance(r["scans"], int) and r["scans"] >= 0 else 0
        return (method_rank, -s)
    suite_rows.sort(key=sort_key)

    # Build code rows (one per Code, carrying suite-level mapping)
    code_rows = []
    suite_lookup = {r["suite_id"]: r for r in suite_rows}
    for s in tx_suites:
        suite_match = suite_lookup[s["id"]]
        codes = suite_codes.get(s["id"], [])
        for c in codes:
            code_rows.append(
                {
                    "folder": s.get("folderPath") or "(root)",
                    "suite_name": s.get("name"),
                    "suite_id": s["id"],
                    "code_id": c.get("id"),
                    "code_name": c.get("name"),
                    "short_url": c.get("shortUrl"),
                    "destination_url": (c.get("metadata") or {}).get("url"),
                    "matched_client_name": suite_match["matched_client_name"],
                    "matched_mm_id": suite_match["matched_mm_id"],
                    "match_method": suite_match["match_method"],
                }
            )

    # Stats
    mm_id_matches = [r for r in suite_rows if r["match_method"] == "MM ID prefix"]
    confident = [r for r in suite_rows if r["match_method"] in ("fuzzy (only TX candidate)", "fuzzy (only one with orders)")]
    review = [r for r in suite_rows if "candidates" in r["match_method"]]
    unmatched = [r for r in suite_rows if r["match_method"] == "no match"]
    prefilled = [r for r in suite_rows if r["matched_mm_id"]]

    # ---------------- Write workbook ----------------
    today = date.today()
    date_str = f"{today.month}-{today.day}-{today.year}"
    out_path = Path(f"output/[C] TX Flowcode Mapping Audit {date_str}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    style_header(ws, ["Metric", "Count"], [50, 14])
    rows = [
        ("Total TX Suites", len(tx_suites)),
        ("  -> MM ID prefix match (highest confidence)", len(mm_id_matches)),
        ("  -> fuzzy match with clear winner (orders or sole TX hit)", len(confident)),
        ("  -> fuzzy match with multiple candidates (review)", len(review)),
        ("  -> no TX match found (likely archive)", len(unmatched)),
        ("Override columns pre-filled (delete to skip, edit to change)", len(prefilled)),
        ("Total Codes inside TX Suites", total_codes),
        ("", ""),
        ("Workflow", ""),
        ("  1. Sheet 'Suite Mappings' is sorted: no-match rows at top, MM-ID prefix at bottom", ""),
        ("  2. Override cols A/B are PRE-FILLED with best TX guess (where one exists)", ""),
        ("  3. Confirm row -> leave as-is. Reject row -> clear both cells. Change -> retype.", ""),
        ("  4. Dropdown on col A offers other TX candidates from this row's fuzzy hits.", ""),
    ]
    for i, (metric, count) in enumerate(rows, start=2):
        write_cell(ws.cell(row=i, column=1), metric)
        write_cell(ws.cell(row=i, column=2), count)

    # Suite Mappings sheet
    ws = wb.create_sheet("Suite Mappings")
    headers = [
        "Override Client Name",        # A — pre-filled, dropdown of alternates
        "Override MM Global ID",       # B — pre-filled
        "Suggested # Orders",          # C — quick credibility check
        "Suggested Status",            # D
        "Folder", "Flowcode Suite Name", "State", "Lifetime Scans",
        "Match Method",
        "Other TX Candidates (name | mm_id | status | orders)",
        "Parsed MM ID", "Parsed Brand",
        "Suite ID", "# Codes",
    ]
    widths = [34, 16, 9, 12, 30, 50, 10, 12, 36, 70, 14, 28, 38, 8]
    style_header(ws, headers, widths)

    for r_idx, r in enumerate(suite_rows, start=2):
        # Background fill on the Match Method cell so the eye sees urgency.
        if r["match_method"] == "no match":
            fill = ERR_FILL
        elif "candidates" in r["match_method"]:
            fill = REVIEW_FILL
        else:
            fill = OK_FILL

        cands_str = (
            "\n".join(
                f"{c['name']} | {c.get('mm_global_id','-')} | {c.get('status','-')} | {c.get('_order_ct', 0)} ord"
                for c in r["candidates"]
            )
            if r.get("candidates")
            else ""
        )
        scans_display = r["scans"] if isinstance(r["scans"], int) and r["scans"] >= 0 else "ERR"

        # PRE-FILL override columns when we have a TX suggestion.
        # User can: leave as-is (accept), clear both (reject/archive), or retype (change).
        prefill_name = r["matched_client_name"]
        prefill_mm = r["matched_mm_id"]

        vals = [
            prefill_name, prefill_mm,
            r["matched_orders"] if prefill_name else "",
            r["matched_status"] or "" if prefill_name else "",
            r["folder"], r["name"], r["state"], scans_display,
            r["match_method"],
            cands_str,
            r["parsed_mm_id"], r["parsed_brand"],
            r["suite_id"], r["num_codes"],
        ]
        for col, v in enumerate(vals, start=1):
            cell_fill = None
            if col == 9:  # Match Method col gets row-status color
                cell_fill = fill
            if col in (1, 2) and prefill_name:
                # Light green tint on pre-filled override cells so user sees "accept by default"
                cell_fill = OK_FILL
            write_cell(ws.cell(row=r_idx, column=col), v, fill=cell_fill)

        # Per-row dropdown of alternates on the Override Client Name cell.
        # Excel data-validation list formula uses commas; we strip commas from
        # candidate names and quote-wrap to keep the formula valid.
        if r["candidates"]:
            opts = []
            if prefill_name:
                opts.append(prefill_name.replace(",", " "))
            for c in r["candidates"]:
                opts.append(c["name"].replace(",", " "))
            formula = '"' + ",".join(opts)[:250] + '"'  # 255-char cap on inline lists
            dv = DataValidation(
                type="list", formula1=formula, allow_blank=True,
                showErrorMessage=False,  # don't block manual typing
            )
            ws.add_data_validation(dv)
            dv.add(f"A{r_idx}")

    # Code Mappings sheet
    ws = wb.create_sheet("Code Mappings")
    headers = [
        "Override Client Name",
        "Folder", "Suite Name", "Code Name", "Code ID", "Short URL",
        "Destination URL", "Matched Client", "Matched MM ID", "Match Method",
    ]
    widths = [30, 30, 40, 50, 38, 32, 50, 30, 14, 28]
    style_header(ws, headers, widths)
    for r_idx, r in enumerate(code_rows, start=2):
        vals = ["", r["folder"], r["suite_name"], r["code_name"], r["code_id"],
                r["short_url"], r["destination_url"], r["matched_client_name"],
                r["matched_mm_id"], r["match_method"]]
        for col, v in enumerate(vals, start=1):
            write_cell(ws.cell(row=r_idx, column=col), v)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print()
    print(f"MM ID prefix matches:           {len(mm_id_matches)}")
    print(f"Confident fuzzy (auto-accept):  {len(confident)}")
    print(f"Multi-candidate (review):       {len(review)}")
    print(f"No TX match (likely archive):   {len(unmatched)}")
    print(f"Override columns pre-filled:    {len(prefilled)} / {len(tx_suites)}")
    print(f"Wrote -> {out_path}")


if __name__ == "__main__":
    main()
