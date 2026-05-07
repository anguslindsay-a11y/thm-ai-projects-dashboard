"""Lightweight helper for THM Data Hub analytics reports.

Usage from any script in the repo:

    from scripts.analyze import query, to_xlsx

    rows = query("SELECT name, status FROM clients WHERE NOT is_mapping_stub LIMIT 5")

    # single-tab xlsx
    to_xlsx("Client Sample", rows)

    # multi-tab xlsx with a custom summary
    to_xlsx(
        "Austin April Retention",
        sheets={"Summary": summary_rows, "All Clients": detail_rows},
    )

Writes land in `Supabase Data Hub/output/[C] {name} {M-D-YYYY}.xlsx` by default.
All queries run read-only by default — set `allow_writes=True` to override.
"""
from __future__ import annotations

import os
from datetime import date
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent.parent
OUTPUT_DIR = REPO / "output"

load_dotenv(REPO / ".env")

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _connect():
    """Open a psycopg connection using DATABASE_URL from .env."""
    try:
        import psycopg  # psycopg3
    except ImportError as e:
        raise RuntimeError(
            "psycopg not installed. Run: pip install \"psycopg[binary]\""
        ) from e

    url = os.environ.get("DATABASE_URL")
    if not url:
        # Fall back to assembling from discrete parameter fields.
        # Supports both the Supabase "parameters" block (host/port/dbname/user/password)
        # and the uppercase SUPABASE_DB_* variants.
        host = os.environ.get("host") or os.environ.get("SUPABASE_DB_HOST")
        port = os.environ.get("port") or os.environ.get("SUPABASE_DB_PORT", "5432")
        dbname = os.environ.get("dbname") or os.environ.get("SUPABASE_DB_NAME", "postgres")
        user = os.environ.get("user") or os.environ.get("SUPABASE_DB_USER")
        pw = os.environ.get("password") or os.environ.get("SUPABASE_DB_PASSWORD")
        if host and user and pw:
            import urllib.parse as _u
            url = (
                f"postgresql://{_u.quote(user, safe='')}"
                f":{_u.quote(pw, safe='')}@{host}:{port}/{dbname}"
            )
    if not url:
        raise RuntimeError(
            "No DB connection found. Add DATABASE_URL to .env, or the Supabase "
            "parameter block (host/port/dbname/user/password). Get it from "
            "Supabase dashboard → Settings → Database → Connection string."
        )
    return psycopg.connect(url)


def query(sql: str, params: Sequence[Any] | Mapping[str, Any] | None = None,
          *, allow_writes: bool = False) -> list[dict]:
    """Run SQL and return rows as list of dicts.

    By default, rejects any statement other than SELECT/WITH. Set
    ``allow_writes=True`` to permit DML/DDL (use the VS Code chat for those).
    """
    stripped = sql.lstrip().lower()
    if not allow_writes and not (stripped.startswith("select") or stripped.startswith("with")):
        raise RuntimeError(
            "analyze.query() is read-only by default. Use VS Code chat for writes, "
            "or pass allow_writes=True if you're sure."
        )

    with _connect() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            if cur.description is None:
                return []
            cols = [d[0] for d in cur.description]
            return [dict(zip(cols, row)) for row in cur.fetchall()]


def _default_filename(name: str) -> Path:
    today = date.today()
    stamp = f"{today.month}-{today.day}-{today.year}"
    return OUTPUT_DIR / f"[C] {name} {stamp}.xlsx"


def _infer_number_format(values: Iterable[Any]) -> str | None:
    """Guess a number format from a column's values (currency for $ hints, etc.)."""
    # Keep it simple — let the caller pass explicit formats when needed.
    return None


def _write_sheet(ws, rows: list[dict], *, money_cols: set[str] | None = None) -> None:
    if not rows:
        ws.cell(row=1, column=1, value="(no rows)").font = Font(italic=True, color="808080")
        return
    headers = list(rows[0].keys())
    money_cols = money_cols or set()
    # Auto-detect obvious money columns
    for h in headers:
        lower = h.lower()
        if any(tok in lower for tok in ("gross", "net", "amount", "spend", "revenue", "cost")):
            money_cols.add(h)

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for r, row in enumerate(rows, start=2):
        for c, h in enumerate(headers, 1):
            val = row.get(h)
            cell = ws.cell(row=r, column=c, value=val)
            if h in money_cols and isinstance(val, (int, float)):
                cell.number_format = "$#,##0.00"

    ws.freeze_panes = "A2"
    try:
        ws.auto_filter.ref = ws.dimensions
    except Exception:
        pass
    # Auto-width
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(width + 2, 60)


def to_xlsx(
    name: str,
    rows: list[dict] | None = None,
    *,
    sheets: dict[str, list[dict]] | None = None,
    path: str | Path | None = None,
    money_cols: set[str] | None = None,
) -> Path:
    """Write rows to an xlsx under output/.

    Pass either ``rows`` (for a single-tab sheet named 'Data') or ``sheets``
    (a dict of tab-name -> rows). Returns the written path.
    """
    if rows is None and not sheets:
        raise ValueError("Pass either rows=... or sheets={...}")
    if rows is not None and sheets:
        raise ValueError("Pass rows OR sheets, not both")

    out_path = Path(path) if path else _default_filename(name)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    if rows is not None:
        ws = wb.active
        ws.title = "Data"
        _write_sheet(ws, rows, money_cols=money_cols)
    else:
        # Multi-tab
        first = True
        for tab_name, tab_rows in sheets.items():
            if first:
                ws = wb.active
                ws.title = tab_name[:31]
                first = False
            else:
                ws = wb.create_sheet(tab_name[:31])
            _write_sheet(ws, tab_rows, money_cols=money_cols)

    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    # Smoke test: if run directly, verify the connection and print the zone list.
    try:
        rows = query("SELECT abbreviation, name, state FROM zones ORDER BY state, name")
        print(f"Connected. {len(rows)} zones:")
        for r in rows:
            print(f"  {r['abbreviation']:<5} {r['name']:<20} ({r['state']})")
    except Exception as e:
        print(f"Connection check failed: {e}")
        raise SystemExit(1)
