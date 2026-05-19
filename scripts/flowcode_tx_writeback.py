"""Write reviewed TX Flowcode -> client mappings into Supabase.

Reads the user-edited audit xlsx (Suite Mappings sheet) and creates one
`client_platform_ids` row per Suite that has an Override MM Global ID filled
in. Blank-override rows are skipped (they're the "no TX match -- archive"
pile).

external_id convention for TX Suite-level rows: `FC-S-{suite_uuid}`. The
`S-` prefix distinguishes Suite-level from the existing 446 CO+UT Code-level
rows (`FC-{code_uuid}`). The ETL uses this prefix to decide whether to query
Flowcode at Suite or Code granularity.

Idempotent: re-runnable, upserts on (client_id, platform, external_id).
Default is --dry-run; pass --commit to write.
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook

from config import supabase as sb

LEDGER_PATH = Path("output/flowcode_tx_writeback_ledger.jsonl")

# Column indexes on the Suite Mappings sheet (0-based, matches build script)
COL_OVERRIDE_NAME = 0
COL_OVERRIDE_MM = 1
COL_FOLDER = 4
COL_SUITE_NAME = 5
COL_STATE = 6
COL_SUITE_ID = 12


def load_clients_by_mm() -> dict[str, str]:
    """Map mm_global_id -> client_id, including junction-linked identities."""
    out: dict[str, str] = {}
    offset = 0
    while True:
        batch = (
            sb.table("clients")
            .select("id, mm_global_id")
            .not_.is_("mm_global_id", "null")
            .range(offset, offset + 999)
            .execute()
            .data
        )
        for r in batch:
            out[r["mm_global_id"]] = r["id"]
        if len(batch) < 1000:
            break
        offset += 1000

    # Multi-tenant brands: an mm_global_id may live in the junction table
    # while the canonical client row carries a different one.
    jct = sb.table("client_mm_identities").select("client_id, mm_global_id").execute().data
    for r in jct:
        mm = r.get("mm_global_id")
        if mm and mm not in out:
            out[mm] = r["client_id"]
    return out


def read_audit_rows(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["Suite Mappings"]
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        override_name = (row[COL_OVERRIDE_NAME] or "").strip() if row[COL_OVERRIDE_NAME] else ""
        override_mm = (row[COL_OVERRIDE_MM] or "").strip() if row[COL_OVERRIDE_MM] else ""
        rows.append({
            "override_name": override_name,
            "override_mm": override_mm,
            "folder": row[COL_FOLDER] or "",
            "suite_name": row[COL_SUITE_NAME] or "",
            "state": row[COL_STATE] or "",
            "suite_id": row[COL_SUITE_ID] or "",
        })
    return rows


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="output/[C] TX Flowcode Mapping Audit 5-19-2026.xlsx")
    ap.add_argument("--commit", action="store_true", help="Actually write to Supabase")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: audit xlsx not found at {xlsx_path}")
        return 1

    print(f"Reading {xlsx_path} ...")
    audit_rows = read_audit_rows(xlsx_path)
    print(f"  {len(audit_rows)} total rows")

    print("Loading client mm_global_id -> client_id map ...")
    by_mm = load_clients_by_mm()
    print(f"  {len(by_mm)} mm_global_ids known")

    # Classify
    to_map: list[dict] = []
    skipped_blank: list[dict] = []
    skipped_unknown_mm: list[dict] = []
    for r in audit_rows:
        if not r["override_mm"]:
            skipped_blank.append(r)
            continue
        client_id = by_mm.get(r["override_mm"])
        if not client_id:
            skipped_unknown_mm.append(r)
            continue
        r["client_id"] = client_id
        to_map.append(r)

    print()
    print(f"To map:               {len(to_map)} Suites")
    print(f"Skipped (blank):      {len(skipped_blank)}")
    print(f"Skipped (unknown MM): {len(skipped_unknown_mm)}")
    if skipped_unknown_mm:
        print("  Unknown MM IDs (override entered but not found in clients):")
        for r in skipped_unknown_mm[:10]:
            print(f"    {r['override_mm']:14s}  {r['suite_name']!r}")

    if not args.commit:
        print()
        print("DRY RUN -- no rows written. Pass --commit to execute.")
        print()
        print("Preview of first 8 upserts:")
        for r in to_map[:8]:
            ext_id = f"FC-S-{r['suite_id']}"
            print(f"  client_id={r['client_id'][:8]}...  external_id={ext_id}  name={r['override_name']!r}")
        return 0

    # Commit
    LEDGER_PATH.parent.mkdir(parents=True, exist_ok=True)
    ledger = LEDGER_PATH.open("a")
    upserted = 0
    failed: list[dict] = []
    for r in to_map:
        ext_id = f"FC-S-{r['suite_id']}"
        notes_obj = {
            "suite_id": r["suite_id"],
            "flow_name": r["suite_name"],
            "folder": r["folder"],
            "state": r["state"],
            "kind": "suite",  # marks this as Suite-level (vs Code-level)
            "archived": "Archive" in (r["folder"] or "") or "Not In Use" in (r["folder"] or "")
                        or r["state"] == "ARCHIVED",
        }
        row = {
            "client_id": r["client_id"],
            "platform": "flowcode",
            "external_id": ext_id,
            "external_name": r["override_name"] or r["suite_name"],
            "notes": json.dumps(notes_obj),
        }
        try:
            sb.table("client_platform_ids").upsert(
                row,
                on_conflict="client_id,platform,external_id",
            ).execute()
            upserted += 1
            ledger.write(json.dumps({
                "ts": datetime.now(timezone.utc).isoformat(),
                "event": "tx_cpi_upserted",
                "external_id": ext_id,
                "client_id": r["client_id"],
                "suite_name": r["suite_name"],
                "override_mm": r["override_mm"],
            }) + "\n")
        except Exception as e:
            failed.append({"row": r, "err": str(e)[:200]})
            ledger.write(json.dumps({
                "ts": datetime.now(timezone.utc).isoformat(),
                "event": "tx_cpi_failed",
                "external_id": ext_id,
                "err": str(e)[:200],
            }) + "\n")

    ledger.close()
    print()
    print(f"Upserted: {upserted}")
    print(f"Failed:   {len(failed)}")
    for f in failed[:5]:
        print(f"  {f['err']}  ({f['row']['suite_name']!r})")
    print(f"Ledger -> {LEDGER_PATH}")
    return 0 if not failed else 2


if __name__ == "__main__":
    sys.exit(main())
