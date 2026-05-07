"""TX Market Analysis — recurring drop-off report.

Flags TX (AU + SA) clients whose ads are dropping off in the next 60 days or
have dropped off in the last 330 days. One tab per rep, with product-level
spend per issue, last-order context, and three notes columns.

Notes column behavior:
  - Previous Notes = most recent note found in the prior report's sheet
    (prefers that sheet's "New Notes", falls back to its "Previous Notes").
  - New Notes / Management Notes = blank for the rep to fill in.

Run:
    python -m scripts.tx_market_analysis
"""
from __future__ import annotations

from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from scripts.analyze import query, OUTPUT_DIR

# --- Config --------------------------------------------------------------
TODAY = date.today()
# Directories where prior "TX Market Analysis *.xlsx" files live, in priority order
PRIOR_FILE_SEARCH_DIRS = [
    Path(r"C:\Users\MasenSpring\OneDrive - TheHomeMagWest"),
    Path(__file__).resolve().parent.parent / "output",
]


def find_prior_file() -> Path | None:
    """Find the most recent TX Market Analysis file that ISN'T today's output."""
    candidates: list[Path] = []
    today_stamp = f"{TODAY.month}-{TODAY.day}-{TODAY.year}"
    for d in PRIOR_FILE_SEARCH_DIRS:
        if not d.exists():
            continue
        for pattern in ("TX Market Analysis *.xlsx", "[[]C[]] TX Market Analysis *.xlsx"):
            for p in d.glob(pattern):
                # Skip today's own output to avoid self-reference
                if today_stamp in p.stem:
                    continue
                candidates.append(p)
    if not candidates:
        return None
    # Pick the most recently modified file
    return max(candidates, key=lambda p: p.stat().st_mtime)


PRIOR_FILE = find_prior_file()

# Map display-tab-name → sales_rep value(s) in orders.sales_rep
REP_MAP: list[tuple[str, list[str]]] = [
    ("Lora", ["Lora Butler"]),
    ("Brittany", ["Brittany Cox"]),
    ("Zach", ["Zach Sherman"]),
    ("Claire", ["House Accounts"]),
    ("National Accounts", ["hm t"]),   # improvement: surface national book too
]

def _month_iter(start_year: int, start_month: int, count: int):
    """Yield (year, month) pairs starting from (start_year, start_month)."""
    y, m = start_year, start_month
    for _ in range(count):
        yield y, m
        m += 1
        if m > 12:
            m = 1
            y += 1


def build_issue_cols(today: date) -> list[tuple[str, str]]:
    """Auto-build a 5-month issue window: 2 months back + current + 2 forward.
    Inserts the Spring special (April) between Mar and xApr in any year where
    April falls inside the window.

    Matches the density of the source Market Analysis format (~6 columns total
    when Spring is included).
    """
    month_names = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    start_month = today.month - 2
    start_year = today.year
    if start_month < 1:
        start_month += 12
        start_year -= 1
    months_list = []
    for y, m in _month_iter(start_year, start_month, 5):
        yy = str(y)[-2:]
        mm = f"{m:02d}"
        name = month_names[m - 1]
        code = f"{yy}.{mm}.x{name}" if m == 4 else f"{yy}.{mm}.{name}"
        label = f"x{name}" if m == 4 else name
        months_list.append((code, label))
        # Insert Spring right before xApr
        if m == 4:
            months_list.insert(-1, (f"{yy}.04.Spr", "Spr"))
    return months_list


# Issue columns to display (auto-computed from today's date)
ISSUE_COLS = build_issue_cols(TODAY)

# Date bounds for the pivot window — derived from ISSUE_COLS (first & last month)
def _code_to_date(code: str, end_of_month: bool = False) -> str:
    """'26.02.Feb' → '2026-02-01' (or '2026-02-28' if end_of_month)."""
    yy, mm, _ = code.split(".")
    year = 2000 + int(yy)
    month = int(mm)
    if end_of_month:
        # Use 28 for simplicity; all issue_date_parsed are on the 15th anyway
        day = 28
    else:
        day = 1
    return f"{year}-{month:02d}-{day:02d}"


WINDOW_START = _code_to_date(ISSUE_COLS[0][0])
WINDOW_END = _code_to_date(ISSUE_COLS[-1][0], end_of_month=True)

# YoY comparison window = last 3 monthly issues in ISSUE_COLS (current + next 2)
# versus the same 3 months one year prior.
_TY_MONTHLY = [c for c, _ in ISSUE_COLS if ".Spr" not in c][-3:]
_LY_MONTHLY = [c.replace(c[:2], f"{int(c[:2]) - 1:02d}", 1) for c in _TY_MONTHLY]
YOY_TY_CODES = tuple(_TY_MONTHLY)
YOY_LY_CODES = tuple(_LY_MONTHLY)
YOY_WINDOW_TY_START = _code_to_date(YOY_TY_CODES[0])
YOY_WINDOW_TY_END = _code_to_date(YOY_TY_CODES[-1], end_of_month=True)
YOY_WINDOW_LY_START = _code_to_date(YOY_LY_CODES[0])
YOY_WINDOW_LY_END = _code_to_date(YOY_LY_CODES[-1], end_of_month=True)

# --- Styling -------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
TOTAL_FONT = Font(bold=True)
DROPPED_FILL = PatternFill("solid", fgColor="F8CBAD")    # already dropped
UPCOMING_FILL = PatternFill("solid", fgColor="FFF2CC")   # drops off in next 60d
MONEY_FMT = "$#,##0"
THIN = Side(style="thin", color="BFBFBF")
THICK = Side(style="medium", color="404040")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_SECTION_BOTTOM = Border(left=THIN, right=THIN, top=THIN, bottom=THICK)

# --- 1. Load prior notes ------------------------------------------------
def load_prior_notes() -> dict[tuple[str, str], str]:
    """Key: (rep_tab_lower, client_name_lower) → note text to carry forward."""
    if PRIOR_FILE is None or not PRIOR_FILE.exists():
        print("  (no prior TX Market Analysis file found; Previous Notes will be empty)")
        return {}
    print(f"  Using prior file: {PRIOR_FILE.name}")

    wb = load_workbook(PRIOR_FILE, data_only=True)
    notes: dict[tuple[str, str], str] = {}
    for sn in wb.sheetnames:
        if sn.lower() == "year to year":
            continue
        ws = wb[sn]
        headers = [c.value for c in ws[1]]
        try:
            i_client = headers.index("Client")
            i_prod = headers.index("Product")
            i_prev = headers.index("Previous Notes")
            i_new = headers.index("New Notes")
        except ValueError:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            client = row[i_client]
            prod = row[i_prod]
            if not client or prod != "Total":
                continue
            new_note = row[i_new]
            prev_note = row[i_prev]
            note = (new_note or prev_note or "")
            note = str(note).strip() if note else ""
            if note:
                notes[(sn.lower(), str(client).strip().lower())] = note
    print(f"  Loaded {len(notes)} notes from prior report")
    return notes


# --- 2. Query drop-off clients + pivot orders ----------------------------
DROP_OFF_SQL = """
WITH tx_client_max AS (
  SELECT
    o.client_id,
    o.sales_rep,
    MAX(o.issue_date_parsed) as last_order_date,
    MIN(o.issue_date_parsed) FILTER (WHERE o.issue_date_parsed > CURRENT_DATE + INTERVAL '60 days') as next_after_window
  FROM orders o
  JOIN markets m ON o.market_id = m.id
  WHERE m.code IN ('AU','SA')
  GROUP BY o.client_id, o.sales_rep
),
dropoff AS (
  SELECT client_id, sales_rep, last_order_date
  FROM tx_client_max
  WHERE last_order_date BETWEEN CURRENT_DATE - INTERVAL '330 days'
                            AND CURRENT_DATE + INTERVAL '60 days'
    AND next_after_window IS NULL
)
SELECT
  d.sales_rep,
  c.name as client_name,
  d.last_order_date,
  o.issue_date as issue_code,
  z.abbreviation as zone_abbr,
  o.size as ad_size,
  o.net
FROM dropoff d
JOIN clients c ON c.id = d.client_id
JOIN orders o ON o.client_id = d.client_id AND o.sales_rep = d.sales_rep
LEFT JOIN zones z ON o.zone_id = z.id
JOIN markets m ON o.market_id = m.id
WHERE m.code IN ('AU','SA')
  AND o.issue_date_parsed BETWEEN '{WINDOW_START}' AND '{WINDOW_END}'
  AND c.name NOT LIKE 'THM Branch Page%'  -- internal filler, $0 ads
ORDER BY d.sales_rep, c.name, o.issue_date, z.abbreviation, o.size
""".format(WINDOW_START=WINDOW_START, WINDOW_END=WINDOW_END)


def build_rep_data(rows, sales_reps: list[str]) -> dict:
    """Reshape rows for one rep into: client → {product: {issue: net}, meta}."""
    matching = [r for r in rows if r["sales_rep"] in sales_reps]
    clients: dict[str, dict] = {}
    for r in matching:
        cname = r["client_name"]
        c = clients.setdefault(cname, {
            "last_order_date": r["last_order_date"],
            "products": defaultdict(lambda: defaultdict(float)),  # product → issue → $
        })
        zone = r["zone_abbr"] or "—"
        product = f"{zone} {r['ad_size']}" if r["ad_size"] else zone
        c["products"][product][r["issue_code"]] += float(r["net"] or 0)
        c["last_order_date"] = max(c["last_order_date"], r["last_order_date"])
    return clients


def status_label(last_order: date) -> tuple[str, PatternFill | None]:
    delta = (last_order - TODAY).days
    if delta < 0:
        return (f"Dropped {-delta}d ago ({last_order:%b %Y})", DROPPED_FILL)
    elif delta <= 60:
        return (f"Drops off in {delta}d ({last_order:%b %Y})", UPCOMING_FILL)
    return ("", None)


def write_rep_tab(ws, rep_display: str, clients: dict):
    """Write one rep tab. Columns:
       Client | Product | Last Order | Status | [issue cols] | Total | Prev | New | Mgmt
    """
    base_cols = ["Client", "Product", "Last Order", "Status"]
    issue_labels = [lbl for _, lbl in ISSUE_COLS]
    tail_cols = ["Total", "Previous Notes", "New Notes", "Management Notes"]
    headers = base_cols + issue_labels + tail_cols

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    row_i = 2
    money_col_start = len(base_cols) + 1
    total_col = money_col_start + len(ISSUE_COLS)
    total_col_idx = len(headers)  # last column
    # Column indices (1-based) for client-level merged cells
    COL_CLIENT = 1
    COL_LAST_ORDER = 3
    COL_STATUS = 4
    COL_PREV = total_col + 1
    COL_NEW = total_col + 2
    COL_MGMT = total_col + 3

    prior = load_prior_notes_cache
    for client_name in sorted(clients.keys()):
        info = clients[client_name]
        last = info["last_order_date"]
        status_txt, status_fill = status_label(last)
        note = prior.get((rep_display.lower(), client_name.lower()), "")

        # Compute totals + usable products first, so we know how many rows
        issue_totals = {code: 0.0 for code, _ in ISSUE_COLS}
        usable_products = []
        for prod in sorted(info["products"].keys()):
            issues = info["products"][prod]
            row_total = sum(issues.get(code, 0) for code, _ in ISSUE_COLS)
            if row_total == 0:
                continue
            usable_products.append((prod, issues, row_total))
            for code, amt in issues.items():
                if code in issue_totals:
                    issue_totals[code] += amt
        grand_total = sum(issue_totals.values())

        section_start = row_i
        section_end = row_i + len(usable_products)  # Total row + N product rows

        # ---- Total row ----
        # NOTE: No peach/orange fill. Only bold font + border. Orange/yellow
        # highlighting is reserved for the Status column per user preference.
        values = [client_name, "Total", last, status_txt]
        values += [issue_totals[code] or None for code, _ in ISSUE_COLS]
        values += [grand_total, note, None, None]
        for col_i, v in enumerate(values, 1):
            cell = ws.cell(row=row_i, column=col_i, value=v)
            cell.font = TOTAL_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col_i == COL_PREV))
            if col_i >= money_col_start and col_i <= total_col:
                cell.number_format = MONEY_FMT
            if col_i == COL_LAST_ORDER and isinstance(v, date):
                cell.number_format = "mmm d, yyyy"
        # Status column is the ONLY place orange/yellow fill is applied
        if status_fill:
            ws.cell(row=row_i, column=COL_STATUS).fill = status_fill
        row_i += 1

        # ---- Product rows ----
        for prod, issues, row_total in usable_products:
            values = [None, prod, None, None]
            values += [issues.get(code) or None for code, _ in ISSUE_COLS]
            values += [row_total or None, None, None, None]
            for col_i, v in enumerate(values, 1):
                cell = ws.cell(row=row_i, column=col_i, value=v)
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center")
                if col_i >= money_col_start and col_i <= total_col:
                    cell.number_format = MONEY_FMT
            # Extend the Status fill down through the merged section
            if status_fill:
                ws.cell(row=row_i, column=COL_STATUS).fill = status_fill
            row_i += 1

        # ---- Merge client-level columns across the section ----
        if section_end > section_start:
            for col in (COL_CLIENT, COL_LAST_ORDER, COL_STATUS, COL_PREV, COL_NEW, COL_MGMT):
                ws.merge_cells(start_row=section_start, end_row=section_end,
                               start_column=col, end_column=col)
            # Set alignment on the merged top-left cells
            ws.cell(row=section_start, column=COL_CLIENT).alignment = Alignment(
                vertical="center", wrap_text=True)
            ws.cell(row=section_start, column=COL_LAST_ORDER).alignment = Alignment(
                vertical="center", horizontal="center")
            ws.cell(row=section_start, column=COL_STATUS).alignment = Alignment(
                vertical="center", horizontal="center", wrap_text=True)
            ws.cell(row=section_start, column=COL_PREV).alignment = Alignment(
                vertical="center", wrap_text=True)
            ws.cell(row=section_start, column=COL_NEW).alignment = Alignment(
                vertical="center", wrap_text=True)
            ws.cell(row=section_start, column=COL_MGMT).alignment = Alignment(
                vertical="center", wrap_text=True)

        # ---- Thick bottom border on the last row of the section ----
        for col_i in range(1, total_col_idx + 1):
            cell = ws.cell(row=section_end, column=col_i)
            existing = cell.border
            cell.border = Border(
                left=existing.left, right=existing.right,
                top=existing.top, bottom=THICK,
            )

    # --- Formatting ---
    ws.freeze_panes = "E2"
    widths = [38, 30, 13, 28] + [11] * len(ISSUE_COLS) + [12, 50, 30, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30


# --- 3. Year-to-Year comparison ------------------------------------------
# Pulls BOTH 2025 and 2026 Apr-Jun for ALL TX clients so we can find clients who
# ran last year but have zero spend in the same window this year (true lost revenue).
YOY_SQL = """
SELECT
  c.name as client_name,
  o.issue_date as issue_code,
  SUM(o.net) as net
FROM orders o
JOIN clients c ON o.client_id = c.id
JOIN markets m ON o.market_id = m.id
WHERE m.code IN ('AU','SA')
  AND (
    o.issue_date_parsed BETWEEN '{LY_START}' AND '{LY_END}'
    OR o.issue_date_parsed BETWEEN '{TY_START}' AND '{TY_END}'
  )
  AND c.name NOT LIKE 'THM Branch Page%'
GROUP BY c.name, o.issue_date
""".format(
    LY_START=YOY_WINDOW_LY_START, LY_END=YOY_WINDOW_LY_END,
    TY_START=YOY_WINDOW_TY_START, TY_END=YOY_WINDOW_TY_END,
)


def build_yoy_sheet(ws, yoy_rows):
    """Show clients who ran in Apr-Jun 2025 but have ZERO Apr-Jun 2026 spend.

    This is the 'lost revenue' list — the reps should know what they used to have
    that they don't have anymore in the same window.
    """
    # Regular monthly issues only — excludes Spring special editions (inflates
    # 'lost revenue' because many clients only run Spring, not monthly).
    def _fmt(code: str) -> str:  # "25.04.xApr" → "25.04 xApr"
        parts = code.split(".")
        return f"{parts[0]}.{parts[1]} {parts[2]}"
    headers = ["Client"]
    headers += [_fmt(c) for c in YOY_LY_CODES] + [f"{YOY_LY_CODES[0][:2]} Total"]
    headers += [_fmt(c) for c in YOY_TY_CODES] + [f"{YOY_TY_CODES[0][:2]} Total"]
    headers += ["YoY Δ"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    # Regular monthly issue codes to consider (skip Spring special) — computed
    # dynamically at module load based on today's date.
    LY_CODES = YOY_LY_CODES
    TY_CODES = YOY_TY_CODES

    last_yr = defaultdict(lambda: defaultdict(float))
    this_yr = defaultdict(lambda: defaultdict(float))
    for r in yoy_rows:
        code = r["issue_code"]
        net = float(r["net"] or 0)
        if code in LY_CODES:
            last_yr[r["client_name"]][code] += net
        elif code in TY_CODES:
            this_yr[r["client_name"]][code] += net

    # Only include clients with 2025 monthly spend AND zero 2026 monthly spend
    lost_clients = []
    for cname, ly in last_yr.items():
        ly_tot = sum(ly.values())
        ty_tot = sum(this_yr.get(cname, {}).values())
        if ly_tot > 0 and ty_tot == 0:
            lost_clients.append((cname, ly, ly_tot))

    lost_clients.sort(key=lambda x: -x[2])

    row_i = 2
    grand_by_code = {c: 0.0 for c in LY_CODES}
    delta_col = len(headers)  # last column
    for cname, ly, ly_tot in lost_clients:
        monthly_vals = [ly.get(c, 0) for c in LY_CODES]
        for c, v in zip(LY_CODES, monthly_vals):
            grand_by_code[c] += v
        vals = [cname]
        vals += [v or None for v in monthly_vals] + [ly_tot]
        vals += [None] * len(TY_CODES) + [None]
        vals += [-ly_tot]
        for col_i, v in enumerate(vals, 1):
            cell = ws.cell(row=row_i, column=col_i, value=v)
            cell.border = BORDER
            if col_i > 1:
                cell.number_format = MONEY_FMT
            if col_i == delta_col and isinstance(v, (int, float)) and v < 0:
                cell.font = Font(color="C00000", bold=True)
        row_i += 1

    grand_tot = sum(grand_by_code.values())
    ly_label = LY_CODES[0][:2]  # "25"
    totals = [f"TOTAL (lost revenue YoY, monthly issues only)"]
    totals += [grand_by_code[c] or None for c in LY_CODES] + [grand_tot or None]
    totals += [None] * len(TY_CODES) + [None]
    totals += [-grand_tot or None]
    # YoY grand total row — bold + thick border, NO peach fill (per style rule:
    # orange highlighting reserved for Status column on rep tabs only)
    for col_i, v in enumerate(totals, 1):
        cell = ws.cell(row=row_i, column=col_i, value=v)
        cell.font = Font(bold=True)
        cell.border = BORDER_SECTION_BOTTOM
        if col_i > 1:
            cell.number_format = MONEY_FMT
        if col_i == delta_col and isinstance(v, (int, float)) and v < 0:
            cell.font = Font(color="C00000", bold=True)

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 45
    for c in "BCDEFGHIJ":
        ws.column_dimensions[c].width = 13


# --- 4. Summary tab (new) ------------------------------------------------
def write_summary_tab(ws, all_clients_by_rep):
    headers = ["Rep", "Drop-Off Clients", "At Risk $ (Feb-Jun)",
               "Already Dropped", "Drops Off <=60d"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    row_i = 2
    for rep, clients in all_clients_by_rep.items():
        dropped = sum(1 for c in clients.values() if c["last_order_date"] < TODAY)
        upcoming = sum(1 for c in clients.values() if TODAY <= c["last_order_date"])
        total_at_risk = 0.0
        for info in clients.values():
            for issues in info["products"].values():
                total_at_risk += sum(v for k, v in issues.items()
                                     if k in {c for c, _ in ISSUE_COLS})
        vals = [rep, len(clients), total_at_risk, dropped, upcoming]
        for col_i, v in enumerate(vals, 1):
            cell = ws.cell(row=row_i, column=col_i, value=v)
            if col_i == 3:
                cell.number_format = MONEY_FMT
        row_i += 1
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18


# --- 5. Main --------------------------------------------------------------
load_prior_notes_cache: dict[tuple[str, str], str] = {}


def main():
    global load_prior_notes_cache
    print("Loading prior report notes...")
    load_prior_notes_cache = load_prior_notes()

    print("Querying TX drop-off clients...")
    rows = query(DROP_OFF_SQL)
    print(f"  {len(rows)} order lines")

    print("Querying YoY comparison window...")
    yoy_rows = query(YOY_SQL)
    print(f"  {len(yoy_rows)} prior-year order lines")

    stamp = f"{TODAY.month}-{TODAY.day}-{TODAY.year}"
    out_path = OUTPUT_DIR / f"[C] TX Market Analysis {stamp}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    all_by_rep = {}
    for rep_display, sales_reps in REP_MAP:
        clients = build_rep_data(rows, sales_reps)
        all_by_rep[rep_display] = clients
        ws = wb.create_sheet(rep_display[:31])
        write_rep_tab(ws, rep_display, clients)
        print(f"  {rep_display}: {len(clients)} drop-off clients")

    ws = wb.create_sheet("Summary")
    write_summary_tab(ws, all_by_rep)

    ws = wb.create_sheet("Year to Year")
    build_yoy_sheet(ws, yoy_rows)

    # Put Summary first for manager visibility, then rep tabs, then YoY
    ordered = ["Summary"] + [r for r, _ in REP_MAP] + ["Year to Year"]
    wb._sheets = [wb[n] for n in ordered if n in wb.sheetnames]

    wb.save(out_path)
    print(f"\nWrote: {out_path}")


if __name__ == "__main__":
    main()
