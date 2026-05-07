"""
Apply CallRail mapping fixes from the audit spreadsheet.

Reads the "CallRail Mapping Audit" sheet, and for each row where
"Correct CallRail Account (edit)" is filled in, applies the corresponding
change: attach, reassign, or remove a CallRail platform mapping. Also moves
existing calls so attribution follows the mapping.

Accepted values in the edit column:
  (blank)                    = no change
  {exact CallRail name}      = attach/reassign that CallRail account to this client
  REMOVE                     = detach this client's current CallRail mapping
  SKIP                       = explicitly skip (noop, useful for notes)

Usage:
  python scripts/apply_callrail_mapping_fixes.py                 # dry-run (default)
  python scripts/apply_callrail_mapping_fixes.py --apply          # actually apply
  python scripts/apply_callrail_mapping_fixes.py --file <path>    # custom spreadsheet path
"""

import sys
import os
import argparse
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
load_dotenv()

from supabase import create_client
from openpyxl import load_workbook

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"


def latest_audit_file():
    candidates = sorted(OUTPUT_DIR.glob("callrail_mapping_audit_*.xlsx"), reverse=True)
    return candidates[0] if candidates else None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="Actually apply changes (default: dry-run)")
    parser.add_argument("--file", type=str, help="Path to audit xlsx (default: latest in output/)")
    parser.add_argument("--skip-mm", type=str, default="", help="Comma-separated MM IDs to skip")
    args = parser.parse_args()
    skip_mm_ids = {s.strip() for s in args.skip_mm.split(",") if s.strip()}

    file_path = Path(args.file) if args.file else latest_audit_file()
    if not file_path or not file_path.exists():
        print("No audit file found. Run audit_callrail_mappings.py first.")
        sys.exit(1)

    print(f"Reading: {file_path}")
    wb = load_workbook(file_path, data_only=True)
    ws = wb["CallRail Mapping Audit"]

    # Build CR name -> company_id lookup from the "CallRail Accounts" reference sheet
    ws_ref = wb["CallRail Accounts"]
    cr_name_to_id = {}
    ref_headers = [c.value for c in ws_ref[1]]
    name_col = ref_headers.index("CallRail Account Name")
    id_col = ref_headers.index("Company ID")
    for row in ws_ref.iter_rows(min_row=2, values_only=True):
        name = row[name_col]
        cr_id = row[id_col]
        if name and cr_id:
            cr_name_to_id[name.strip()] = cr_id

    # Main sheet headers
    main_headers = [c.value for c in ws[1]]
    mm_id_col = main_headers.index("MM ID")
    db_name_col = main_headers.index("DB Client Name")
    cr_name_col_idx = main_headers.index("CallRail Account")
    cr_cid_col = main_headers.index("CallRail Company ID")
    correct_col = main_headers.index("Correct CallRail Account (edit)")

    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    actions = []  # (kind, client_id, current_cr_id, target_cr_id, target_name, row_label)
    row_num = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        correct = row[correct_col]
        if not correct:
            continue
        correct = str(correct).strip()
        if correct.upper() in ("", "SKIP"):
            continue

        mm_id = row[mm_id_col]
        db_name = row[db_name_col]
        current_cr_id = row[cr_cid_col]
        label = f"{db_name} ({mm_id})"

        if mm_id in skip_mm_ids:
            print(f"  SKIP (--skip-mm): {label}")
            continue

        # Resolve the client_id by MM external_id
        mm_rows = (
            sb.table("client_platform_ids")
            .select("client_id")
            .eq("platform", "magazine_manager")
            .eq("external_id", mm_id)
            .execute()
            .data
        )
        if not mm_rows:
            print(f"  SKIP (no client for MM {mm_id}): {label}")
            continue
        client_id = mm_rows[0]["client_id"]

        if correct.upper() == "REMOVE":
            if not current_cr_id:
                print(f"  SKIP (nothing to remove): {label}")
                continue
            actions.append(("REMOVE", client_id, current_cr_id, None, None, label))
            continue

        # Attach/reassign — resolve the target CR name to company_id
        target_cr_id = cr_name_to_id.get(correct)
        if not target_cr_id:
            print(f"  SKIP (CR name not found in reference sheet): '{correct}' for {label}")
            continue

        actions.append(("ATTACH", client_id, current_cr_id, target_cr_id, correct, label))

    print(f"\n{len(actions)} changes to apply:")
    for kind, client_id, cur_cr_id, target_cr_id, target_name, label in actions:
        if kind == "REMOVE":
            print(f"  REMOVE  {label}: detach current CR {cur_cr_id}")
        else:
            prefix = "REASSIGN" if cur_cr_id else "ATTACH  "
            print(f"  {prefix} {label}: -> {target_name}")

    if not args.apply:
        print("\nDRY RUN. Re-run with --apply to execute.")
        return

    print("\nApplying...")
    for kind, client_id, cur_cr_id, target_cr_id, target_name, label in actions:
        if kind == "REMOVE":
            # Detach the current CR from this client, and unassign its calls
            sb.table("client_platform_ids").delete().match({
                "client_id": client_id,
                "platform": "callrail",
                "external_id": cur_cr_id,
            }).execute()
            sb.table("calls").update({"client_id": None}).match({
                "client_id": client_id,
                "callrail_company_id": cur_cr_id,
            }).execute()
            print(f"  REMOVED CR {cur_cr_id} from {label}")
            continue

        # ATTACH or REASSIGN
        # 1. Detach target CR from any existing client (if already mapped elsewhere)
        existing = (
            sb.table("client_platform_ids")
            .select("client_id,external_name")
            .eq("platform", "callrail")
            .eq("external_id", target_cr_id)
            .execute()
            .data
        )
        if existing:
            old_owner = existing[0]["client_id"]
            if old_owner != client_id:
                sb.table("client_platform_ids").delete().match({
                    "platform": "callrail",
                    "external_id": target_cr_id,
                    "client_id": old_owner,
                }).execute()
                # Reassign calls from old owner for this CR company to new client
                sb.table("calls").update({"client_id": client_id}).match({
                    "client_id": old_owner,
                    "callrail_company_id": target_cr_id,
                }).execute()
                print(f"  Detached {target_name} from old owner {old_owner}")

        # 2. Detach the client's current CR if they had one and it's different
        if cur_cr_id and cur_cr_id != target_cr_id:
            sb.table("client_platform_ids").delete().match({
                "client_id": client_id,
                "platform": "callrail",
                "external_id": cur_cr_id,
            }).execute()
            sb.table("calls").update({"client_id": None}).match({
                "client_id": client_id,
                "callrail_company_id": cur_cr_id,
            }).execute()

        # 3. Attach the target CR to this client (upsert to be idempotent)
        sb.table("client_platform_ids").upsert({
            "client_id": client_id,
            "platform": "callrail",
            "external_id": target_cr_id,
            "external_name": target_name,
        }, on_conflict="client_id,platform,external_id").execute()

        # 4. Pull over any unassigned calls for this CR company
        sb.table("calls").update({"client_id": client_id}).match({
            "callrail_company_id": target_cr_id,
        }).is_("client_id", "null").execute()

        print(f"  ATTACHED {target_name} -> {label}")

    print("\nDone.")


if __name__ == "__main__":
    main()
