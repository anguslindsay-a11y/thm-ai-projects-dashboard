"""Phase 2 smoke test — verify SharePoint auth + file download work end-to-end.

Connects to the IA SharePoint site, lists the IA folder, downloads the
spreadsheet to memory, and prints sheet names + row counts. Writes nothing
to disk and nothing to Supabase. If row counts match what you see manually
in SharePoint, Phase 2 is done.

Reads from .env:
- AZURE_TENANT_ID, AZURE_CLIENT_ID, AZURE_CLIENT_SECRET
- SHAREPOINT_IA_SITE_URL  — e.g. https://thehomemagwest.sharepoint.com/sites/SystemsOps
- SHAREPOINT_IA_FILE_PATH — path inside the default document library,
                            e.g. Reports/InBox Advantage/IA Data 2024.12.5 CO,UT,TX.xlsx
"""

import os
import sys
from io import BytesIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
load_dotenv()

from etl.sharepoint_client import (
    SharePointClient,
    SharePointAuthError,
    SharePointAPIError,
)
from openpyxl import load_workbook


def _fail(stage: str, err: Exception) -> None:
    print(f"\nFAILED at: {stage}")
    print(f"  {type(err).__name__}: {err}")
    sys.exit(1)


def main() -> None:
    site_url = os.getenv("SHAREPOINT_IA_SITE_URL")
    file_path = os.getenv("SHAREPOINT_IA_FILE_PATH")
    if not site_url or not file_path:
        print(
            "ERROR: SHAREPOINT_IA_SITE_URL and SHAREPOINT_IA_FILE_PATH must "
            "be set in .env."
        )
        sys.exit(1)

    print("=" * 70)
    print("SharePoint Phase 2 smoke test")
    print("=" * 70)
    print(f"Site URL:  {site_url}")
    print(f"File path: {file_path}")
    print()

    # 1) Build client (validates env vars)
    try:
        sp = SharePointClient()
    except SharePointAuthError as e:
        _fail("client init", e)

    # 2) Acquire token
    print("[1/4] Acquiring access token...")
    try:
        sp.get_token()
        print("      OK")
    except SharePointAuthError as e:
        _fail("token acquisition", e)

    # 3) Resolve site
    print("[2/4] Resolving site ID...")
    try:
        site_id = sp.resolve_site_id(site_url)
        print(f"      OK  ({site_id})")
    except (SharePointAPIError, SharePointAuthError) as e:
        _fail("site resolution", e)

    # 4) List the parent folder so we can eyeball the file is visible
    parts = file_path.strip("/").split("/")
    folder = "/".join(parts[:-1])
    filename = parts[-1]

    if folder:
        print(f"[3/4] Listing folder '{folder}'...")
        try:
            items = sp.list_folder(site_id, folder, top=200)
            print(f"      Found {len(items)} item(s). Showing first 15:")
            for it in items[:15]:
                kind = "DIR " if "folder" in it else "FILE"
                print(f"        {kind}  {it.get('name')}")
            if len(items) > 15:
                print(f"        ... +{len(items) - 15} more")
            visible = any(it.get("name") == filename for it in items)
            if not visible:
                print(
                    f"      WARNING: '{filename}' was not in the listing — "
                    "name may be misspelled or live in a different folder. "
                    "Will still attempt direct download."
                )
        except (SharePointAPIError, SharePointAuthError) as e:
            print(f"      WARN: folder listing failed ({e}). Continuing.")
    else:
        print("[3/4] (no folder — file is at doc-library root, skipping list)")

    # 5) Download to memory and inspect
    print(f"[4/4] Downloading '{filename}'...")
    try:
        data = sp.download_file_bytes(site_id, file_path)
    except (SharePointAPIError, SharePointAuthError) as e:
        _fail("file download", e)
    print(f"      OK  ({len(data):,} bytes)")

    print("\nOpening workbook in memory...")
    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        _fail("openpyxl load_workbook", e)

    print(f"  Sheet names ({len(wb.sheetnames)}): {wb.sheetnames}")
    for name in wb.sheetnames:
        ws = wb[name]
        # In read-only mode, max_row can be unreliable for some xlsx files,
        # so iterate and count actual non-empty rows.
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            if any(c is not None and c != "" for c in row):
                row_count += 1
        print(f"    '{name}': {row_count:,} non-empty rows")

    print("\n" + "=" * 70)
    print("SUCCESS — Phase 2 connection verified.")
    print("=" * 70)


if __name__ == "__main__":
    main()
