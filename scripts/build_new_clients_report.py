"""Build the [C] New Clients Last 12 Months xlsx from the 3 market JSON files."""
from __future__ import annotations

import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(r"C:\Users\MasenSpring\OneDrive - TheHomeMagWest\Supabase Data Hub")
DATA = Path(r"C:\Users\MasenSpring\OneDrive - TheHomeMagWest\.claude\worktrees\peaceful-saha-83c48e\data")
OUT = REPO / "output" / "[C] New Clients Last 12 Months 4-22-2026.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
SUB_FONT = Font(bold=True)


def load_rows() -> list[dict]:
    rows: list[dict] = []
    for fname in ("new_clients_CO.json", "new_clients_TX.json", "new_clients_UT.json"):
        rows.extend(json.loads((DATA / fname).read_text(encoding="utf-8")))
    # Normalize types
    for r in rows:
        r["issues_run"] = int(r["issues_run"] or 0)
        r["months_retained"] = float(r["months_retained"] or 0)
        r["gross_to_date"] = float(r["gross_to_date"] or 0)
    return rows


def style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def auto_width(ws, max_w: int = 55) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(width + 2, max_w)


def build(rows: list[dict]) -> Workbook:
    wb = Workbook()

    # === Summary ===
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "New Clients — Last 12 Months"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated 2026-04-22 · {len(rows)} clients with first order on/after 2025-04-22 · TX = AU+SA combined · stubs excluded"
    ws["A2"].font = Font(italic=True, color="595959")

    ws["A4"] = "By Market"
    ws["A4"].font = Font(bold=True)
    mkt_h = ["Market", "New Clients", "Still Active", "Cancelled", "Expired", "Other", "Retention %", "Gross-to-date"]
    for i, h in enumerate(mkt_h, 1):
        ws.cell(row=5, column=i, value=h)
    style_header(ws, 5, len(mkt_h))

    r = 6
    tot = {"n": 0, "a": 0, "c": 0, "e": 0, "o": 0, "g": 0.0}
    for mk in ("CO", "TX", "UT"):
        sub = [x for x in rows if x["market"] == mk]
        n = len(sub)
        a = sum(1 for x in sub if x["status"] == "active")
        c = sum(1 for x in sub if x["status"] == "cancelled")
        e = sum(1 for x in sub if x["status"] == "expired")
        o = n - a - c - e
        g = sum(x["gross_to_date"] for x in sub)
        ws.cell(row=r, column=1, value=mk)
        ws.cell(row=r, column=2, value=n)
        ws.cell(row=r, column=3, value=a)
        ws.cell(row=r, column=4, value=c)
        ws.cell(row=r, column=5, value=e)
        ws.cell(row=r, column=6, value=o)
        ws.cell(row=r, column=7, value=(a / n) if n else 0).number_format = "0%"
        ws.cell(row=r, column=8, value=g).number_format = "$#,##0.00"
        tot["n"] += n; tot["a"] += a; tot["c"] += c; tot["e"] += e; tot["o"] += o; tot["g"] += g
        r += 1
    ws.cell(row=r, column=1, value="Total")
    ws.cell(row=r, column=2, value=tot["n"])
    ws.cell(row=r, column=3, value=tot["a"])
    ws.cell(row=r, column=4, value=tot["c"])
    ws.cell(row=r, column=5, value=tot["e"])
    ws.cell(row=r, column=6, value=tot["o"])
    ws.cell(row=r, column=7, value=(tot["a"] / tot["n"]) if tot["n"] else 0).number_format = "0%"
    ws.cell(row=r, column=8, value=tot["g"]).number_format = "$#,##0.00"
    for c in range(1, len(mkt_h) + 1):
        ws.cell(row=r, column=c).fill = SUB_FILL
        ws.cell(row=r, column=c).font = SUB_FONT

    # By rep rollup
    r += 3
    ws.cell(row=r, column=1, value="By Opening Rep").font = Font(bold=True)
    r += 1
    rep_h = ["Opening Rep", "Market", "New Clients", "Still Active", "Churned", "Retention %", "Gross-to-date"]
    for i, h in enumerate(rep_h, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, len(rep_h))
    r += 1
    rep_agg: dict[tuple[str, str], dict] = {}
    for x in rows:
        k = (x["opening_rep"], x["market"])
        a = rep_agg.setdefault(k, {"n": 0, "active": 0, "churned": 0, "gross": 0.0})
        a["n"] += 1
        if x["status"] == "active":
            a["active"] += 1
        elif x["status"] in ("cancelled", "expired", "dormant", "inactive"):
            a["churned"] += 1
        a["gross"] += x["gross_to_date"]
    for (rep, mk), a in sorted(rep_agg.items(), key=lambda kv: -kv[1]["n"]):
        ws.cell(row=r, column=1, value=rep)
        ws.cell(row=r, column=2, value=mk)
        ws.cell(row=r, column=3, value=a["n"])
        ws.cell(row=r, column=4, value=a["active"])
        ws.cell(row=r, column=5, value=a["churned"])
        ws.cell(row=r, column=6, value=(a["active"] / a["n"]) if a["n"] else 0).number_format = "0%"
        ws.cell(row=r, column=7, value=a["gross"]).number_format = "$#,##0.00"
        r += 1
    auto_width(ws)

    # === All Clients ===
    ws2 = wb.create_sheet("All Clients")
    dh = ["Client", "Market", "Category", "Status", "Opening Rep",
          "First Issue", "Last Issue Run", "Last Issue Booked",
          "Issues Run", "Months Retained", "Gross to Date"]
    for i, h in enumerate(dh, 1):
        ws2.cell(row=1, column=i, value=h)
    style_header(ws2, 1, len(dh))
    rows_sorted = sorted(rows, key=lambda x: (x["first_issue"], x["gross_to_date"]), reverse=True)
    for i, x in enumerate(rows_sorted, start=2):
        ws2.cell(row=i, column=1, value=x["client"])
        ws2.cell(row=i, column=2, value=x["market"])
        ws2.cell(row=i, column=3, value=x["category"])
        ws2.cell(row=i, column=4, value=x["status"])
        ws2.cell(row=i, column=5, value=x["opening_rep"])
        ws2.cell(row=i, column=6, value=x["first_issue"])
        ws2.cell(row=i, column=7, value=x["last_issue_run"])
        ws2.cell(row=i, column=8, value=x["last_issue_booked"])
        ws2.cell(row=i, column=9, value=x["issues_run"])
        ws2.cell(row=i, column=10, value=x["months_retained"])
        ws2.cell(row=i, column=11, value=x["gross_to_date"]).number_format = "$#,##0.00"
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions
    auto_width(ws2)

    # === By Rep ===
    ws3 = wb.create_sheet("By Rep")
    ws3.cell(row=1, column=1, value="Clients grouped by opening rep (sorted by new-client count desc, newest first)").font = Font(italic=True, color="595959")
    rh = ["Client", "Market", "Category", "Status", "First Issue", "Last Issue Run", "Issues Run", "Months Retained", "Gross to Date"]
    for i, h in enumerate(rh, 1):
        ws3.cell(row=3, column=i, value=h)
    style_header(ws3, 3, len(rh))
    r = 4
    by_rep: dict[str, list[dict]] = {}
    for x in rows:
        by_rep.setdefault(x["opening_rep"], []).append(x)
    for rep in sorted(by_rep, key=lambda k: -len(by_rep[k])):
        group = by_rep[rep]
        active_n = sum(1 for x in group if x["status"] == "active")
        churned_n = sum(1 for x in group if x["status"] in ("cancelled", "expired", "dormant", "inactive"))
        gross = sum(x["gross_to_date"] for x in group)
        ws3.cell(row=r, column=1, value=f"{rep} — {len(group)} new · {active_n} active · {churned_n} churned · ${gross:,.0f} gross").font = SUB_FONT
        for c in range(1, len(rh) + 1):
            ws3.cell(row=r, column=c).fill = SUB_FILL
        r += 1
        for x in sorted(group, key=lambda v: v["first_issue"], reverse=True):
            ws3.cell(row=r, column=1, value=x["client"])
            ws3.cell(row=r, column=2, value=x["market"])
            ws3.cell(row=r, column=3, value=x["category"])
            ws3.cell(row=r, column=4, value=x["status"])
            ws3.cell(row=r, column=5, value=x["first_issue"])
            ws3.cell(row=r, column=6, value=x["last_issue_run"])
            ws3.cell(row=r, column=7, value=x["issues_run"])
            ws3.cell(row=r, column=8, value=x["months_retained"])
            ws3.cell(row=r, column=9, value=x["gross_to_date"]).number_format = "$#,##0.00"
            r += 1
        r += 1
    ws3.freeze_panes = "A4"
    auto_width(ws3)

    return wb


if __name__ == "__main__":
    rows = load_rows()
    print(f"Loaded {len(rows)} rows")
    wb = build(rows)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
