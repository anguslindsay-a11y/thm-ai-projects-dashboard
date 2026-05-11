"""Sample MM activities to understand what we're working with.

Pulls a larger sample (~3000 rows) across all 3 tenants, then categorizes
and writes summary findings to data/mm_api_probes/activity_analysis.md
plus a stratified xlsx sample with representative rows for human review.

Read-only. No DB writes.
"""

from __future__ import annotations

import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from statistics import median

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from etl.magmanager_client import MagManagerClient

load_dotenv()

DATABASES = [
    "thehomemagcolorado",
    "thehomemagutah",
    "thehomemagsanantonio",
]

OUT_DIR = Path(__file__).resolve().parent.parent / "data" / "mm_api_probes"
OUT_DIR.mkdir(parents=True, exist_ok=True)


def classify_activity(a: dict) -> str:
    """Best-guess category for an activity based on flags + content."""
    notes = (a.get("Notes") or "").strip()
    is_call = a.get("IsCall")
    is_email = a.get("IsEmail")
    is_letter = a.get("IsLetter")
    is_mass = a.get("IsMassEmail")
    is_sys = a.get("IsSystem")
    activity_type = a.get("ActivityType")
    has_meeting = a.get("Meeting") is not None
    has_callback = a.get("CallBack") is not None

    if not notes and not has_meeting and not has_callback:
        return "empty"
    if is_sys:
        return "system_autogen"
    if is_mass:
        return "mass_email"
    if is_email and "<br>" in notes:
        return "email_thread_chrome"
    if is_email:
        return "email_other"
    if is_call or (activity_type and "call" in activity_type.lower()):
        return "call_note"
    if has_meeting or (activity_type and "meet" in activity_type.lower()):
        return "meeting_note"
    if is_letter:
        return "letter"
    if has_callback:
        return "callback_reminder"
    if activity_type:
        return f"typed_{activity_type[:30]}"
    return "freeform_note"


def strip_html(s: str) -> str:
    """Quick & dirty HTML strip for length measurement."""
    if not s:
        return ""
    s = re.sub(r"<[^>]+>", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def main():
    print("=" * 70)
    print("MM ACTIVITY ANALYSIS — sampling for ETL design")
    print("=" * 70)

    mm = MagManagerClient()

    # Pull last 30 days, all 3 tenants, paginated to exhaustion (calibrate volume)
    thirty_days_ago = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    all_rows: list[dict] = []
    per_db_pages: dict[str, int] = {}
    for db in DATABASES:
        page = 1
        while True:
            body = mm.get_activities_page(page=page, from_date=thirty_days_ago, database_name=db)
            rows = body.get("Data") or []
            for r in rows:
                r.setdefault("DatabaseName", db)
            all_rows.extend(rows)
            print(f"  {db} page {page}: {len(rows)}")
            if len(rows) < 1000:
                break
            page += 1
            if page > 30:
                print(f"    safety stop")
                break
        per_db_pages[db] = page

    print(f"\nTotal activities pulled (last 30 days): {len(all_rows):,}\n")

    # ----- Stats -----
    cls_counts: Counter[str] = Counter()
    cls_by_db: dict[str, Counter] = defaultdict(Counter)
    note_lengths_by_cls: dict[str, list[int]] = defaultdict(list)
    note_lengths_by_cls_clean: dict[str, list[int]] = defaultdict(list)
    types_seen: Counter[str] = Counter()
    rep_counts: Counter[str] = Counter()
    samples_per_cls: dict[str, list[dict]] = defaultdict(list)
    sys_with_notes: list[dict] = []
    blank_notes_count = 0

    for a in all_rows:
        cls = classify_activity(a)
        cls_counts[cls] += 1
        cls_by_db[a["DatabaseName"]][cls] += 1
        notes = a.get("Notes") or ""
        note_lengths_by_cls[cls].append(len(notes))
        note_lengths_by_cls_clean[cls].append(len(strip_html(notes)))
        if a.get("ActivityType"):
            types_seen[a["ActivityType"]] += 1
        if a.get("Rep"):
            rep_counts[a["Rep"]] += 1
        if not notes.strip():
            blank_notes_count += 1
        # Keep up to 5 samples per class for human review
        if len(samples_per_cls[cls]) < 5:
            samples_per_cls[cls].append({
                "DatabaseName": a.get("DatabaseName"),
                "ActivityID": a.get("ActivityID"),
                "Customer": a.get("Customer"),
                "Rep": a.get("Rep"),
                "ActivityType": a.get("ActivityType"),
                "DateAdded": a.get("DateAdded"),
                "IsCall": a.get("IsCall"), "IsEmail": a.get("IsEmail"),
                "IsLetter": a.get("IsLetter"), "IsMassEmail": a.get("IsMassEmail"),
                "IsSystem": a.get("IsSystem"),
                "Notes": (a.get("Notes") or "")[:500],
                "NotesFullLength": len(a.get("Notes") or ""),
            })
        if a.get("IsSystem") and a.get("Notes"):
            sys_with_notes.append({
                "Notes": (a.get("Notes") or "")[:200],
                "ActivityType": a.get("ActivityType"),
            })

    print("=== CLASSIFICATION DISTRIBUTION ===")
    for cls, n in cls_counts.most_common():
        med_raw = int(median(note_lengths_by_cls[cls])) if note_lengths_by_cls[cls] else 0
        med_clean = int(median(note_lengths_by_cls_clean[cls])) if note_lengths_by_cls_clean[cls] else 0
        max_raw = max(note_lengths_by_cls[cls]) if note_lengths_by_cls[cls] else 0
        print(f"  {cls:<28} {n:>6,}   "
              f"median notes raw={med_raw:>5}  clean={med_clean:>5}  max_raw={max_raw:>6}")

    print(f"\n  Blank-notes rows: {blank_notes_count:,}")
    print(f"  Total unique reps: {len(rep_counts)}")
    print(f"  Top 5 reps by activity volume:")
    for rep, n in rep_counts.most_common(5):
        print(f"    {rep}: {n}")

    print(f"\n=== ACTIVITY TYPES (top 15) ===")
    for t, n in types_seen.most_common(15):
        print(f"  {t}: {n}")

    print(f"\n=== PER-DB DISTRIBUTION ===")
    for db in DATABASES:
        print(f"  {db}: {sum(cls_by_db[db].values()):,}")
        for cls, n in cls_by_db[db].most_common(5):
            print(f"    {cls}: {n}")

    # ----- Write samples xlsx -----
    wb = Workbook()
    s = wb.active
    s.title = "Summary"
    s.column_dimensions["A"].width = 35
    s.column_dimensions["B"].width = 15
    s.column_dimensions["C"].width = 80

    NAVY = "1F3A5F"
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill("solid", fgColor=NAVY)

    s["A1"] = "MM Activity Analysis — last 30 days"
    s["A1"].font = Font(bold=True, size=16, color=NAVY)
    s["A2"] = f"Pulled {datetime.now().strftime('%Y-%m-%d %H:%M')} — {len(all_rows):,} total rows"
    s["A2"].font = Font(italic=True)

    row = 4
    s.cell(row=row, column=1, value="Classification").font = HEADER_FONT
    s.cell(row=row, column=1).fill = HEADER_FILL
    s.cell(row=row, column=2, value="Count").font = HEADER_FONT
    s.cell(row=row, column=2).fill = HEADER_FILL
    s.cell(row=row, column=3, value="Median note length (clean text)").font = HEADER_FONT
    s.cell(row=row, column=3).fill = HEADER_FILL
    row += 1
    for cls, n in cls_counts.most_common():
        s.cell(row=row, column=1, value=cls)
        s.cell(row=row, column=2, value=n)
        s.cell(row=row, column=3, value=int(median(note_lengths_by_cls_clean[cls])) if note_lengths_by_cls_clean[cls] else 0)
        row += 1

    # Per-class sample sheets
    for cls, sample_list in samples_per_cls.items():
        sheet_name = cls[:31]  # Excel limit
        sh = wb.create_sheet(sheet_name)
        cols = ["DatabaseName", "ActivityID", "Customer", "Rep", "ActivityType",
                "DateAdded", "IsCall", "IsEmail", "IsLetter", "IsMassEmail",
                "IsSystem", "NotesFullLength", "Notes"]
        widths = [22, 12, 35, 22, 22, 18, 8, 8, 8, 8, 8, 12, 90]
        for i, (c, w) in enumerate(zip(cols, widths), start=1):
            cell = sh.cell(row=1, column=i, value=c)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            sh.column_dimensions[get_column_letter(i)].width = w
        for r_idx, samp in enumerate(sample_list, start=2):
            for c_idx, key in enumerate(cols, start=1):
                v = samp.get(key)
                cell = sh.cell(row=r_idx, column=c_idx, value=v)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        sh.freeze_panes = "A2"

    out_xlsx = OUT_DIR / "activity_analysis_samples.xlsx"
    wb.save(out_xlsx)
    print(f"\nWrote samples: {out_xlsx}")

    # ----- Markdown findings -----
    md_lines = [
        "# MM Activity Analysis — last 30 days",
        f"_Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}_\n",
        f"## Volume",
        f"- Total activities (last 30 days): **{len(all_rows):,}**",
        f"- Per tenant (pages pulled):",
    ]
    for db, pages in per_db_pages.items():
        md_lines.append(f"  - `{db}`: {sum(cls_by_db[db].values()):,} ({pages} page(s))")
    md_lines.append("")
    md_lines.append("## Classification breakdown")
    md_lines.append("| Category | Count | Median clean notes | Notes |")
    md_lines.append("|---|---:|---:|---|")
    interpretations = {
        "empty": "Useless — no notes, no meeting, no callback",
        "system_autogen": "MM auto-generated (order logs, status events). LOW value individually, useful in aggregate",
        "mass_email": "Marketing blasts. LOW info per row, useful for engagement scoring",
        "email_thread_chrome": "Chrome Extension copy of full email thread. HTML, may be LONG, HIGH value",
        "email_other": "Other email entries. Probably manual rep notes about emails",
        "call_note": "Rep call summary. HIGH value for pre-call intel + coaching",
        "meeting_note": "Sales meeting notes. HIGH value",
        "letter": "Physical letters. Rare",
        "callback_reminder": "Future-dated reminders. Useful for surfacing 'rep promised to call back'",
        "freeform_note": "Untyped note. Variable quality. HIGH value when present",
    }
    for cls, n in cls_counts.most_common():
        med = int(median(note_lengths_by_cls_clean[cls])) if note_lengths_by_cls_clean[cls] else 0
        md_lines.append(
            f"| `{cls}` | {n:,} | {med} | {interpretations.get(cls, '')} |"
        )
    md_lines.append("")
    md_lines.append("## Top reps by activity volume (30 days)")
    md_lines.append("| Rep | Activities |")
    md_lines.append("|---|---:|")
    for rep, n in rep_counts.most_common(15):
        md_lines.append(f"| {rep} | {n:,} |")
    md_lines.append("")
    md_lines.append(f"## ActivityType distribution (top 15)")
    md_lines.append("| Type | Count |")
    md_lines.append("|---|---:|")
    for t, n in types_seen.most_common(15):
        md_lines.append(f"| {t} | {n:,} |")
    md_lines.append("")
    md_lines.append(f"## Stats")
    md_lines.append(f"- Blank-notes rows: **{blank_notes_count:,}** ({blank_notes_count*100//max(1,len(all_rows))}%)")
    md_lines.append(f"- IsSystem=true rows with non-empty notes: **{len(sys_with_notes):,}** "
                     f"(sample below)")
    md_lines.append(f"- Total unique reps: **{len(rep_counts)}**")
    md_lines.append("")
    md_lines.append("### System-autogen note sample (what are these?)")
    md_lines.append("```")
    for s_sample in sys_with_notes[:10]:
        md_lines.append(
            f"  [{s_sample.get('ActivityType') or '(no type)'}] "
            f"{s_sample['Notes'][:150]}"
        )
    md_lines.append("```")

    out_md = OUT_DIR / "activity_analysis.md"
    out_md.write_text("\n".join(md_lines), encoding="utf-8")
    print(f"Wrote findings: {out_md}")


if __name__ == "__main__":
    main()
