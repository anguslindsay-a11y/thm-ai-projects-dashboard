"""
Build an Excel audit of MagManager -> CallRail mappings.

Exports all MagManager accounts in the database with their linked CallRail
account (if any), plus call counts, order activity, and flags for likely
misattribution cases. Sorted by total calls desc so high-volume clients
can be audited first.

Output: output/callrail_mapping_audit_YYYY-MM-DD.xlsx

Usage: python scripts/audit_callrail_mappings.py
"""

import sys
import os
from pathlib import Path
from datetime import date

import requests

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
load_dotenv()

from supabase import create_client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
CALLRAIL_API_KEY = os.getenv("CALLRAIL_API_KEY")

CALLRAIL_BASE_URL = "https://api.callrail.com/v3"
CALLRAIL_ACCOUNTS = [
    ("ACCe42c98d3446c4dc898467150060f870c", "CO"),
    ("ACCb1f04de7a28941f4827eb25f18d5e810", "UT"),
    ("ACC60a4cf8cf0514a45acfde9c07fa1275b", "TX"),
]


def fetch_callrail_companies():
    """Pull all companies from all 3 CallRail accounts. Returns list of dicts."""
    all_companies = []
    headers = {"Authorization": f'Token token="{CALLRAIL_API_KEY}"'}
    for acct_id, acct_code in CALLRAIL_ACCOUNTS:
        page = 1
        while True:
            url = f"{CALLRAIL_BASE_URL}/a/{acct_id}/companies.json"
            resp = requests.get(url, headers=headers, params={"per_page": 250, "page": page})
            resp.raise_for_status()
            data = resp.json()
            for c in data.get("companies", []):
                all_companies.append({
                    "id": c["id"],
                    "name": c["name"],
                    "account": acct_code,
                })
            if page >= data.get("total_pages", 1):
                break
            page += 1
    return all_companies

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)


def fetch_all(sb, table, select, batch=1000, **filters):
    rows, offset = [], 0
    while True:
        q = sb.table(table).select(select).range(offset, offset + batch - 1)
        for k, v in filters.items():
            q = q.eq(k, v)
        data = q.execute().data
        rows.extend(data)
        if len(data) < batch:
            break
        offset += batch
    return rows


def main():
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    print("Loading MagManager platform IDs...")
    mm_rows = fetch_all(sb, "client_platform_ids", "client_id,external_id,external_name", platform="magazine_manager")

    print("Loading CallRail platform IDs...")
    cr_rows = fetch_all(sb, "client_platform_ids", "client_id,external_id,external_name", platform="callrail")
    cr_by_client = {}
    for r in cr_rows:
        cr_by_client.setdefault(r["client_id"], []).append(r)

    print("Loading clients...")
    client_rows = fetch_all(sb, "clients", "id,name,status,primary_market_id")
    clients_by_id = {c["id"]: c for c in client_rows}

    print("Loading markets...")
    market_rows = sb.table("markets").select("id,code").execute().data
    markets_by_id = {m["id"]: m["code"] for m in market_rows}

    PAGE = 1000

    print("Loading call stats (this may take a minute)...")
    call_stats = {}
    offset = 0
    while True:
        batch = sb.table("calls").select("client_id,call_time").range(offset, offset + PAGE - 1).execute().data
        for c in batch:
            cid = c["client_id"]
            if cid not in call_stats:
                call_stats[cid] = {"total": 0, "25_26": 0, "last_call": None}
            call_stats[cid]["total"] += 1
            if c["call_time"] and c["call_time"] >= "2025-01-01":
                call_stats[cid]["25_26"] += 1
            if c["call_time"] and (call_stats[cid]["last_call"] is None or c["call_time"] > call_stats[cid]["last_call"]):
                call_stats[cid]["last_call"] = c["call_time"]
        if len(batch) < PAGE:
            break
        offset += PAGE

    print(f"  {len(call_stats)} clients with calls")

    print("Loading order stats...")
    order_stats = {}
    offset = 0
    while True:
        batch = sb.table("orders").select("client_id,issue_date_parsed").range(offset, offset + PAGE - 1).execute().data
        for o in batch:
            cid = o["client_id"]
            if cid not in order_stats:
                order_stats[cid] = {"orders": 0, "last_order": None, "active_booking": False}
            order_stats[cid]["orders"] += 1
            if o["issue_date_parsed"]:
                if order_stats[cid]["last_order"] is None or o["issue_date_parsed"] > order_stats[cid]["last_order"]:
                    order_stats[cid]["last_order"] = o["issue_date_parsed"]
                if o["issue_date_parsed"] >= "2026-04-01":
                    order_stats[cid]["active_booking"] = True
        if len(batch) < PAGE:
            break
        offset += PAGE

    print(f"  {len(order_stats)} clients with orders")

    # Build rows
    print("Assembling audit rows...")
    rows = []
    for mm in mm_rows:
        cid = mm["client_id"]
        client = clients_by_id.get(cid, {})
        cr_list = cr_by_client.get(cid, [])
        cr_name = " | ".join(r.get("external_name") or "" for r in cr_list) if cr_list else ""
        cr_id = " | ".join(r.get("external_id") or "" for r in cr_list) if cr_list else ""

        stats = call_stats.get(cid, {})
        ostats = order_stats.get(cid, {})
        calls_total = stats.get("total", 0)
        calls_25_26 = stats.get("25_26", 0)
        orders = ostats.get("orders", 0)

        # Flag logic
        flags = []
        if calls_total > 0 and not cr_list:
            flags.append("HAS CALLS BUT NO CR MAPPING")
        if cr_name and mm.get("external_name"):
            mm_base = (mm["external_name"].split(" - ")[0] or "").lower().replace(" ", "")
            cr_base = cr_name.lower().replace(" ", "")
            if mm_base and cr_base and mm_base not in cr_base and cr_base not in mm_base:
                # Also allow matching on first 6+ chars to catch minor variants
                if len(mm_base) >= 6 and mm_base[:6] not in cr_base and (len(cr_base) < 6 or cr_base[:6] not in mm_base):
                    flags.append("NAME MISMATCH - VERIFY")
        if calls_total == 0 and orders > 0 and ostats.get("active_booking"):
            flags.append("ACTIVE BUT ZERO CALLS")
        if len(cr_list) > 1:
            flags.append("MULTIPLE CALLRAIL ACCOUNTS")

        rows.append({
            "mm_account": mm.get("external_name") or "",
            "mm_id": mm.get("external_id") or "",
            "db_name": client.get("name") or "",
            "status": client.get("status") or "",
            "market": markets_by_id.get(client.get("primary_market_id"), ""),
            "callrail_account": cr_name,
            "callrail_company_id": cr_id,
            "calls_total": calls_total,
            "calls_25_26": calls_25_26,
            "orders": orders,
            "last_order": ostats.get("last_order"),
            "last_call": stats.get("last_call"),
            "active_booking": "Y" if ostats.get("active_booking") else "",
            "flag": "; ".join(flags),
        })

    rows.sort(key=lambda r: (-r["calls_total"], r["mm_account"]))

    print("Fetching CallRail companies from API...")
    cr_companies = fetch_callrail_companies()
    print(f"  {len(cr_companies)} CallRail companies across all 3 accounts")

    # Map CR company_id -> (name, account)
    cr_by_id = {c["id"]: c for c in cr_companies}
    mapped_ids = {r["external_id"] for r in cr_rows}

    # Count unassigned calls per CR company
    print("Counting unassigned calls per CR company...")
    unassigned_counts = {}
    offset = 0
    while True:
        batch = (
            sb.table("calls")
            .select("callrail_company_id")
            .is_("client_id", "null")
            .range(offset, offset + PAGE - 1)
            .execute()
            .data
        )
        for c in batch:
            cid = c["callrail_company_id"]
            if cid:
                unassigned_counts[cid] = unassigned_counts.get(cid, 0) + 1
        if len(batch) < PAGE:
            break
        offset += PAGE

    # Write Excel
    print("Writing Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = "CallRail Mapping Audit"

    headers = [
        "MM Account", "MM ID", "DB Client Name", "Status", "Market",
        "CallRail Account", "CallRail Company ID",
        "Calls (Total)", "Calls (2025-26)",
        "Orders", "Last Order", "Last Call",
        "Active Booking", "Flag",
        "Correct CallRail Account (edit)", "Notes (edit)",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    flag_fills = {
        "HAS CALLS BUT NO CR MAPPING": PatternFill("solid", fgColor="FFC7CE"),
        "NAME MISMATCH - VERIFY": PatternFill("solid", fgColor="FFEB9C"),
        "MULTIPLE CALLRAIL ACCOUNTS": PatternFill("solid", fgColor="FFD966"),
        "ACTIVE BUT ZERO CALLS": PatternFill("solid", fgColor="DDEBF7"),
    }

    for r in rows:
        ws.append([
            r["mm_account"], r["mm_id"], r["db_name"], r["status"], r["market"],
            r["callrail_account"], r["callrail_company_id"],
            r["calls_total"], r["calls_25_26"],
            r["orders"], r["last_order"], r["last_call"],
            r["active_booking"], r["flag"],
            "", "",  # Correct CallRail Account, Notes — user-editable
        ])
        if r["flag"]:
            # Color the flag cell based on the highest-priority flag present
            flag_cell = ws.cell(row=ws.max_row, column=len(headers))
            for key, fill in flag_fills.items():
                if key in r["flag"]:
                    flag_cell.fill = fill
                    break

    # Column widths
    widths = {
        "A": 42, "B": 15, "C": 42, "D": 12, "E": 8,
        "F": 38, "G": 38, "H": 12, "I": 13,
        "J": 10, "K": 13, "L": 13, "M": 14, "N": 38,
        "O": 40, "P": 30,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # CallRail Accounts reference sheet (shows all CR companies + unmapped call counts)
    ws_ref = wb.create_sheet("CallRail Accounts")
    ws_ref.append(["CallRail Account Name", "Account", "Company ID", "Currently Mapped?", "Unassigned Calls"])
    for cell in ws_ref[1]:
        cell.font = header_font
        cell.fill = header_fill

    cr_ref_rows = []
    for c in cr_companies:
        cr_ref_rows.append({
            "name": c["name"],
            "account": c["account"],
            "id": c["id"],
            "mapped": "Yes" if c["id"] in mapped_ids else "No",
            "unassigned": unassigned_counts.get(c["id"], 0),
        })
    cr_ref_rows.sort(key=lambda x: (x["mapped"] == "Yes", -x["unassigned"], x["name"]))
    for r in cr_ref_rows:
        ws_ref.append([r["name"], r["account"], r["id"], r["mapped"], r["unassigned"]])
        if r["mapped"] == "No" and r["unassigned"] > 0:
            ws_ref.cell(row=ws_ref.max_row, column=5).fill = PatternFill("solid", fgColor="FFC7CE")
    ws_ref.column_dimensions["A"].width = 50
    ws_ref.column_dimensions["B"].width = 10
    ws_ref.column_dimensions["C"].width = 40
    ws_ref.column_dimensions["D"].width = 17
    ws_ref.column_dimensions["E"].width = 18
    ws_ref.freeze_panes = "A2"
    ws_ref.auto_filter.ref = ws_ref.dimensions

    # Dropdown validation on the "Correct CallRail Account" column pointing at the reference sheet
    all_cr_names = sorted({c["name"] for c in cr_companies})
    # Named range for dropdown — Excel limits list-string validation to 255 chars, so use a range reference
    from openpyxl.workbook.defined_name import DefinedName
    # Write names into a hidden sheet for the range
    ws_list = wb.create_sheet("_cr_names_hidden")
    ws_list.sheet_state = "hidden"
    for i, name in enumerate(all_cr_names, start=1):
        ws_list.cell(row=i, column=1, value=name)
    last_row = len(all_cr_names)
    range_ref = f"_cr_names_hidden!$A$1:$A${last_row}"
    dv = DataValidation(type="list", formula1=f"={range_ref}", allow_blank=True)
    dv.add(f"O2:O{1 + len(rows)}")
    ws.add_data_validation(dv)

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Count"])
    ws2["A1"].font = header_font
    ws2["A1"].fill = header_fill
    ws2["B1"].font = header_font
    ws2["B1"].fill = header_fill

    summary = [
        ("Total MagManager accounts", len(rows)),
        ("With CallRail mapping", sum(1 for r in rows if r["callrail_account"])),
        ("Without CallRail mapping", sum(1 for r in rows if not r["callrail_account"])),
        ("Has calls but NO CR mapping", sum(1 for r in rows if "HAS CALLS BUT NO CR MAPPING" in r["flag"])),
        ("Name mismatches to verify", sum(1 for r in rows if "NAME MISMATCH" in r["flag"])),
        ("Multiple CallRail accounts", sum(1 for r in rows if "MULTIPLE" in r["flag"])),
        ("Active but zero calls", sum(1 for r in rows if "ACTIVE BUT ZERO CALLS" in r["flag"])),
        ("", ""),
        ("Total calls across all", sum(r["calls_total"] for r in rows)),
        ("Calls in 2025-26", sum(r["calls_25_26"] for r in rows)),
    ]
    for k, v in summary:
        ws2.append([k, v])
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 14

    out_path = OUTPUT_DIR / f"callrail_mapping_audit_{date.today().isoformat()}.xlsx"
    wb.save(out_path)
    print(f"\nDone. Saved: {out_path}")
    print(f"  {len(rows)} MM accounts audited")


if __name__ == "__main__":
    main()
