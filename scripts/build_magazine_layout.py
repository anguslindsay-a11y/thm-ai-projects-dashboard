"""
Build a draft magazine layout for a given market/zone + issue.

Inputs:
  - Issue-board xlsx export (from the THM Colorado issue board)
  - Prior-issue placements from Supabase (for rotation)
  - clients.category from Supabase (for conflict detection)

Output:
  output/layout_{zone}_{issue}.xlsx
    - Layout     : page-by-page assignments with size, category, prior page, notes
    - By Advertiser: every ad + where it landed + where it was last issue
    - Conflicts  : any same-spread/same-page category overlaps left in the draft
    - Unplaced   : ads that couldn't be placed
    - Specials   : page-request ads and whether the request was honored

Usage:
  python scripts/build_magazine_layout.py \
    --file "data/List EPC May 2026 - THM Colorado - IssueBoard.xlsx" \
    --zone EPC --prior-issue 2603 --pages 52
"""

import sys
import os
import re
import argparse
from pathlib import Path
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
load_dotenv()

from supabase import create_client
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)


# ---------- Name matching ----------

def normalize(name: str) -> str:
    if not name:
        return ""
    s = name.lower()
    s = re.sub(r"\bthe\b", " ", s)
    s = re.sub(r"[,&/().\-]", " ", s)
    s = re.sub(r"\b(llc|inc|co|company|ltd|corp)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def strip_suffix(name: str) -> str:
    """Drop trailing ' - EPC', ' - NLA Media', ' (Main Street Media)', etc."""
    if not name:
        return ""
    s = re.sub(r"\s*\([^)]*\)\s*$", "", name)
    s = re.sub(r"\s*-\s*(EPC|NLA Media|Spartan Marketing|White Hat Marketing|Main Street Media|GWRK|Flat Branch)\s*$", "", s, flags=re.IGNORECASE)
    return s.strip()


def match_advertiser_to_client(advertiser: str, clients_by_id: dict) -> dict | None:
    """Return best-matching client record or None. Prefers suffix-preserving
    matches (e.g. '...- EPC') over stripped base names."""
    if not advertiser or advertiser.lower() == "available":
        return None
    adv_norm = normalize(advertiser)
    stripped_norm = normalize(strip_suffix(advertiser))

    # 1. Exact normalized match to full advertiser string (preserves suffix)
    for c in clients_by_id.values():
        if normalize(c["name"]) == adv_norm:
            return c
    # 2. Exact normalized match to stripped version
    for c in clients_by_id.values():
        if normalize(c["name"]) == stripped_norm:
            return c
    # 3. Contained match — prefer clients whose category is populated and name length is closest
    best = None
    best_score = (-1, 0)  # (has_category, overlap_len)
    for c in clients_by_id.values():
        cn = normalize(c["name"])
        if not cn:
            continue
        for t in (adv_norm, stripped_norm):
            if t and (t in cn or cn in t):
                overlap = min(len(t), len(cn))
                has_cat = 1 if c.get("category") else 0
                score = (has_cat, overlap)
                if score > best_score:
                    best_score = score
                    best = c
    return best


# ---------- Load issue board ----------

def load_issue_board(path: Path) -> tuple[list[dict], list[dict]]:
    """Returns (book_ads, excluded). Auto-detects header row.
    Marketplace-feature/spotlight and Certified Directory are kept in the book
    and flagged front_section=True so they land before page 15."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header_row = 1
    for i, row in enumerate(ws.iter_rows(max_row=3, values_only=True), 1):
        if row and row[0] and "Pg Type" in str(row[0]):
            header_row = i
            break
    rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))
    ads = []
    excluded = []
    for r in rows:
        if not r[2]:
            continue
        pg_type = r[0] or ""
        advertiser = r[2]
        size = str(r[3])
        status = r[4]
        price = r[5] or 0
        pg_request = r[7] if len(r) > 7 else None
        weight = float(r[8] or 0) if len(r) > 8 else 0
        rep = r[9] if len(r) > 9 else None
        notes = (r[10] or "") if len(r) > 10 else ""
        front_section = (
            "Marketplace" in str(pg_type)
            or "Certified Directory" in str(pg_type)
            or "Certified Directory" in str(advertiser)
        )
        ad = {
            "pg_type": pg_type,
            "advertiser": advertiser,
            "size": size,
            "status": status,
            "price": price,
            "pg_request": pg_request,
            "weight": weight,
            "rep": rep,
            "notes": notes,
            "front_section": front_section,
        }
        ads.append(ad)
    return ads, excluded


# ---------- Supabase pulls ----------

def fetch_clients(sb):
    rows = []
    offset = 0
    while True:
        batch = (
            sb.table("clients")
            .select("id,name,category,status,primary_market_id")
            .range(offset, offset + 999)
            .execute()
            .data
        )
        rows.extend(batch)
        if len(batch) < 1000:
            break
        offset += 1000
    return {r["id"]: r for r in rows}


def fetch_prior_placements(sb, zone_id: str, issue_code: str) -> dict:
    """Returns {normalized_source_client_name: page}."""
    rows = (
        sb.table("ad_placements")
        .select("source_client_name,page,client_id,ad_size")
        .eq("zone_id", zone_id)
        .eq("issue_code", issue_code)
        .execute()
        .data
    )
    by_name = {}
    by_client = defaultdict(list)
    for r in rows:
        n = normalize(r.get("source_client_name") or "")
        if n:
            by_name[n] = r["page"]
        if r.get("client_id"):
            by_client[r["client_id"]].append(r["page"])
    return by_name, by_client


# ---------- Category helpers ----------

def category_tokens(cat: str | None) -> set[str]:
    if not cat:
        return set()
    return {p.strip() for p in cat.split(",") if p.strip()}


def categories_conflict(a: str | None, b: str | None) -> bool:
    ta, tb = category_tokens(a), category_tokens(b)
    return bool(ta & tb)


# ---------- Layout ----------

def rotate_zone(prior_page: int | None, total_pages: int) -> str:
    """Given prior page, where should we aim this issue?
    Returns one of: 'front', 'middle', 'back', 'any'."""
    if not prior_page:
        return "any"
    third = total_pages // 3
    if prior_page <= third:
        return "back"
    if prior_page >= 2 * third:
        return "front"
    # Middle last time -> push toward front or back (we'll pick based on space)
    return "edge"


def pages_in_zone(zone: str, total_pages: int) -> list[int]:
    third = total_pages // 3
    if zone == "front":
        return list(range(2, third + 1))
    if zone == "back":
        return list(range(total_pages - third, total_pages))
    if zone == "middle":
        return list(range(third + 1, 2 * third + 1))
    if zone == "edge":
        return list(range(2, third + 1)) + list(range(total_pages - third, total_pages))
    return list(range(2, total_pages))


def spread_of(page: int) -> int:
    """Spread index — p2+p3 share spread 1, p4+p5 share spread 2, etc."""
    if page <= 1 or page % 2 == 1:
        # odd pages 3,5,... are RIGHT-HAND; pair with page-1
        return (page - 1) // 2 if page > 1 else 0
    # even pages 2,4,... are LEFT-HAND
    return page // 2


def build_layout(ads: list[dict], clients_by_id: dict, prior_by_name: dict, prior_by_client: dict, total_pages: int = 52, reserved_pages: dict | None = None, proximity: int = 3):
    """
    Returns:
      placements: list of dicts keyed by page with list of ads on that page
      unplaced: ads not placed
    """
    # Attach category + prior-page metadata to each ad
    for ad in ads:
        client = match_advertiser_to_client(ad["advertiser"], clients_by_id)
        ad["client_id"] = client["id"] if client else None
        ad["category"] = client["category"] if client else None
        ad["db_name"] = client["name"] if client else None
        # Prior page
        prior = prior_by_name.get(normalize(ad["advertiser"]))
        if not prior and ad["client_id"]:
            pages = prior_by_client.get(ad["client_id"]) or []
            prior = pages[0] if pages else None
        ad["prior_page"] = prior
        ad["target_zone"] = rotate_zone(prior, total_pages)

    # Initialize page grid. Each page has a list of slots (positions).
    # p1 = front cover (1 slot full)
    # p52 = back cover (slots: 2/3 and 1/3)
    # p2..p51 = interior; each is a "full" slot that can be split into 2 halves or 4 quarters etc
    pages = {p: {"slots": [], "categories": []} for p in range(1, total_pages + 1)}

    def page_remaining(p):
        used = sum(s["weight"] for s in pages[p]["slots"])
        return 1.0 - used

    def can_place(p, weight, category, within_page_ok=False, proximity_override: int | None = None):
        """Hard-rules: no same category on same spread or same page.
        Soft-rule: avoid same category within `proximity` pages (defaults to module-level proximity).
        Set proximity_override=0 to disable soft rule for fallback placement."""
        if page_remaining(p) + 1e-6 < weight:
            return False
        # Same-page conflict (two halves same category)
        if not within_page_ok:
            for s in pages[p]["slots"]:
                if categories_conflict(s.get("category"), category):
                    return False
        # Same-spread (hard rule)
        spread = spread_of(p)
        for other_p in (p - 1, p + 1):
            if 1 < other_p < total_pages and spread_of(other_p) == spread:
                for s in pages[other_p]["slots"]:
                    if categories_conflict(s.get("category"), category):
                        return False
        # Proximity (soft rule) — same category within N pages
        prox = proximity if proximity_override is None else proximity_override
        if prox > 0:
            for delta in range(1, prox + 1):
                for other_p in (p - delta, p + delta):
                    if 1 < other_p < total_pages and spread_of(other_p) != spread:
                        for s in pages[other_p]["slots"]:
                            if categories_conflict(s.get("category"), category):
                                return False
        return True

    def place(p, ad, position_label=""):
        pages[p]["slots"].append({
            "ad": ad,
            "weight": ad["weight"],
            "category": ad["category"],
            "position": position_label,
        })
        ad["_placed"] = p

    unplaced = []

    # --- Pin reserved (editorial) pages ---
    reserved_pages = reserved_pages or {}
    for p, label in reserved_pages.items():
        if 1 <= p <= total_pages:
            pages[p]["slots"].append({
                "ad": {"advertiser": f"[Editorial — {label}]", "size": "Full", "price": 0, "rep": "", "pg_request": "", "notes": "reserved", "category": None, "prior_page": None, "db_name": None},
                "weight": 1.0,
                "category": None,
                "position": f"Reserved: {label}",
            })

    # --- Pin covers ---
    landscaper_front = next((a for a in ads if a["pg_type"] == "Front"), None)
    if landscaper_front:
        place(1, landscaper_front, "Front Cover")

    for a in ads:
        if a["pg_type"] == "Back":
            place(total_pages, a, "Back Cover")

    # --- Pin front-section ads (Marketplace / Certified Directory) to pages 2-14 ---
    front_ads = [a for a in ads if a.get("front_section") and not a.get("_placed")]
    # Spreads first (need facing pages), then fulls, then halves
    front_ads.sort(key=lambda a: (-a.get("weight", 0), -a.get("price", 0)))
    for a in front_ads:
        is_spread = a.get("weight", 0) >= 2 or str(a.get("size", "")).lower() == "spread"
        placed_here = False
        if is_spread:
            for p in range(2, 14):  # left-page candidates 2..13, right goes to 3..14
                if p % 2 != 0:
                    continue
                odd_p = p + 1
                if odd_p > 14:
                    continue
                if p in reserved_pages or odd_p in reserved_pages:
                    continue
                if page_remaining(p) != 1.0 or page_remaining(odd_p) != 1.0:
                    continue
                if not can_place(p, 1.0, a.get("category"), within_page_ok=True):
                    continue
                if not can_place(odd_p, 1.0, a.get("category"), within_page_ok=True):
                    continue
                pages[p]["slots"].append({"ad": a, "weight": 1.0, "category": a.get("category"),
                                          "position": f"Front-section Spread (L) with p{odd_p}"})
                pages[odd_p]["slots"].append({"ad": a, "weight": 1.0, "category": a.get("category"),
                                              "position": f"Front-section Spread (R) with p{p}"})
                a["_placed"] = p
                placed_here = True
                break
        else:
            for p in range(2, 15):
                if p in reserved_pages:
                    continue
                if page_remaining(p) < a.get("weight", 1.0):
                    continue
                if not can_place(p, a.get("weight", 1.0), a.get("category")):
                    continue
                place(p, a, f"Front-section ({a.get('pg_type')})")
                placed_here = True
                break
        if not placed_here:
            # Couldn't fit before page 15; fall through to general placement
            a["_front_section_overflow"] = True

    # --- Pin 2-page spreads (occupy facing pages: even + odd), distributed evenly ---
    spreads = [a for a in ads if (a.get("pg_type") == "Spread" or str(a.get("size", "")).lower() == "spread") and not a.get("_placed")]
    if spreads:
        # Spread them across the interior, excluding front-cover-adjacent & back-cover-adjacent spreads
        mid_start = max(6, total_pages // 4)
        mid_end = total_pages - 6
        # Generate evenly-spaced target left-page slots
        n = len(spreads)
        if n == 1:
            targets = [((mid_start + mid_end) // 2) // 2 * 2]  # nearest even to middle
            if targets[0] % 2 == 1:
                targets[0] -= 1
        else:
            span = mid_end - mid_start
            step = span / (n + 1)
            targets = []
            for i in range(1, n + 1):
                raw = int(round(mid_start + step * i))
                # round to nearest even (left-page)
                if raw % 2 == 1:
                    raw -= 1
                targets.append(raw)
        # Sort spreads by priority (e.g., cover sponsor spread last to avoid conflict with cover, highest price first)
        spreads_sorted = sorted(spreads, key=lambda a: -a.get("price", 0))
        for sp, target in zip(spreads_sorted, targets):
            # Starting from target, expand outward to find a clear pair of pages
            placed = False
            for offset in range(0, total_pages, 2):
                for direction in (+1, -1):
                    even_p = target + direction * offset
                    if even_p % 2 == 1:
                        continue
                    if even_p < 2 or even_p + 1 >= total_pages:
                        continue
                    odd_p = even_p + 1
                    if even_p in reserved_pages or odd_p in reserved_pages:
                        continue
                    if page_remaining(even_p) != 1.0 or page_remaining(odd_p) != 1.0:
                        continue
                    # No spread-level conflict with neighbors
                    if not can_place(even_p, 1.0, sp.get("category"), within_page_ok=True):
                        continue
                    if not can_place(odd_p, 1.0, sp.get("category"), within_page_ok=True):
                        continue
                    pages[even_p]["slots"].append({
                        "ad": sp, "weight": 1.0, "category": sp.get("category"),
                        "position": f"Spread (L) with p{odd_p}",
                    })
                    pages[odd_p]["slots"].append({
                        "ad": sp, "weight": 1.0, "category": sp.get("category"),
                        "position": f"Spread (R) with p{even_p}",
                    })
                    sp["_placed"] = even_p
                    placed = True
                    break
                if placed:
                    break
            if not placed:
                unplaced.append(sp)

    # --- Pin special requests ---
    def parse_request(ad):
        req = (ad.get("pg_request") or "") + " " + (ad.get("notes") or "")
        req_l = req.lower()
        pinned_pages = []
        # Explicit page numbers
        for m in re.finditer(r"page\s*(\d+)(?:\s*or\s*(\d+))?", req_l):
            pinned_pages.append(int(m.group(1)))
            if m.group(2):
                pinned_pages.append(int(m.group(2)))
        m_first = re.search(r"(?:first|1st)\s*(\d+)\s*(?:page|pg)", req_l)
        if m_first:
            n = int(m_first.group(1))
            pinned_pages.extend(range(2, n + 2))
        elif "first 12" in req_l:
            pinned_pages.extend(range(2, 14))
        elif "first 10" in req_l:
            pinned_pages.extend(range(2, 12))
        if "inside front" in req_l:
            pinned_pages.extend([2, 3, 4])
        if "inside back" in req_l:
            pinned_pages.extend([total_pages - 1, total_pages - 3])
        if "near back" in req_l or "near the back" in req_l:
            pinned_pages.extend(range(total_pages - 5, total_pages))
        want_right = bool(re.search(r"(right.hand|rhr|right)", req_l))
        want_left = bool(re.search(r"(left.hand|lhr|left)", req_l)) and not want_right
        return pinned_pages, want_right, want_left

    # Sort specials by specificity (explicit page numbers first, then positional requests)
    def has_placement_intent(a):
        if a.get("pg_request"):
            return True
        notes_l = (a.get("notes") or "").lower()
        return any(k in notes_l for k in ["right hand", "left hand", "rhr", "lhr", "first 1", "first 2", "inside front", "inside back", "near back", "near front", "page 3", "page 5", "page 7", "page 9"])

    specials = [a for a in ads if has_placement_intent(a) and not a.get("_placed")]
    def specificity(a):
        pinned, _, _ = parse_request(a)
        explicit = any(p < 20 or p > 40 for p in pinned[:1]) and len(pinned) <= 4
        return (0 if explicit else 1, -a["price"])
    specials.sort(key=specificity)

    for ad in specials:
        if ad.get("_placed"):
            continue
        pinned_pages, want_right, want_left = parse_request(ad)
        # If side-only request with no pinned pages, use front-half right/left pages as target
        if not pinned_pages and (want_right or want_left):
            pinned_pages = list(range(2, total_pages))
        # Filter candidates
        candidates = []
        for p in pinned_pages:
            if p < 2 or p >= total_pages:
                continue
            if want_right and p % 2 == 0:
                continue
            if want_left and p % 2 == 1:
                continue
            if can_place(p, ad["weight"], ad["category"]):
                candidates.append(p)
        # Fallback: allow same-page conflict if positional request is tight
        if not candidates:
            for p in pinned_pages:
                if p < 2 or p >= total_pages:
                    continue
                if page_remaining(p) + 1e-6 >= ad["weight"]:
                    candidates.append(p)
        if candidates:
            place(candidates[0], ad, f"Special: {ad.get('pg_request') or 'notes'}")
        # If still no candidate, leave for normal placement below

    # --- Fill partially filled pages first (specials that took a half left 0.5 open) ---
    remaining = [a for a in ads if not a.get("_placed")]

    def fill_partial_pages(remaining):
        """Fill pages that have some space but aren't full, preferring exact-size fills.
        Returns updated remaining list."""
        still_remaining = []
        for ad in remaining:
            placed = False
            # Look for partially filled pages where this ad fits
            partial_pages = sorted(
                [p for p in range(2, total_pages) if 0 < 1.0 - page_remaining(p) < 1.0 and page_remaining(p) + 1e-6 >= ad["weight"]],
                key=lambda p: page_remaining(p),  # tightest fit first
            )
            for p in partial_pages:
                if can_place(p, ad["weight"], ad["category"]):
                    place(p, ad)
                    placed = True
                    break
            if not placed:
                still_remaining.append(ad)
        return still_remaining

    # First pass: fill partial pages with halves (halves are more flexible than quarters)
    halves_first = sorted(remaining, key=lambda a: (-a["weight"], -a["price"]))
    remaining = fill_partial_pages(halves_first)

    # --- Build page units from remaining ads ---
    # Each unit is a composition that fills exactly 1 page
    def build_units(ads_left):
        fulls = [a for a in ads_left if a["weight"] > 0.9]
        halves = [a for a in ads_left if 0.4 < a["weight"] < 0.6]
        quarters = [a for a in ads_left if 0.2 < a["weight"] < 0.3]

        # Sort halves/quarters by price desc (premium ads get paired first)
        halves.sort(key=lambda a: -a["price"])
        quarters.sort(key=lambda a: -a["price"])

        units = []

        # Full-page units
        for f in fulls:
            units.append({
                "ads": [f],
                "type": "full",
                "priority": -f["price"],
            })

        # NEW STRATEGY: pair every 2 quarters with 1 half FIRST to avoid 4-quarter pages.
        # Then pair remaining halves with halves.
        q_used = set()
        h_used = set()

        # Step 1: quarter-pairs + half, half at a time
        quarter_pairs = []
        i = 0
        while i + 1 < len(quarters):
            q1, q2 = quarters[i], quarters[i + 1]
            # Try to keep 2 quarters non-conflicting; if they conflict, skip q2 and try later
            if not categories_conflict(q1.get("category"), q2.get("category")):
                quarter_pairs.append((i, i + 1))
                i += 2
            else:
                # advance q2 search
                found = False
                for j in range(i + 2, len(quarters)):
                    if not categories_conflict(q1.get("category"), quarters[j].get("category")):
                        quarter_pairs.append((i, j))
                        # swap quarters[j] with quarters[i+1] so the loop can skip i+1 later
                        quarters[i + 1], quarters[j] = quarters[j], quarters[i + 1]
                        i += 2
                        found = True
                        break
                if not found:
                    # lonely quarter — treat as pair-of-one
                    quarter_pairs.append((i, None))
                    i += 1
        if i < len(quarters):  # trailing single quarter
            quarter_pairs.append((i, None))

        # For each quarter-pair, find the best half partner (non-conflicting)
        for qi1, qi2 in quarter_pairs:
            q1 = quarters[qi1]
            q2 = quarters[qi2] if qi2 is not None else None
            qs = [q1] + ([q2] if q2 else [])
            qs_categories = [q.get("category") for q in qs]
            partner_h_idx = None
            for hi, h in enumerate(halves):
                if hi in h_used:
                    continue
                if any(categories_conflict(h.get("category"), qc) for qc in qs_categories):
                    continue
                partner_h_idx = hi
                break
            if partner_h_idx is not None:
                units.append({
                    "ads": [halves[partner_h_idx]] + qs,
                    "type": "half+quarters",
                    "priority": -halves[partner_h_idx]["price"],
                })
                h_used.add(partner_h_idx)
                q_used.add(qi1)
                if qi2 is not None:
                    q_used.add(qi2)
            else:
                # No available non-conflicting half. Place quarters together (max 2, never 4)
                units.append({
                    "ads": qs,
                    "type": "quarters-pair",
                    "priority": -sum(q["price"] for q in qs) / len(qs),
                })
                q_used.add(qi1)
                if qi2 is not None:
                    q_used.add(qi2)

        # Step 2: pair remaining halves with halves
        remaining_halves = [(i, h) for i, h in enumerate(halves) if i not in h_used]
        paired_h = set()
        for idx, (hi, h1) in enumerate(remaining_halves):
            if hi in paired_h:
                continue
            partner_idx = None
            for jdx in range(idx + 1, len(remaining_halves)):
                hj, h2 = remaining_halves[jdx]
                if hj in paired_h:
                    continue
                if not categories_conflict(h1.get("category"), h2.get("category")):
                    partner_idx = jdx
                    break
            if partner_idx is not None:
                hj, h2 = remaining_halves[partner_idx]
                units.append({
                    "ads": [h1, h2],
                    "type": "halves",
                    "priority": -(h1["price"] + h2["price"]) / 2,
                })
                paired_h.add(hi)
                paired_h.add(hj)

        # Step 3: any leftover orphan halves (rare)
        for hi, h in remaining_halves:
            if hi not in paired_h:
                units.append({
                    "ads": [h],
                    "type": "half-alone",
                    "priority": -h["price"],
                })

        # Compute target_zone per unit (majority of ads' target_zones; "any" if mixed)
        for u in units:
            zones = [a.get("target_zone", "any") for a in u["ads"]]
            # Score: front=-1, back=+1, middle=0, edge=random
            zone_counts = {}
            for z in zones:
                zone_counts[z] = zone_counts.get(z, 0) + 1
            # Pick the most common non-"any" zone, else "any"
            non_any = {z: c for z, c in zone_counts.items() if z != "any"}
            if non_any:
                u["target_zone"] = max(non_any.items(), key=lambda x: x[1])[0]
            else:
                u["target_zone"] = "any"
        return units

    units = build_units(remaining)

    # --- Interleave units by type when placing ---
    # Split by type class for interleaving: full vs non-full (half/mixed/quarter pages)
    fulls_q = sorted([u for u in units if u["type"] == "full"], key=lambda u: u["priority"])
    nonfull_q = sorted([u for u in units if u["type"] != "full"], key=lambda u: u["priority"])

    # Interleave queue: alternate full / non-full so adjacent pages blend sizes
    placement_queue = []
    nf_ratio = len(nonfull_q) / max(1, len(fulls_q) + len(nonfull_q))
    fi = ni = 0
    idx = 0
    while fi < len(fulls_q) or ni < len(nonfull_q):
        # Proportion of non-fulls placed so far
        placed_total = fi + ni
        placed_nf_ratio = ni / placed_total if placed_total else 0
        want_nonfull = (placed_nf_ratio < nf_ratio) and ni < len(nonfull_q)
        if want_nonfull and ni < len(nonfull_q):
            placement_queue.append(nonfull_q[ni])
            ni += 1
        elif fi < len(fulls_q):
            placement_queue.append(fulls_q[fi])
            fi += 1
        else:
            placement_queue.append(nonfull_q[ni])
            ni += 1
        idx += 1

    # --- Assign pages to units using rotation + conflict awareness ---
    def zone_pages(zone):
        return [p for p in pages_in_zone(zone, total_pages) if 1 < p < total_pages]

    def candidate_pages_for_unit(unit):
        tz = unit["target_zone"]
        if tz == "any":
            order = list(range(2, total_pages))
        elif tz == "edge":
            order = zone_pages("front") + zone_pages("back") + zone_pages("middle")
        else:
            opposite = "back" if tz == "front" else "front"
            order = zone_pages(tz) + zone_pages("middle") + zone_pages(opposite)
        # Dedupe preserving order
        seen = set()
        return [p for p in order if not (p in seen or seen.add(p))]

    def unit_fits_empty(p, unit, prox_override=None):
        total_w = sum(a["weight"] for a in unit["ads"])
        if page_remaining(p) + 1e-6 < total_w:
            return False
        for a in unit["ads"]:
            if not can_place(p, a["weight"], a.get("category"), within_page_ok=True, proximity_override=prox_override):
                return False
        return True

    def try_place_unit(unit, prox_override=None):
        candidates = candidate_pages_for_unit(unit)
        # Prefer empty pages
        for p in candidates:
            if page_remaining(p) == 1.0 and unit_fits_empty(p, unit, prox_override=prox_override):
                for a in unit["ads"]:
                    place(p, a)
                return True
        # Try partially-filled
        total_w = sum(a["weight"] for a in unit["ads"])
        for p in candidates:
            if page_remaining(p) + 1e-6 >= total_w and page_remaining(p) < 1.0:
                if all(can_place(p, a["weight"], a.get("category"), proximity_override=prox_override) for a in unit["ads"]):
                    for a in unit["ads"]:
                        place(p, a)
                    return True
        return False

    # Pass 1: strict proximity. Pass 2: relax to 1. Pass 3: proximity=0 (hard rules only).
    pending_units = list(placement_queue)
    for prox in (proximity, max(1, proximity // 2), 0):
        still_pending = []
        for unit in pending_units:
            if not try_place_unit(unit, prox_override=prox):
                still_pending.append(unit)
        pending_units = still_pending
        if not pending_units:
            break

    # Last resort: place individual ads with conflict override (try ANY page with space)
    for unit in pending_units:
        for a in unit["ads"]:
            placed_ad = False
            # Prefer candidate pages first, then ALL pages, then overflow
            order_a = candidate_pages_for_unit(unit)
            order_b = list(range(2, total_pages))
            for p in order_a + [p for p in order_b if p not in order_a]:
                if page_remaining(p) + 1e-6 >= a["weight"]:
                    place(p, a, "CONFLICT OVERRIDE")
                    placed_ad = True
                    break
            if not placed_ad:
                unplaced.append(a)

    return pages, unplaced


# ---------- Output ----------

def write_output(pages, ads, unplaced, prior_by_name, prior_by_client, zone, issue, total_pages, out_path):
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    special_fill = PatternFill("solid", fgColor="FFF2CC")
    conflict_fill = PatternFill("solid", fgColor="FFC7CE")
    cover_fill = PatternFill("solid", fgColor="D9E1F2")

    def style_header(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # ---- Layout sheet ----
    ws = wb.active
    ws.title = "Layout"
    ws.append(["Page", "Spread", "Side", "Position", "Advertiser", "Size", "Category", "Prior Page", "Rotation", "Rep", "Special Request", "Notes"])
    style_header(ws)

    # Conflict check — skip self-conflicts (spread ad referencing itself)
    conflict_pages = set()
    for p in range(2, total_pages):
        spread = spread_of(p)
        for s1 in pages[p]["slots"]:
            for other_p in (p - 1, p + 1):
                if 1 < other_p < total_pages and spread_of(other_p) == spread:
                    for s2 in pages[other_p]["slots"]:
                        if s1["ad"] is s2["ad"]:
                            continue
                        if categories_conflict(s1.get("category"), s2.get("category")):
                            conflict_pages.add(p)
                            conflict_pages.add(other_p)
        # Within-page
        slots = pages[p]["slots"]
        for i in range(len(slots)):
            for j in range(i + 1, len(slots)):
                if slots[i]["ad"] is slots[j]["ad"]:
                    continue
                if categories_conflict(slots[i].get("category"), slots[j].get("category")):
                    conflict_pages.add(p)

    for p in range(1, total_pages + 1):
        side = "Cover" if p in (1, total_pages) else ("L" if p % 2 == 0 else "R")
        spread = spread_of(p) if p not in (1, total_pages) else ""
        slots = pages[p]["slots"]
        if not slots:
            ws.append([p, spread, side, "OPEN", "", "", "", "", "", "", "", ""])
            continue
        for s in slots:
            ad = s["ad"]
            rot = "—"
            if ad.get("prior_page"):
                rot = f"was p{ad['prior_page']} → p{p}"
            elif ad.get("prior_page") is None:
                rot = "new/no prior"
            ws.append([
                p, spread, side, s.get("position") or "",
                ad["advertiser"], ad["size"],
                ad.get("category") or "—",
                ad.get("prior_page") or "",
                rot,
                ad.get("rep") or "",
                ad.get("pg_request") or "",
                ad.get("notes") or "",
            ])
            row = ws.max_row
            if p in (1, total_pages):
                for c in ws[row]:
                    c.fill = cover_fill
            elif p in conflict_pages:
                for c in ws[row]:
                    c.fill = conflict_fill
            elif ad.get("pg_request"):
                for c in ws[row]:
                    c.fill = special_fill

    widths = {"A": 6, "B": 8, "C": 6, "D": 18, "E": 45, "F": 10, "G": 35, "H": 10, "I": 24, "J": 18, "K": 24, "L": 50}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ---- By Advertiser sheet ----
    ws2 = wb.create_sheet("By Advertiser")
    ws2.append(["Advertiser", "DB Name", "Category", "Size", "Price", "Rep", "Placed Page", "Prior Page", "Move", "Special Request", "Notes"])
    style_header(ws2)
    ads_sorted = sorted(ads, key=lambda a: (a.get("_placed") or 9999, a["advertiser"]))
    for ad in ads_sorted:
        placed = ad.get("_placed") or "UNPLACED"
        prior = ad.get("prior_page") or ""
        if prior and isinstance(placed, int):
            delta = placed - prior
            move = f"{delta:+d}"
        elif not prior:
            move = "new"
        else:
            move = ""
        ws2.append([
            ad["advertiser"], ad.get("db_name") or "—", ad.get("category") or "—",
            ad["size"], ad["price"], ad.get("rep") or "",
            placed, prior, move,
            ad.get("pg_request") or "", ad.get("notes") or "",
        ])
    for col, w in {"A": 42, "B": 42, "C": 32, "D": 10, "E": 10, "F": 18, "G": 12, "H": 12, "I": 10, "J": 26, "K": 50}.items():
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    # ---- Conflicts sheet ----
    ws3 = wb.create_sheet("Conflicts")
    ws3.append(["Page", "Spread", "Advertiser", "Category", "Conflicts With (Page)", "Conflicts With (Advertiser)", "Conflicts With (Category)"])
    style_header(ws3)
    for p in sorted(conflict_pages):
        spread = spread_of(p)
        for s in pages[p]["slots"]:
            # Within-page
            for s2 in pages[p]["slots"]:
                if s is s2 or s["ad"] is s2["ad"]:
                    continue
                if categories_conflict(s.get("category"), s2.get("category")):
                    ws3.append([p, spread, s["ad"]["advertiser"], s.get("category") or "", p, s2["ad"]["advertiser"], s2.get("category") or ""])
            # Spread
            for other_p in (p - 1, p + 1):
                if 1 < other_p < total_pages and spread_of(other_p) == spread:
                    for s2 in pages[other_p]["slots"]:
                        if s["ad"] is s2["ad"]:
                            continue
                        if categories_conflict(s.get("category"), s2.get("category")):
                            ws3.append([p, spread, s["ad"]["advertiser"], s.get("category") or "", other_p, s2["ad"]["advertiser"], s2.get("category") or ""])
    for col, w in {"A": 6, "B": 8, "C": 38, "D": 28, "E": 8, "F": 38, "G": 28}.items():
        ws3.column_dimensions[col].width = w

    # ---- Unplaced sheet ----
    ws4 = wb.create_sheet("Unplaced")
    ws4.append(["Advertiser", "Size", "Price", "Category", "Special Request", "Notes"])
    style_header(ws4)
    for ad in unplaced:
        ws4.append([ad["advertiser"], ad["size"], ad["price"], ad.get("category") or "", ad.get("pg_request") or "", ad.get("notes") or ""])

    # ---- Specials sheet ----
    ws5 = wb.create_sheet("Specials")
    ws5.append(["Advertiser", "Size", "Request", "Notes", "Placed Page", "Honored?"])
    style_header(ws5)
    for ad in ads:
        if not ad.get("pg_request") and "page" not in (ad.get("notes") or "").lower() and "inside" not in (ad.get("notes") or "").lower() and "right" not in (ad.get("notes") or "").lower():
            continue
        placed = ad.get("_placed")
        req = ad.get("pg_request") or ""
        notes = ad.get("notes") or ""
        honored = "—"
        if placed:
            req_l = (req + " " + notes).lower()
            if "first 12" in req_l and 2 <= placed <= 13:
                honored = "YES"
            elif "first 10" in req_l and 2 <= placed <= 11:
                honored = "YES"
            elif "inside front" in req_l and placed <= 5:
                honored = "YES"
            elif "inside back" in req_l and placed >= total_pages - 3:
                honored = "YES"
            elif "near back" in req_l and placed >= total_pages - 6:
                honored = "YES"
            elif "right" in req_l and placed % 2 == 1:
                honored = "YES"
            elif re.search(r"page\s*(\d+)", req_l):
                nums = [int(x) for x in re.findall(r"\d+", req_l)]
                honored = "YES" if placed in nums else "NO"
            else:
                honored = "CHECK"
        ws5.append([ad["advertiser"], ad["size"], req, notes, placed or "UNPLACED", honored])
    for col, w in {"A": 38, "B": 10, "C": 26, "D": 50, "E": 12, "F": 12}.items():
        ws5.column_dimensions[col].width = w

    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True, help="Path to issue board xlsx")
    parser.add_argument("--zone", required=True, help="Zone abbreviation (e.g., EPC)")
    parser.add_argument("--prior-issue", default=None, help="Prior issue_code for rotation (e.g., 2603)")
    parser.add_argument("--pages", type=int, default=56)
    parser.add_argument("--reserve", action="append", default=[], help="Reserve a page, e.g. --reserve 14:Certified Directory (repeatable)")
    parser.add_argument("--proximity", type=int, default=3, help="Avoid same category within this many pages (soft rule). Default 3.")
    parser.add_argument("--out", default=None)
    args = parser.parse_args()

    file_path = Path(args.file)
    if not file_path.exists():
        print(f"File not found: {file_path}")
        sys.exit(1)

    print(f"Loading issue board: {file_path.name}")
    ads, excluded = load_issue_board(file_path)
    print(f"  {len(ads)} book ads + {len(excluded)} excluded (marketplace/digital)")

    sb = create_client(SUPABASE_URL, SUPABASE_KEY)
    print("Loading clients from Supabase...")
    clients_by_id = fetch_clients(sb)
    print(f"  {len(clients_by_id)} clients loaded")

    prior_by_name, prior_by_client = {}, {}
    if args.prior_issue:
        zrow = sb.table("zones").select("id").eq("abbreviation", args.zone).execute().data
        if zrow:
            zone_id = zrow[0]["id"]
            print(f"Loading prior placements for {args.zone} {args.prior_issue}...")
            prior_by_name, prior_by_client = fetch_prior_placements(sb, zone_id, args.prior_issue)
            print(f"  {len(prior_by_name)} prior placements loaded")

    reserved = {}
    for spec in args.reserve:
        if ":" in spec:
            p_str, label = spec.split(":", 1)
        else:
            p_str, label = spec, "Editorial"
        try:
            reserved[int(p_str.strip())] = label.strip()
        except ValueError:
            print(f"Bad --reserve value: {spec}")

    print(f"Building layout ({args.pages} pages, reserved: {reserved or 'none'}, proximity={args.proximity})...")
    pages, unplaced = build_layout(ads, clients_by_id, prior_by_name, prior_by_client, args.pages, reserved, args.proximity)

    out = args.out or OUTPUT_DIR / f"layout_{args.zone}_{file_path.stem.replace(' ', '_')}.xlsx"
    write_output(pages, ads, unplaced, prior_by_name, prior_by_client, args.zone, args.prior_issue, args.pages, out)
    print(f"\nDone. Saved: {out}")
    print(f"  Unplaced: {len(unplaced)}")


if __name__ == "__main__":
    main()