"""Build a per-market client mapping spreadsheet for cross-checking against IT's PowerBI mapping.

Output: output/[C] Client Platform Mapping M-D-YYYY.xlsx
Sheets: Summary, CO, UT, TX (AU + SA combined per company convention), Cross-Market.
Columns: Market | Zones | Client | Status | Stub | Magazine Manager | CallRail | Uniqode | Inbox Advantage
"""

import os
import sys
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv

load_dotenv(override=True)

from supabase import create_client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
OUT_DIR = Path(__file__).resolve().parent.parent / "output"
OUT_DIR.mkdir(exist_ok=True)
DATE_TAG = datetime.now().strftime("%-m-%-d-%Y") if os.name != "nt" else datetime.now().strftime("%#m-%#d-%Y")
OUT_PATH = OUT_DIR / f"[C] Client Platform Mapping {DATE_TAG}.xlsx"

NAVY = "1F3A5F"
ORANGE = "E07A1F"
LIGHT = "F3F3F3"
THIN_BORDER = Border(
    left=Side(style="thin", color="999999"),
    right=Side(style="thin", color="999999"),
    top=Side(style="thin", color="999999"),
    bottom=Side(style="thin", color="999999"),
)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
STATUS_FILLS = {
    "active": PatternFill("solid", fgColor="C6EFCE"),
    "cancelled": PatternFill("solid", fgColor=ORANGE),
    "expired": PatternFill("solid", fgColor="FCE4D6"),
    "dormant": PatternFill("solid", fgColor="E7E6E6"),
    "prospect": PatternFill("solid", fgColor="DDEBF7"),
    "inactive": PatternFill("solid", fgColor="E7E6E6"),
}

COLUMNS = [
    ("Market", 8),
    ("Zones", 12),
    ("Client", 50),
    ("Status", 12),
    ("Stub", 7),
    ("Magazine Manager", 32),
    ("CallRail", 36),
    ("Uniqode", 22),
    ("Inbox Advantage", 32),
]


def _fetch_all(sb, table: str, columns: str = "*"):
    rows, off = [], 0
    while True:
        batch = sb.table(table).select(columns).range(off, off + 999).execute().data
        rows.extend(batch)
        if len(batch) < 1000:
            return rows
        off += 1000


def fetch_rows(sb):
    """Pull all source tables, pivot client_platform_ids in Python."""
    print("Fetching markets, zones, clients, client_zones, client_platform_ids...")
    markets = {r["id"]: r["code"] for r in sb.table("markets").select("id,code").execute().data}
    zones = {r["id"]: r["abbreviation"] for r in sb.table("zones").select("id,abbreviation").execute().data}
    clients = _fetch_all(sb, "clients", "id,name,status,is_mapping_stub,primary_market_id")
    client_zones = _fetch_all(sb, "client_zones", "client_id,zone_id")
    platform_ids = _fetch_all(sb, "client_platform_ids", "client_id,platform,external_id")

    zones_by_client: dict[str, set] = {}
    for cz in client_zones:
        z = zones.get(cz["zone_id"])
        if z:
            zones_by_client.setdefault(cz["client_id"], set()).add(z)

    pids_by_client: dict[str, dict[str, list]] = {}
    for p in platform_ids:
        bucket = pids_by_client.setdefault(p["client_id"], {})
        bucket.setdefault(p["platform"], []).append(p["external_id"])

    rows = []
    for c in clients:
        zs = sorted(zones_by_client.get(c["id"], []))
        pids = pids_by_client.get(c["id"], {})
        rows.append({
            "market": markets.get(c["primary_market_id"]),
            "name": c["name"],
            "status": c["status"],
            "is_mapping_stub": c["is_mapping_stub"],
            "zones": "; ".join(zs) if zs else None,
            "magazine_manager": "; ".join(sorted(pids.get("magazine_manager", []))) or None,
            "callrail":         "; ".join(sorted(pids.get("callrail", []))) or None,
            "uniqode":          "; ".join(sorted(pids.get("uniqode", []))) or None,
            "inbox_advantage":  "; ".join(sorted(pids.get("inbox_advantage", []))) or None,
        })

    rows.sort(key=lambda r: (r["market"] or "ZZZ", r["name"].lower()))
    return rows


def write_sheet(wb, sheet_name: str, rows: list[dict]):
    ws = wb.create_sheet(sheet_name)

    # Header
    for idx, (label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Body
    for r, row in enumerate(rows, start=2):
        values = [
            row.get("market") or "—",
            row.get("zones") or "",
            row.get("name") or "",
            row.get("status") or "",
            "Yes" if row.get("is_mapping_stub") else "",
            row.get("magazine_manager") or "",
            row.get("callrail") or "",
            row.get("uniqode") or "",
            row.get("inbox_advantage") or "",
        ]
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if c == 4:
                fill = STATUS_FILLS.get(val)
                if fill:
                    cell.fill = fill
                    cell.font = Font(bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"


def write_summary(wb, all_rows: list[dict]):
    ws = wb.create_sheet("Summary", 0)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 14

    ws.cell(row=1, column=1, value="Client Platform Mapping").font = Font(bold=True, size=16, color=NAVY)
    ws.cell(row=2, column=1, value=f"Generated {datetime.now().strftime('%B %d, %Y at %I:%M %p')}").font = Font(italic=True, color="666666")

    headers = ["Market", "Total Clients", "Real (non-stub)", "MM Mapped", "CallRail", "Uniqode", "Inbox Advantage", "Stubs"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER

    def market_label(m):
        if m in ("AU", "SA"):
            return "TX"
        return m or "Cross-Market"

    grouped: dict[str, list[dict]] = {}
    for row in all_rows:
        grouped.setdefault(market_label(row.get("market")), []).append(row)

    r = 5
    for label in ("CO", "UT", "TX", "Cross-Market"):
        rows = grouped.get(label, [])
        if not rows:
            continue
        real = [x for x in rows if not x.get("is_mapping_stub")]
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=len(rows))
        ws.cell(row=r, column=3, value=len(real))
        ws.cell(row=r, column=4, value=sum(1 for x in rows if x.get("magazine_manager")))
        ws.cell(row=r, column=5, value=sum(1 for x in rows if x.get("callrail")))
        ws.cell(row=r, column=6, value=sum(1 for x in rows if x.get("uniqode")))
        ws.cell(row=r, column=7, value=sum(1 for x in rows if x.get("inbox_advantage")))
        ws.cell(row=r, column=8, value=sum(1 for x in rows if x.get("is_mapping_stub")))
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    # Total row
    ws.cell(row=r, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=r, column=2, value=len(all_rows))
    ws.cell(row=r, column=3, value=sum(1 for x in all_rows if not x.get("is_mapping_stub")))
    ws.cell(row=r, column=4, value=sum(1 for x in all_rows if x.get("magazine_manager")))
    ws.cell(row=r, column=5, value=sum(1 for x in all_rows if x.get("callrail")))
    ws.cell(row=r, column=6, value=sum(1 for x in all_rows if x.get("uniqode")))
    ws.cell(row=r, column=7, value=sum(1 for x in all_rows if x.get("inbox_advantage")))
    ws.cell(row=r, column=8, value=sum(1 for x in all_rows if x.get("is_mapping_stub")))
    for c in range(1, 9):
        ws.cell(row=r, column=c).border = THIN_BORDER
        ws.cell(row=r, column=c).font = Font(bold=True)
    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=LIGHT)


def main():
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)
    rows = fetch_rows(sb)
    print(f"Fetched {len(rows)} client rows")

    wb = Workbook()
    wb.remove(wb.active)  # we'll add Summary as the first sheet

    # Bucket rows by display market
    co_rows = [r for r in rows if r.get("market") == "CO"]
    ut_rows = [r for r in rows if r.get("market") == "UT"]
    tx_rows = [r for r in rows if r.get("market") in ("AU", "SA")]
    cross_rows = [r for r in rows if r.get("market") not in ("CO", "UT", "AU", "SA")]

    write_summary(wb, rows)
    if co_rows:    write_sheet(wb, "CO", co_rows)
    if ut_rows:    write_sheet(wb, "UT", ut_rows)
    if tx_rows:    write_sheet(wb, "TX (AU + SA)", tx_rows)
    if cross_rows: write_sheet(wb, "Cross-Market", cross_rows)

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"  CO: {len(co_rows)}  UT: {len(ut_rows)}  TX: {len(tx_rows)}  Cross: {len(cross_rows)}")


if __name__ == "__main__":
    main()
