"""Audit MagManager Category strings against our category_aliases map.

Read-only. Pulls every contact across all tenants, extracts the Category JSON
array, counts distinct values, diffs against existing category_aliases.

Output:
  - output/[C] MM Category Alias Audit M-D-YYYY.xlsx
    - Sheet 1 (Summary): top-line counts
    - Sheet 2 (Mapped): MM strings we already have an alias for
    - Sheet 3 (UNMAPPED): MM strings we need to add — most important sheet

Idempotent: read-only, no DB writes.

Usage:
  python scripts/audit_mm_category_aliases.py
"""

from __future__ import annotations

import json
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
from supabase import create_client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from etl.magmanager_client import MagManagerClient

load_dotenv()

DATABASES = [
    "thehomemagcolorado",
    "thehomemagutah",
    "thehomemagsanantonio",
]

OUT_DIR = Path(__file__).resolve().parent.parent / "output"
OUT_DIR.mkdir(exist_ok=True)
DATE_TAG = (
    datetime.now().strftime("%#m-%#d-%Y")
    if os.name == "nt"
    else datetime.now().strftime("%-m-%-d-%Y")
)
OUT_PATH = OUT_DIR / f"[C] MM Category Alias Audit {DATE_TAG}.xlsx"

NAVY = "1F3A5F"
LIGHT_BLUE = "DCE6F1"
LIGHT = "F3F3F3"
THIN = Border(*(Side(style="thin", color="999999"),) * 4)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)


def extract_categories(json_str: str | None) -> list[str]:
    """MM returns Category as '[{"category":"Furniture"}]'. Parse it."""
    if not json_str or json_str.strip() in ("", "[]"):
        return []
    try:
        arr = json.loads(json_str)
        return [
            item.get("category")
            for item in arr
            if isinstance(item, dict) and item.get("category")
        ]
    except (json.JSONDecodeError, TypeError):
        return []


def main():
    print("=" * 70)
    print("MM CATEGORY ALIAS AUDIT (read-only)")
    print("=" * 70)

    mm = MagManagerClient()
    sb = create_client(os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_KEY"))

    # ----------------------------------------------------------------
    # 1) Pull existing alias map (alias_lower -> category_id)
    # ----------------------------------------------------------------
    print("\nLoading existing category_aliases...")
    aliases = sb.table("category_aliases").select("alias,category_id,source").execute().data
    alias_map = {a["alias"].strip().lower(): a for a in aliases}
    print(f"  {len(alias_map):,} existing aliases")

    # Pull category id -> name for display
    cats = sb.table("categories").select("id,name,slug,level,parent_id").execute().data
    cat_by_id = {c["id"]: c for c in cats}

    def cat_path(cat_id):
        """Return 'Group > Category > Sub' style path."""
        if not cat_id or cat_id not in cat_by_id:
            return None
        node = cat_by_id[cat_id]
        parts = [node["name"]]
        while node.get("parent_id"):
            parent = cat_by_id.get(node["parent_id"])
            if not parent:
                break
            parts.append(parent["name"])
            node = parent
        return " > ".join(reversed(parts))

    # ----------------------------------------------------------------
    # 2) Pull every contact, extract Category values, count
    # ----------------------------------------------------------------
    print("\nPulling all contacts from MM API (3 tenants)...")
    category_counts: Counter[str] = Counter()
    category_by_db: dict[str, Counter] = defaultdict(Counter)
    blank_or_uncategorized = 0
    total_contacts = 0
    multi_tag_contacts = 0

    for db in DATABASES:
        page = 1
        while True:
            body = mm.get_contacts_page(page=page, database_name=db)
            rows = body.get("Data") or []
            for r in rows:
                total_contacts += 1
                cats_for_row = extract_categories(r.get("Category"))
                if not cats_for_row:
                    blank_or_uncategorized += 1
                    continue
                if len(cats_for_row) > 1:
                    multi_tag_contacts += 1
                for c in cats_for_row:
                    category_counts[c] += 1
                    category_by_db[db][c] += 1
            print(f"  {db} page {page}: {len(rows)} rows (running total: {total_contacts:,})")
            if len(rows) < 10000:
                break
            page += 1
            if page > 10:
                break

    print(f"\nDistinct MM Category values: {len(category_counts):,}")
    print(f"Total contacts: {total_contacts:,}")
    print(f"  with at least 1 category: {total_contacts - blank_or_uncategorized:,}")
    print(f"  blank/uncategorized: {blank_or_uncategorized:,}")
    print(f"  multi-tag contacts: {multi_tag_contacts:,}")

    # ----------------------------------------------------------------
    # 3) Diff against alias map
    # ----------------------------------------------------------------
    mapped_rows = []
    unmapped_rows = []
    for value, n in category_counts.most_common():
        lc = value.strip().lower()
        a = alias_map.get(lc)
        per_db = {db: category_by_db[db].get(value, 0) for db in DATABASES}
        if a:
            mapped_rows.append({
                "mm_category": value,
                "n_contacts": n,
                "co": per_db["thehomemagcolorado"],
                "ut": per_db["thehomemagutah"],
                "sa": per_db["thehomemagsanantonio"],
                "target_category": cat_path(a["category_id"]) or "(unknown)",
                "alias_source": a.get("source", ""),
            })
        else:
            unmapped_rows.append({
                "mm_category": value,
                "n_contacts": n,
                "co": per_db["thehomemagcolorado"],
                "ut": per_db["thehomemagutah"],
                "sa": per_db["thehomemagsanantonio"],
            })

    print(f"\nMapped: {len(mapped_rows):,} (covering "
          f"{sum(r['n_contacts'] for r in mapped_rows):,} contact-tags)")
    print(f"UNMAPPED: {len(unmapped_rows):,} (covering "
          f"{sum(r['n_contacts'] for r in unmapped_rows):,} contact-tags) — must add before sync")

    # ----------------------------------------------------------------
    # 4) Build workbook
    # ----------------------------------------------------------------
    wb = Workbook()

    # Summary sheet
    s = wb.active
    s.title = "Summary"
    s.column_dimensions["A"].width = 40
    s.column_dimensions["B"].width = 18

    summary_rows = [
        ("MM Category Alias Audit", Font(bold=True, size=18, color=NAVY)),
        (f"Generated {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
         Font(italic=True, color="666666")),
        ("", None),
        ("Total MM contacts", total_contacts),
        ("  with at least 1 category", total_contacts - blank_or_uncategorized),
        ("  blank/uncategorized", blank_or_uncategorized),
        ("  multi-tag contacts", multi_tag_contacts),
        ("", None),
        ("Distinct MM Category values", len(category_counts)),
        ("Already mapped via category_aliases", len(mapped_rows)),
        ("UNMAPPED (need to add)", len(unmapped_rows)),
        ("", None),
        ("Mapped contact-tags covered", sum(r['n_contacts'] for r in mapped_rows)),
        ("Unmapped contact-tags (uncovered)", sum(r['n_contacts'] for r in unmapped_rows)),
        ("", None),
        ("Existing aliases in DB", len(alias_map)),
        ("Categories defined in DB", len(cat_by_id)),
    ]
    for r_idx, (label, val) in enumerate(summary_rows, start=1):
        s.cell(row=r_idx, column=1, value=label)
        if isinstance(val, Font):
            s.cell(row=r_idx, column=1).font = val
        else:
            s.cell(row=r_idx, column=2, value=val)
            if r_idx > 3:
                s.cell(row=r_idx, column=1).font = Font(bold=True)

    # UNMAPPED sheet (most important — put it second so reviewer sees it after summary)
    u = wb.create_sheet("UNMAPPED")
    headers = ["MM Category", "Total #", "CO", "UT", "SA", "Suggested Target", "Notes"]
    widths = [40, 9, 7, 7, 7, 35, 30]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = u.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = THIN
        u.column_dimensions[get_column_letter(i)].width = w

    for r_idx, row in enumerate(unmapped_rows, start=2):
        vals = [
            row["mm_category"], row["n_contacts"],
            row["co"], row["ut"], row["sa"],
            "", "",
        ]
        for c_idx, v in enumerate(vals, start=1):
            cell = u.cell(row=r_idx, column=c_idx, value=v)
            cell.border = THIN
            if c_idx == 1:
                cell.font = Font(bold=True)
    u.freeze_panes = "A2"
    u.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(unmapped_rows) + 1}"

    # MAPPED sheet (full reference)
    m = wb.create_sheet("Mapped")
    headers_m = ["MM Category", "Total #", "CO", "UT", "SA", "→ Target Category", "Alias Source"]
    widths_m = [35, 9, 7, 7, 7, 40, 18]
    for i, (h, w) in enumerate(zip(headers_m, widths_m), start=1):
        c = m.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = THIN
        m.column_dimensions[get_column_letter(i)].width = w

    for r_idx, row in enumerate(mapped_rows, start=2):
        vals = [
            row["mm_category"], row["n_contacts"],
            row["co"], row["ut"], row["sa"],
            row["target_category"], row["alias_source"],
        ]
        for c_idx, v in enumerate(vals, start=1):
            cell = m.cell(row=r_idx, column=c_idx, value=v)
            cell.border = THIN
            if c_idx == 1:
                cell.font = Font(bold=True)
    m.freeze_panes = "A2"
    m.auto_filter.ref = f"A1:{get_column_letter(len(headers_m))}{len(mapped_rows) + 1}"

    wb.save(OUT_PATH)
    print(f"\nWrote {OUT_PATH}")
    print(f"\nNext step:")
    print(f"  1. Open UNMAPPED sheet, fill in 'Suggested Target' for high-volume rows")
    print(f"  2. Use setup/import_category_approvals.py pattern to bulk-add aliases")
    print(f"  3. Then run the contacts ETL")


if __name__ == "__main__":
    main()
