"""
Import THM Ad Placement spreadsheets (CO + UT) into Supabase.

Parses filenames into structured data:
  - Client name (with fuzzy matching against clients table)
  - Ad size (Full/Half/Quarter/Front Cover/etc.)
  - Issue year/month/spring flag
  - Zone

Handles multiple historical filename formats and the "DN" zone code
(which means an ad ran in both ND and SD - creates two rows).

Usage:
  python setup/import_ad_placements.py --dry-run
  python setup/import_ad_placements.py
"""

import sys
import os
import re
import argparse
from pathlib import Path
from difflib import SequenceMatcher

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
load_dotenv()

from supabase import create_client
from openpyxl import load_workbook

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

DATA_DIR = Path(__file__).resolve().parent.parent / "data"


def detect_placement_files():
    """Auto-detect latest CO and UT ad placement spreadsheets."""
    files = []
    for label, pattern in [("CO", "THM Colorado Ad Placement*"), ("UT", "THM Utah Ad Placement*")]:
        candidates = sorted(DATA_DIR.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
        if candidates:
            files.append((candidates[0], label))
    return files


DEFAULT_FILES = [
    (DATA_DIR / "THM Colorado Ad Placement (1).xlsx", "CO"),
    (DATA_DIR / "THM Utah Ad Placement.xlsx", "UT"),
]

# Spreadsheet zone codes -> our zone abbreviations
# DN means "both Denver zones" - we'll create 2 rows
ZONE_MAP = {
    "EPC": ["EPC"],
    "NCO": ["NOCO"],
    "NDN": ["ND"],
    "SDN": ["SD"],
    "DN":  ["ND", "SD"],   # both denver zones
    "CW":  ["CW"],
    "NW":  ["NW"],
    "SW":  ["SW"],
}

# Ad size code -> human readable name
# Note: FB (Full Bleed) and DB (Double Bleed) are designer/proofing syntax for the
# same products as F (Full Page) and D (Double Spread). We collapse them.
SIZE_NAMES = {
    "F":  "Full Page",
    "H":  "Half Page",
    "Q":  "Quarter Page",
    "D":  "Double Spread",
    "FC": "Front Cover",
    "BC": "Back Cover",
    "BB": "Back Cover Banner",
    "FB": "Full Page",        # collapse Full Bleed -> Full Page
    "DB": "Double Spread",    # collapse Double Bleed -> Double Spread
    "HV": "Half Vertical",
}

# When inserting, normalize the code so FB stores as F, DB as D
SIZE_CODE_NORMALIZE = {
    "FB": "F",
    "DB": "D",
}

# Known zone codes and size codes - used to identify which is which
KNOWN_ZONES = {"EPC", "NCO", "NDN", "SDN", "DN", "CW", "NW", "SW"}
KNOWN_SIZES = {"F", "H", "Q", "D", "FC", "BC", "BB", "FB", "DB", "HV"}

# Filename patterns - try in order, first match wins
# Each pattern captures: (client_name, size_code)
# Treat space and underscore as equivalents to hyphen for parsing - we replace them up front
FILENAME_PATTERNS = [
    # New format: "ClientName-THMCO-F-EPC-2403.pdf" / "ClientName-THMUT-H-CW-2401.pdf"
    re.compile(r'^(.+?)-THM(?:CO|UT)-([A-Z]{1,2})-[A-Z]+-\d{4}', re.IGNORECASE),
    # New format with THM CO (space) or just THM: "ClientName-THM CO-F-NDN-2603.pdf" / "ClientName-THM-Q-NCO-2403.pdf"
    re.compile(r'^(.+?)-THM(?:\s*(?:CO|UT))?-([A-Z]{1,2})-[A-Z]+-\d{4}', re.IGNORECASE),
    # New format without THMCO at all: "GregUnsethPainting&Exteriors-H-EPC-2403.pdf"
    re.compile(r'^(.+?)-([A-Z]{1,2})-[A-Z]{2,4}-\d{4}[a-z]?\.(?:pdf|jpg|jpeg|png)', re.IGNORECASE),
    # Doubled zone: "PaversNTurf-THMCO-F-NCO-NCO-2503.pdf"
    re.compile(r'^(.+?)-THM(?:CO|UT)-([A-Z]{1,2})-[A-Z]+-[A-Z]+-\d{4}', re.IGNORECASE),
    # Old format with -F.jpg suffix: "ClientName-EPC-THM24-01-F.jpg"
    re.compile(r'^(.+?)-[A-Z]+-THM\d{2,}-\d{2}[a-z]?-([A-Z]{1,2})\.', re.IGNORECASE),
    # Old format with date in middle: "ClientName-Plumbing-SDN-24-01-H.jpg"
    re.compile(r'^(.+?)-[A-Z]+-\d{2}-\d{2}-([A-Z]{1,2})\.', re.IGNORECASE),
    # Trailing format: "Precision Closets & Garage-FC-NW-2511.pdf" (no THM marker)
    re.compile(r'^(.+?)-([A-Z]{1,2})-[A-Z]{2,4}-\d{4}[a-z]?', re.IGNORECASE),
    # FenceRevolution-SDN-THM22-04-Q-SDN-2407.pdf - extra zone in middle
    re.compile(r'^(.+?)-[A-Z]+-THM\d{2,}-\d{2}-([A-Z]{1,2})-[A-Z]+-\d{4}', re.IGNORECASE),
]


def parse_filename_token_based(filename):
    """
    Token-based fallback parser. Splits the filename and looks for:
      - A known size code anywhere in the segments
      - A known zone code anywhere in the segments
    Anything before the THM marker is the client name.
    """
    if not filename:
        return None, None
    base = re.sub(r'\.(pdf|jpg|jpeg|png)$', '', filename, flags=re.IGNORECASE)
    tokens = base.split("-")

    size_code = None
    client_tokens = []

    for tok in tokens:
        tok_clean = tok.strip().upper()
        # Stop at THM marker - everything before is client name
        if tok_clean.startswith("THM"):
            break
        client_tokens.append(tok)

    # Look for size code in remaining tokens (after THM)
    found_thm = False
    for tok in tokens:
        if tok.upper().startswith("THM"):
            found_thm = True
            continue
        if found_thm and tok.upper() in KNOWN_SIZES:
            size_code = tok.upper()
            break

    # If no THM marker, search whole filename for size code (in token form)
    if not size_code:
        for tok in tokens:
            if tok.upper() in KNOWN_SIZES:
                size_code = tok.upper()
                # Client name is everything before this token
                idx = tokens.index(tok)
                if idx > 0:
                    client_tokens = tokens[:idx]
                break

    if not client_tokens or not size_code:
        return None, None

    raw_client = "-".join(client_tokens)
    clean = split_camel_case(raw_client).replace("_", " ")
    return clean, size_code


def preprocess_filename(filename):
    """Normalize quirky filenames before pattern matching."""
    if not filename:
        return filename
    s = filename
    # Skip rows that don't look like filenames at all (no extension, just text)
    if "." not in s and "-" not in s:
        return None
    # Replace underscores with hyphens (e.g., S_SRoofing, H_SW)
    s = s.replace("_", "-")
    # Replace " THMCO-" or " THMUT-" with "-THMCO-" (some filenames have a space)
    s = re.sub(r'\s+(THM(?:CO|UT))', r'-\1', s)
    return s


def normalize(name):
    if not name:
        return ""
    n = name.lower().strip()
    # Strip suffixes only at word boundaries (not in the middle of words)
    n = re.sub(r'\b(llc|inc|co)\b\.?', '', n)
    # Replace punctuation with spaces
    for s in ["-", "_", ".", "&", "'", "/", ",", "(", ")"]:
        n = n.replace(s, " ")
    return " ".join(n.split())  # collapse whitespace


def alphanumeric_only(name):
    """Strip everything but letters and digits, lowercase. For matching condensed names."""
    if not name:
        return ""
    return "".join(c for c in name.lower() if c.isalnum())


def similarity(a, b):
    return SequenceMatcher(None, normalize(a), normalize(b)).ratio()


def split_camel_case(s):
    """Insert spaces before uppercase letters: 'ApexCleanAir' -> 'Apex Clean Air'"""
    # Add space before uppercase letters that follow lowercase
    s = re.sub(r'([a-z])([A-Z])', r'\1 \2', s)
    # Handle sequences like "RGSExteriors" -> "RGS Exteriors"
    s = re.sub(r'([A-Z]+)([A-Z][a-z])', r'\1 \2', s)
    # Add space between digit and letter: "3DayBlinds" -> "3 Day Blinds", "888HeatingAir" -> "888 Heating Air"
    s = re.sub(r'(\d)([A-Za-z])', r'\1 \2', s)
    s = re.sub(r'([A-Za-z])(\d)', r'\1 \2', s)
    return s


def parse_filename(filename):
    """Returns (client_name_clean, size_code) or (None, None) if no match."""
    pre = preprocess_filename(filename)
    if not pre:
        return None, None

    for pattern in FILENAME_PATTERNS:
        m = pattern.match(pre)
        if m:
            raw_name = m.group(1)
            size = m.group(2).upper()
            # Validate size code is real
            if size not in SIZE_NAMES:
                continue
            # Clean the client name
            clean = split_camel_case(raw_name)
            clean = clean.replace("_", " ")
            return clean, size

    # Fallback: token-based parser
    return parse_filename_token_based(pre)


def parse_issue_code(issue_raw):
    """Parses '2401' or '2404s' -> (year, month, is_spring)."""
    if not issue_raw:
        return None, None, False
    s = str(issue_raw).strip()
    is_spring = s.lower().endswith("s")
    if is_spring:
        s = s[:-1]
    if len(s) != 4 or not s.isdigit():
        return None, None, False
    year = 2000 + int(s[:2])
    month = int(s[2:])
    if month < 1 or month > 12:
        return None, None, False
    return year, month, is_spring


def run(dry_run=True):
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Load lookups
    print("Loading lookups...")
    zones_result = sb.table("zones").select("id,abbreviation,market_id").execute()
    zone_by_abbrev = {z["abbreviation"]: z for z in zones_result.data if z.get("abbreviation")}
    print(f"  {len(zone_by_abbrev)} zones loaded")

    print("  Loading clients...")
    all_clients = []
    offset = 0
    while True:
        batch = sb.table("clients").select("id,name,primary_market_id").range(offset, offset + 999).execute()
        all_clients.extend(batch.data)
        if len(batch.data) < 1000:
            break
        offset += 1000
    print(f"  {len(all_clients)} clients loaded")

    # Build name lookups (multiple keys per client for flexible matching)
    clients_by_norm = {}
    clients_by_alnum = {}  # alphanumeric-only key for camelcase/condensed matching
    for c in all_clients:
        clients_by_norm.setdefault(normalize(c["name"]), []).append(c)
        clients_by_alnum.setdefault(alphanumeric_only(c["name"]), []).append(c)

    # Markets
    markets_result = sb.table("markets").select("id,code").execute()
    market_by_code = {m["code"]: m for m in markets_result.data}

    # Cache fuzzy matches
    fuzzy_cache = {}

    def lookup_client(name, primary_market_id):
        if not name:
            return None
        norm = normalize(name)
        alnum = alphanumeric_only(name)

        # Cache hit
        if norm in fuzzy_cache:
            candidates = fuzzy_cache[norm]
        else:
            candidates = []

        # 1. Exact normalized match
        if not candidates:
            candidates = clients_by_norm.get(norm, [])

        # 2. Alphanumeric exact match (handles camelcase: 3DayBlinds -> "3 Day Blinds c/o ...")
        if not candidates:
            candidates = clients_by_alnum.get(alnum, [])

        # 3. Alphanumeric prefix match (handles "3DayBlinds" -> "3 Day Blinds c/o Incremental Media")
        if not candidates and len(alnum) >= 6:
            for db_alnum, db_clients in clients_by_alnum.items():
                if db_alnum.startswith(alnum) and len(alnum) >= 0.5 * len(db_alnum):
                    candidates = db_clients
                    break
                if alnum.startswith(db_alnum) and len(db_alnum) >= 6:
                    candidates = db_clients
                    break

        # 4. Word-based prefix match (normalized form)
        if not candidates:
            for c in all_clients:
                db_norm = normalize(c["name"])
                if db_norm.startswith(norm + " ") and len(norm) >= 6:
                    candidates = [c]
                    break
                if norm.startswith(db_norm + " ") and len(db_norm) >= 6:
                    candidates = [c]
                    break

        # 5. Fuzzy fallback
        if not candidates:
            best_score = 0
            best_match = None
            for c in all_clients:
                score = similarity(name, c["name"])
                if score > best_score:
                    best_score = score
                    best_match = c
            if best_score >= 0.85 and best_match:
                candidates = [best_match]

        if candidates:
            fuzzy_cache[norm] = candidates

        if not candidates:
            return None
        if len(candidates) == 1:
            return candidates[0]["id"]
        # Multi-market: prefer the matching market
        if primary_market_id:
            for c in candidates:
                if c["primary_market_id"] == primary_market_id:
                    return c["id"]
        return candidates[0]["id"]

    rows_to_insert = []
    skipped_no_filename = 0
    skipped_no_zone = 0
    failed_to_parse = []
    unmatched_clients = {}

    for path, market_code in FILES:
        if not path.exists():
            print(f"\nWARNING: {path} not found, skipping")
            continue

        market = market_by_code.get(market_code)
        market_id = market["id"] if market else None

        print(f"\nReading {path.name}...")
        wb = load_workbook(str(path), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        # Headers are on row 2 (row 1 is the title)
        for row in ws.iter_rows(min_row=3, values_only=True):
            issue_raw = row[0]
            zone_raw = row[1]
            filename = row[2]
            page = row[3]

            if not filename:
                skipped_no_filename += 1
                continue

            # Parse issue
            year, month, is_spring = parse_issue_code(issue_raw)
            if year is None:
                continue
            issue_code = str(issue_raw).strip()

            # Parse zone
            zone_key = (zone_raw or "").strip().upper()
            zone_abbrevs = ZONE_MAP.get(zone_key, [])
            if not zone_abbrevs:
                skipped_no_zone += 1
                continue

            # Parse filename for client + size
            client_name_clean, size_code = parse_filename(filename)
            if not client_name_clean:
                failed_to_parse.append(filename)
                # Still create the row, just without client/size match
                client_id = None
                ad_size = None
            else:
                client_id = lookup_client(client_name_clean, market_id)
                if not client_id:
                    unmatched_clients[client_name_clean] = unmatched_clients.get(client_name_clean, 0) + 1
                ad_size = SIZE_NAMES.get(size_code)
                # Normalize FB->F and DB->D
                size_code = SIZE_CODE_NORMALIZE.get(size_code, size_code)

            # Create one row per zone (DN expands to 2 zones)
            for zone_abbrev in zone_abbrevs:
                zone = zone_by_abbrev.get(zone_abbrev)
                if not zone:
                    skipped_no_zone += 1
                    continue
                rows_to_insert.append({
                    "client_id": client_id,
                    "zone_id": zone["id"],
                    "market_id": market_id,
                    "issue_code": issue_code,
                    "issue_year": year,
                    "issue_month": month,
                    "is_spring": is_spring,
                    "page": int(page) if page else None,
                    "ad_size": ad_size,
                    "ad_size_code": size_code,
                    "filename": filename,
                    "source_client_name": client_name_clean,
                    "source_zone": zone_raw,
                })

        wb.close()

    # Reporting
    matched = sum(1 for r in rows_to_insert if r["client_id"])
    sized = sum(1 for r in rows_to_insert if r["ad_size_code"])

    print(f"\n{'='*60}")
    print(f"  IMPORT PLAN")
    print(f"{'='*60}")
    print(f"  Total rows to insert:    {len(rows_to_insert)}")
    print(f"  Client matched:          {matched}")
    print(f"  Client unmatched:        {len(rows_to_insert) - matched}")
    print(f"  Filename parsed (size):  {sized}")
    print(f"  Filename parse failed:   {len(failed_to_parse)}")
    print(f"  Skipped (no filename):   {skipped_no_filename}")
    print(f"  Skipped (no zone):       {skipped_no_zone}")
    print(f"{'='*60}")

    if unmatched_clients:
        print(f"\n  Top 25 unmatched client names ({len(unmatched_clients)} total):")
        for n, c in sorted(unmatched_clients.items(), key=lambda x: -x[1])[:25]:
            print(f"    '{n}': {c} placements")

    if failed_to_parse:
        print(f"\n  Failed-to-parse filenames ({len(failed_to_parse)} total):")
        for f in failed_to_parse[:15]:
            print(f"    {f}")

    if dry_run:
        print("\n  DRY RUN - no data written")
        return

    # Insert in batches with conflict handling
    print(f"\nInserting {len(rows_to_insert)} placements...")
    BATCH = 200
    inserted = 0
    for i in range(0, len(rows_to_insert), BATCH):
        batch = rows_to_insert[i:i + BATCH]
        try:
            sb.table("ad_placements").upsert(
                batch, on_conflict="filename,zone_id,issue_code"
            ).execute()
            inserted += len(batch)
        except Exception as e:
            print(f"  Batch failed at {i}: {str(e)[:200]}")
            # Try individually
            for rec in batch:
                try:
                    sb.table("ad_placements").upsert(
                        rec, on_conflict="filename,zone_id,issue_code"
                    ).execute()
                    inserted += 1
                except Exception:
                    pass
        if inserted % 2000 == 0 and inserted > 0:
            print(f"  ... {inserted} inserted")

    print(f"\n{'='*60}")
    print(f"  COMPLETE - {inserted} placements inserted")
    print(f"{'='*60}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--co-file", type=str, help="Path to CO Ad Placement Excel file")
    parser.add_argument("--ut-file", type=str, help="Path to UT Ad Placement Excel file")
    args = parser.parse_args()

    global FILES
    if args.co_file or args.ut_file:
        FILES = list(DEFAULT_FILES)
        if args.co_file:
            FILES = [(f, m) if m != "CO" else (Path(args.co_file), "CO") for f, m in FILES]
        if args.ut_file:
            FILES = [(f, m) if m != "UT" else (Path(args.ut_file), "UT") for f, m in FILES]
    else:
        FILES = detect_placement_files()
        if not FILES:
            FILES = list(DEFAULT_FILES)

    if not all([SUPABASE_URL, SUPABASE_KEY]):
        print("ERROR: Missing env vars")
        sys.exit(1)

    run(dry_run=args.dry_run)


if __name__ == "__main__":
    main()
