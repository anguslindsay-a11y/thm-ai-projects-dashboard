"""
Import Priority 4-9 spreadsheet into Supabase.

Updates existing clients and creates new ones with:
  - priority, category, sales_attrib, mm_start_issue
  - sales rep and zone links

Usage:
  python setup/import_priority.py --dry-run   # Preview only
  python setup/import_priority.py              # Run the import
"""

import sys
import os
import argparse
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
load_dotenv()

from supabase import create_client
from openpyxl import load_workbook

EXCEL_PATH = Path(__file__).resolve().parent.parent / "data" / "Priority 4-9.xlsx"

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

# Mkt column values -> market codes in Supabase
MKT_TO_MARKET_CODE = {
    "CO": "CO",
    "UT": "UT",
    "TX": "AU",  # TX rows default to Austin — could be Austin or San Antonio
}


def read_spreadsheet():
    wb = load_workbook(str(EXCEL_PATH), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data = []
    for row in rows[1:]:
        data.append(dict(zip(headers, row)))
    wb.close()
    return data


def parse_date(val):
    """Try to parse Start Issue into a date string (YYYY-MM-DD)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    # Try common formats
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def run_import(rows, dry_run=True):
    total = len(rows)
    unique_companies = set(str(r["Company Name"]).strip() for r in rows if r.get("Company Name"))
    unique_reps = set(
        str(r["Rep Name"]).strip()
        for r in rows
        if r.get("Rep Name") and str(r["Rep Name"]).strip()
    )

    print(f"\n{'='*50}")
    print(f"  PRIORITY SPREADSHEET ANALYSIS")
    print(f"{'='*50}")
    print(f"  Total rows:          {total}")
    print(f"  Unique companies:    {len(unique_companies)}")
    print(f"  Unique reps:         {len(unique_reps)}")
    print(f"{'='*50}\n")

    if dry_run:
        print("  DRY RUN — no data will be written.\n")
        print(f"  Would update/create {len(unique_companies)} clients")
        print(f"  Would create any missing sales reps from {len(unique_reps)} unique names")
        print()
        return

    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Load markets
    markets_result = sb.table("markets").select("*").execute()
    market_by_code = {m["code"]: m for m in markets_result.data}

    # Load existing clients
    print("Loading existing clients...")
    all_clients = []
    offset = 0
    while True:
        batch = sb.table("clients").select("id,name").range(offset, offset + 999).execute()
        all_clients.extend(batch.data)
        if len(batch.data) < 1000:
            break
        offset += 1000
    client_lookup = {c["name"]: c["id"] for c in all_clients}
    print(f"  {len(client_lookup)} clients in DB")

    # Load existing reps
    existing_reps = sb.table("sales_reps").select("id,name").execute()
    rep_lookup = {r["name"]: r["id"] for r in existing_reps.data}

    # Process each row — group by company to get one record per client
    # (a company may appear multiple times if it's in multiple markets)
    # We take the first occurrence's data for priority/category/attrib
    client_data = {}  # company_name -> {priority, category, sales_attrib, mm_start_issue, rep, mkt}
    for r in rows:
        name = str(r.get("Company Name", "")).strip()
        if not name:
            continue
        if name not in client_data:
            client_data[name] = {
                "priority": str(r.get("Priority", "")).strip() if r.get("Priority") else None,
                "category": str(r.get("Category", "")).strip() if r.get("Category") else None,
                "sales_attrib": str(r.get("Sales Attrib", "")).strip() if r.get("Sales Attrib") else None,
                "mm_start_issue": parse_date(r.get("Start Issue")),
                "rep_name": str(r.get("Rep Name", "")).strip() if r.get("Rep Name") else None,
                "mkt": str(r.get("Mkt", "")).strip() if r.get("Mkt") else None,
            }

    # Create missing reps
    print("\nCreating missing sales reps...")
    reps_created = 0
    for name, data in client_data.items():
        rep_name = data.get("rep_name")
        if rep_name and rep_name not in rep_lookup and not rep_name.startswith("*"):
            result = sb.table("sales_reps").insert({"name": rep_name}).execute()
            rep_lookup[rep_name] = result.data[0]["id"]
            reps_created += 1
    print(f"  Created {reps_created} new reps")

    # Update/create clients
    print("\nUpdating/creating clients...")
    updated = 0
    created = 0
    for name, data in client_data.items():
        update_fields = {}
        if data["priority"]:
            update_fields["priority"] = data["priority"]
        if data["category"]:
            update_fields["category"] = data["category"]
        if data["sales_attrib"]:
            update_fields["sales_attrib"] = data["sales_attrib"]
        if data["mm_start_issue"]:
            update_fields["mm_start_issue"] = data["mm_start_issue"]

        # Resolve rep
        rep_name = data.get("rep_name")
        if rep_name and not rep_name.startswith("*") and rep_name in rep_lookup:
            update_fields["sales_rep_id"] = rep_lookup[rep_name]

        if name in client_lookup:
            # Update existing client
            if update_fields:
                sb.table("clients").update(update_fields).eq("id", client_lookup[name]).execute()
                updated += 1
        else:
            # Create new client
            update_fields["name"] = name
            # Resolve market for primary_market_id
            mkt = data.get("mkt")
            if mkt and mkt in MKT_TO_MARKET_CODE:
                market_code = MKT_TO_MARKET_CODE[mkt]
                if market_code in market_by_code:
                    update_fields["primary_market_id"] = market_by_code[market_code]["id"]

            result = sb.table("clients").insert(update_fields).execute()
            client_lookup[name] = result.data[0]["id"]
            created += 1

        if (updated + created) % 200 == 0 and (updated + created) > 0:
            print(f"  ... {updated} updated, {created} created")

    print(f"\n{'='*50}")
    print(f"  PRIORITY IMPORT COMPLETE")
    print(f"{'='*50}")
    print(f"  Reps created:        {reps_created}")
    print(f"  Clients updated:     {updated}")
    print(f"  Clients created:     {created}")
    print(f"  Total in DB:         {len(client_lookup)}")
    print(f"{'='*50}\n")


def main():
    parser = argparse.ArgumentParser(description="Import Priority spreadsheet into Supabase")
    parser.add_argument("--dry-run", action="store_true", help="Preview only")
    args = parser.parse_args()

    if not SUPABASE_URL or not SUPABASE_KEY:
        print("ERROR: SUPABASE_URL and SUPABASE_KEY must be set in .env")
        sys.exit(1)
    if not EXCEL_PATH.exists():
        print(f"ERROR: Spreadsheet not found at {EXCEL_PATH}")
        sys.exit(1)

    print(f"Reading spreadsheet: {EXCEL_PATH.name}")
    rows = read_spreadsheet()
    run_import(rows, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
