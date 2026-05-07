"""
Import Monday Runsheet data into Supabase.

Reads the exported Monday Runsheets Excel file and populates the
runsheet_entries table. Each row = one client in one issue.

Also updates client metadata: call_tracking status, categories.

Idempotent: deletes existing entries for each issue_code before inserting,
so re-importing an updated file replaces stale data.

Usage:
  python setup/import_runsheets.py --dry-run
  python setup/import_runsheets.py
  python setup/import_runsheets.py --file "data/Monday Runsheets for Supabase.xlsx"
"""

import sys
import os
import re
import argparse
from pathlib import Path
from datetime import datetime
from difflib import SequenceMatcher

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
load_dotenv()

from supabase import create_client
from openpyxl import load_workbook

DEFAULT_EXCEL = Path(__file__).resolve().parent.parent / "data" / "Monday Runsheets for Supabase.xlsx"

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

# Issue code pattern: XX-Mon-YY (e.g., TX-May-26, CO-Apr-26)
ISSUE_CODE_RE = re.compile(r'^([A-Z]{2})-([A-Za-z]+)-(\d{2})$')

# Market code from issue prefix
ISSUE_PREFIX_TO_MARKET = {
    "CO": "CO",
    "UT": "UT",
    "TX": None,  # TX splits into AU + SA based on zone
}

# Skip groups with more than this many rows (likely combined multi-month sheets)
MAX_GROUP_SIZE = 500

# Zone name -> abbreviation (strip suffixes like IMP, OPP, PRES, DOM first)
ZONE_SUFFIX_RE = re.compile(r'\s+(IMP|OPP|PRES|DOM)$', re.IGNORECASE)

ZONE_MAP = {
    "sa west": "SAW",
    "sa east": "SAE",
    "au north": "AN",
    "au south": "AS",
    "sctrl": "CW",
    "snorth": "NW",
    "ssouth": "SW",
    "north denver": "ND",
    "south denver": "SD",
    "epc": "EPC",
    "noco": "NOCO",
    "nw": "NW",
    "sw": "SW",
    "cw": "CW",
}

# Ad size name -> code
AD_SIZE_MAP = {
    "full page": "F",
    "full page dirspot": "F",
    "1/2 page": "H",
    "1/2 page vertical": "HV",
    "1/4 page": "Q",
    "double page": "D",
    "double page certfeat": "D",
    "front cover": "FC",
    "back cover banner": "BB",
    "back cover 2/3 page": "BC",
    "spotlight": "SPOT",
    "certified directory": "CD",
    "ia exc 1": "IA_EXC",
    "ia exc 4": "IA_EXC",
    "ia spon 1": "IA_SPON",
    "opp bookmark": "OPP_BM",
    "opp popout": "OPP_PO",
}

# Column indexes
COL = {
    "name": 0,
    "designer": 1,
    "client_strategy": 2,
    "design_contact": 3,
    "zone": 4,
    "past_ad_size": 5,
    "current_ad_size": 6,
    "call_tracking": 7,
    "flip_notes": 8,
    "v1": 9,
    "v2": 10,
    "previous_page": 11,
    "page_counter": 12,
    "other_sizes": 13,
    "sales_rep": 14,
    "all_categories": 15,
    "primary_category": 16,
    "category_notes": 17,
    "position": 18,
    "group": 19,
    "v3": 20,
    "general_notes": 21,
    "in_home_date": 22,
    "deadline_date": 23,
    "ad_colors": 24,
    "state": 25,
}


def normalize(name):
    if not name:
        return ""
    n = str(name).lower().strip()
    n = re.sub(r'\b(llc|inc|co\.?)\b', '', n)
    return n.strip()


def similarity(a, b):
    return SequenceMatcher(None, normalize(a), normalize(b)).ratio()


def to_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def to_int(val):
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def parse_zone(zone_raw):
    """Parse zone string, stripping bundle suffixes. Returns abbreviation or None."""
    if not zone_raw:
        return None
    z = str(zone_raw).strip()
    z = ZONE_SUFFIX_RE.sub('', z)
    return ZONE_MAP.get(z.lower())


def parse_ad_size_code(size_str):
    if not size_str:
        return None
    return AD_SIZE_MAP.get(str(size_str).strip().lower())


def parse_issue_code(code):
    """Parse 'TX-May-26' into (market_prefix, month, year)."""
    m = ISSUE_CODE_RE.match(code)
    if not m:
        return None, None, None
    return m.group(1), m.group(2), int(m.group(3)) + 2000


def get_market_for_row(issue_prefix, zone_abbrev):
    """Determine market code for a row. TX rows need zone to distinguish AU vs SA."""
    if issue_prefix in ("CO", "UT"):
        return issue_prefix
    if issue_prefix == "TX" and zone_abbrev:
        if zone_abbrev in ("AN", "AS"):
            return "AU"
        if zone_abbrev in ("SAE", "SAW"):
            return "SA"
    return None


def run(dry_run=True):
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Load lookups
    print("Loading lookups...")
    zones_result = sb.table("zones").select("id,abbreviation,market_id").execute()
    zone_by_abbrev = {z["abbreviation"]: z for z in zones_result.data if z.get("abbreviation")}

    markets_result = sb.table("markets").select("id,code").execute()
    market_by_code = {m["code"]: m for m in markets_result.data}

    # Load all clients (paginated)
    print("Loading clients...")
    all_clients = []
    offset = 0
    while True:
        batch = sb.table("clients").select("id,name,primary_market_id,status").range(offset, offset + 999).execute()
        all_clients.extend(batch.data)
        if len(batch.data) < 1000:
            break
        offset += 1000
    print(f"  {len(all_clients)} clients loaded")

    clients_by_norm = {}
    for c in all_clients:
        key = normalize(c["name"])
        clients_by_norm.setdefault(key, []).append(c)

    def strip_category_suffix(name):
        """Strip ' - Category' suffix from Monday names like 'J & K Roofing - Windows'."""
        if not name:
            return name
        # Also handle ' AU', ' SA' market suffixes
        name = re.sub(r'\s+(AU|SA)$', '', name)
        # Strip ' - Something' but only if the base name is long enough
        m = re.match(r'^(.{6,}?)\s+-\s+\w[\w\s&/]*$', name)
        if m:
            return m.group(1).strip()
        # Strip '(Formerly) ...' patterns
        name = re.sub(r'\s*\(Formerly\).*$', '', name)
        return name

    def lookup_client(name, market_code):
        if not name:
            return None
        norm = normalize(name)
        candidates = clients_by_norm.get(norm, [])

        # Try with category suffix stripped
        if not candidates:
            stripped = strip_category_suffix(name)
            if stripped != name:
                norm_stripped = normalize(stripped)
                candidates = clients_by_norm.get(norm_stripped, [])
                if candidates:
                    norm = norm_stripped

        # Prefix match
        if not candidates:
            for c in all_clients:
                if normalize(c["name"]).startswith(norm + " ") and len(norm) >= 8:
                    candidates = [c]
                    break

        # Prefix match on stripped name too
        if not candidates:
            stripped = strip_category_suffix(name)
            if stripped != name:
                norm_stripped = normalize(stripped)
                for c in all_clients:
                    if normalize(c["name"]).startswith(norm_stripped + " ") and len(norm_stripped) >= 8:
                        candidates = [c]
                        break

        # Fuzzy fallback
        if not candidates:
            best_score = 0
            best_match = None
            for c in all_clients:
                score = similarity(name, c["name"])
                if score > best_score:
                    best_score = score
                    best_match = c
            if best_score >= 0.85 and best_match:
                candidates = [best_match]

        if len(candidates) == 1:
            return candidates[0]["id"]
        if len(candidates) > 1 and market_code:
            market = market_by_code.get(market_code)
            if market:
                matched = [c for c in candidates if c["primary_market_id"] == market["id"]]
                if matched:
                    return matched[0]["id"]
            return candidates[0]["id"]
        if candidates:
            return candidates[0]["id"]
        return None

    # Read spreadsheet
    print(f"\nReading {EXCEL_PATH.name}...")
    wb = load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Parse groups
    groups = {}  # issue_code -> [rows]
    current_group = None
    header_seen = False

    for row in all_rows:
        val = str(row[0] or "").strip()

        # Check if this is a group header
        if ISSUE_CODE_RE.match(val):
            current_group = val
            header_seen = False
            if current_group not in groups:
                groups[current_group] = []
            continue

        # Skip the column header row
        if val == "Name":
            header_seen = True
            continue

        # Data row — skip section headers that leaked through
        if current_group and header_seen and val:
            # Skip rows that look like section headers or metadata
            if re.match(r'^\d{2}\s*-\s', val):  # "01 - Spr Issue Line Item"
                continue
            if re.match(r'^[A-Z]{2}-', val) and ('OPP' in val or 'Crossout' in val or 'Fell Out' in val):
                continue
            groups[current_group].append(row)

    # Report and filter groups
    print(f"\n{'='*60}")
    print(f"  GROUPS FOUND: {len(groups)}")
    print(f"{'='*60}")

    skipped_groups = []
    active_groups = {}
    for code, rows in sorted(groups.items()):
        if len(rows) > MAX_GROUP_SIZE:
            print(f"  {code}: {len(rows)} rows  ** SKIPPING (likely combined sheet)")
            skipped_groups.append(code)
        else:
            print(f"  {code}: {len(rows)} rows")
            active_groups[code] = rows

    # Process rows
    entries = []
    unmatched_clients = {}
    unmatched_zones = {}

    for issue_code, rows in active_groups.items():
        prefix, month, year = parse_issue_code(issue_code)

        for row in rows:
            client_name = to_str(row[COL["name"]])
            if not client_name:
                continue

            zone_raw = to_str(row[COL["zone"]])
            zone_abbrev = parse_zone(zone_raw)
            zone = zone_by_abbrev.get(zone_abbrev) if zone_abbrev else None

            if zone_raw and not zone_abbrev:
                unmatched_zones[zone_raw] = unmatched_zones.get(zone_raw, 0) + 1

            market_code = get_market_for_row(prefix, zone_abbrev)
            # Fallback: use zone's market_id
            if not market_code and zone:
                for mcode, mdata in market_by_code.items():
                    if mdata["id"] == zone["market_id"]:
                        market_code = mcode
                        break

            market = market_by_code.get(market_code) if market_code else None
            client_id = lookup_client(client_name, market_code)

            if not client_id:
                unmatched_clients[client_name] = unmatched_clients.get(client_name, 0) + 1

            ad_size_str = to_str(row[COL["current_ad_size"]])

            entry = {
                "client_id": client_id,
                "market_id": market["id"] if market else None,
                "zone_id": zone["id"] if zone else None,
                "issue_code": issue_code,
                "issue_market": prefix,
                "issue_month": month,
                "issue_year": year,
                "source_client_name": client_name,
                "designer": to_str(row[COL["designer"]]),
                "client_strategy": to_str(row[COL["client_strategy"]]),
                "design_contact": to_str(row[COL["design_contact"]]),
                "ad_size": ad_size_str,
                "ad_size_code": parse_ad_size_code(ad_size_str),
                "past_ad_size": to_str(row[COL["past_ad_size"]]),
                "other_sizes": to_str(row[COL["other_sizes"]]),
                "call_tracking": to_str(row[COL["call_tracking"]]),
                "sales_rep": to_str(row[COL["sales_rep"]]),
                "all_categories": to_str(row[COL["all_categories"]]),
                "primary_category": to_str(row[COL["primary_category"]]),
                "position_notes": to_str(row[COL["position"]]),
                "flip_notes": to_str(row[COL["flip_notes"]]),
                "general_notes": to_str(row[COL["general_notes"]]),
                "in_home_date": parse_date(row[COL["in_home_date"]]),
                "deadline_date": parse_date(row[COL["deadline_date"]]),
                "ad_colors": to_str(row[COL["ad_colors"]]),
                "previous_page": to_int(row[COL["previous_page"]]),
                "page_counter": to_int(row[COL["page_counter"]]),
                "v1_status": to_str(row[COL["v1"]]),
                "v2_status": to_str(row[COL["v2"]]),
                "v3_status": to_str(row[COL["v3"]]),
                "category_notes": to_str(row[COL["category_notes"]]),
                "group_name": to_str(row[COL["group"]]),
                "state": to_str(row[COL["state"]]),
            }
            entries.append(entry)

    # Stats
    matched = sum(1 for e in entries if e["client_id"])
    total = len(entries)

    print(f"\n{'='*60}")
    print(f"  IMPORT PLAN")
    print(f"{'='*60}")
    print(f"  Total entries:          {total}")
    print(f"  Matched to clients:     {matched} ({matched/total*100:.1f}%)" if total else "")
    print(f"  Unmatched:              {total - matched}")
    print(f"  Issues to load:         {len(active_groups)}")
    print(f"  Issues skipped:         {len(skipped_groups)}")
    print(f"{'='*60}")

    if unmatched_zones:
        print(f"\n  Unmatched zones ({len(unmatched_zones)}):")
        for z, c in sorted(unmatched_zones.items(), key=lambda x: -x[1])[:15]:
            print(f"    '{z}': {c} rows")

    if unmatched_clients:
        top_unmatched = sorted(unmatched_clients.items(), key=lambda x: -x[1])[:20]
        print(f"\n  Top unmatched clients ({len(unmatched_clients)} total):")
        for n, c in top_unmatched:
            print(f"    '{n}': {c} rows")

    if dry_run:
        print("\n  DRY RUN - no data written")
        return

    # Delete existing entries for these issue codes, then insert
    BATCH = 200
    for issue_code in active_groups:
        sb.table("runsheet_entries").delete().eq("issue_code", issue_code).execute()

    print(f"\nInserting {len(entries)} entries...")
    inserted = 0
    for i in range(0, len(entries), BATCH):
        batch = entries[i:i + BATCH]
        try:
            sb.table("runsheet_entries").insert(batch).execute()
            inserted += len(batch)
        except Exception as e:
            print(f"  Batch error at {i}, retrying individually: {str(e)[:100]}")
            for row in batch:
                try:
                    sb.table("runsheet_entries").insert(row).execute()
                    inserted += 1
                except Exception as e2:
                    print(f"    Failed: {row['source_client_name']}: {str(e2)[:80]}")
        if inserted % 1000 == 0 and inserted > 0:
            print(f"  ... {inserted} inserted")

    print(f"\n{'='*60}")
    print(f"  COMPLETE - {inserted} entries imported across {len(active_groups)} issues")
    print(f"{'='*60}")


def main():
    parser = argparse.ArgumentParser(description="Import Monday Runsheet data into Supabase")
    parser.add_argument("--dry-run", action="store_true", help="Preview only")
    parser.add_argument("--file", type=str, help="Path to Monday Runsheets Excel file")
    args = parser.parse_args()

    global EXCEL_PATH
    EXCEL_PATH = Path(args.file) if args.file else DEFAULT_EXCEL

    if not all([SUPABASE_URL, SUPABASE_KEY]):
        print("ERROR: Missing env vars")
        sys.exit(1)
    if not EXCEL_PATH.exists():
        print(f"ERROR: File not found at {EXCEL_PATH}")
        sys.exit(1)

    run(dry_run=args.dry_run)


if __name__ == "__main__":
    main()
