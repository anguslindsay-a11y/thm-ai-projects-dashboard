"""
Import Uniqode QR scan data from historical CSV/Excel exports.

Reads the combined Uniqode export and populates the qr_scans table.
Uses Unique Scan ID as the dedupe key.

Usage:
  python setup/import_uniqode_csv.py --dry-run
  python setup/import_uniqode_csv.py
"""

import sys
import os
import csv
import argparse
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
load_dotenv()

from supabase import create_client
from openpyxl import load_workbook

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
DEFAULT_EXCEL = DATA_DIR / "Uniqode Data 2025 - April 2026.xlsx"


def latest_uniqode_file():
    """Auto-detect the latest Uniqode file (CSV or XLSX)."""
    candidates = list(DATA_DIR.glob("*.csv")) + list(DATA_DIR.glob("*[Uu]niqode*.xlsx"))
    # Prefer CSVs (newer format), sorted by modified time desc
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

# Header names to look for (handles both CSV and XLSX column layouts)
HEADER_MAP = {
    "scan_id": ["Unique Scan ID", "unique scan id"],
    "timestamp": ["Timestamp", "timestamp"],
    "product_id": ["Product ID", "product id"],
    "product_name": ["Product Name", "product name"],
    "product_url": ["Product URL", "product url"],
    "user_agent": ["User Agent", "user agent"],
    "city": ["City", "city"],
    "state": ["State", "state"],
}


def find_col_indices(headers, data_rows):
    """Match header names to column indices. Detects if data has more columns
    than headers (Uniqode sometimes adds an extra unlabeled column) and applies
    an offset to correct."""
    header_lower = [str(h).strip().lower() for h in headers]
    indices = {}
    for field, candidates in HEADER_MAP.items():
        for cand in candidates:
            if cand.lower() in header_lower:
                indices[field] = header_lower.index(cand.lower())
                break
    # Detect column offset: if data rows consistently have more columns than headers,
    # shift all indices > 0 by the difference
    if data_rows:
        data_col_count = len(data_rows[0])
        header_col_count = len(headers)
        offset = data_col_count - header_col_count
        if offset > 0:
            for field in indices:
                if indices[field] > 0:
                    indices[field] += offset
    return indices


def read_file(path):
    """Read CSV or XLSX and return (headers, rows)."""
    if path.suffix.lower() == ".csv":
        with open(path, encoding="utf-8") as f:
            reader = csv.reader(f)
            headers = next(reader)
            rows = list(reader)
        return headers, rows
    else:
        wb = load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        headers = [str(h) if h else "" for h in all_rows[0]]
        rows = [list(r) for r in all_rows[1:]]
        return headers, rows


import re


def extract_client_name(campaign_name):
    """Extract a likely client name from a Uniqode campaign name like
    'APEX Clean Air CO Air Quality PopOut Call 7207064662 QR'."""
    if not campaign_name:
        return None
    s = campaign_name.strip()
    # Remove trailing QR / - QR
    s = re.sub(r'\s*-?\s*QR\s*$', '', s, flags=re.IGNORECASE)
    # Remove phone numbers
    s = re.sub(r'\b\d{3}[-.]?\d{3}[-.]?\d{4}\b', '', s)
    # Remove known suffixes/noise
    for noise in ['PopOut', 'Pop-Out', 'Pop Out', 'OPP', 'Contact Form', 'Contact Page',
                  'Contact', 'Website', 'Homepage', 'Promo', 'Booking Page', 'YouTube Video',
                  'Project Gallery', 'Estimate', 'Air Quality', 'Basement Page',
                  'with UTM codes', 'Call', 'Irrigreen']:
        s = re.sub(r'\b' + re.escape(noise) + r'\b', '', s, flags=re.IGNORECASE)
    # Remove market/zone codes
    s = re.sub(r'\b(CO|UT|AU|SA|NCO|THM|NoCO|EPC|SLC)\b', '', s)
    # Remove date fragments
    s = re.sub(r'\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s*\d{0,4}\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\b\d{2,4}\b', '', s)
    # Clean up whitespace and dashes
    s = re.sub(r'\s*-\s*$', '', s)
    s = re.sub(r'^\s*-\s*', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s if s else None


def match_campaign_to_client(campaign_name, client_lookup):
    """Try to match a Uniqode campaign name to an existing client. Returns client_id or None."""
    extracted = extract_client_name(campaign_name)
    if not extracted:
        return None, extracted

    extracted_lower = extracted.lower()

    # Exact match
    for name, cid in client_lookup.items():
        if name.lower() == extracted_lower:
            return cid, extracted

    # Contained match (client name in extracted or vice versa)
    best = None
    best_len = 0
    for name, cid in client_lookup.items():
        nl = name.lower()
        if extracted_lower in nl or nl in extracted_lower:
            overlap = min(len(extracted_lower), len(nl))
            if overlap > best_len and overlap >= 5:
                best_len = overlap
                best = cid
    return best, extracted


def parse_device_type(user_agent):
    """Crude device type parser from user agent string."""
    if not user_agent:
        return None
    ua = user_agent.lower()
    if "iphone" in ua or "ipad" in ua or "ios" in ua:
        return "iOS"
    if "android" in ua:
        return "Android"
    if "windows" in ua or "macintosh" in ua or "linux" in ua:
        return "Desktop"
    return "Other"


def parse_timestamp(ts):
    """Parse the ISO timestamp from the CSV."""
    if ts is None:
        return None
    if isinstance(ts, datetime):
        return ts.isoformat()
    s = str(ts).strip()
    if not s:
        return None
    try:
        return datetime.fromisoformat(s).isoformat()
    except (ValueError, TypeError):
        return None


def run(dry_run=True):
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Load Uniqode -> client_id mapping
    print("Loading Uniqode mappings...")
    result = sb.table("client_platform_ids").select("external_id,client_id").eq("platform", "uniqode").execute()
    pid_to_client = {}
    for row in result.data:
        ext = row["external_id"]
        if ext.startswith("UQ-XX-"):
            pid_to_client[ext.replace("UQ-XX-", "")] = row["client_id"]
    print(f"  {len(pid_to_client)} Uniqode QR codes mapped to clients")

    # Load existing scan IDs to avoid duplicates
    print("Loading existing scan IDs...")
    existing_ids = set()
    offset = 0
    while True:
        batch = sb.table("qr_scans").select("external_scan_id").range(offset, offset + 999).execute()
        for r in batch.data:
            if r.get("external_scan_id"):
                existing_ids.add(r["external_scan_id"])
        if len(batch.data) < 1000:
            break
        offset += 1000
    print(f"  {len(existing_ids)} existing scans in DB")

    # Read the file (CSV or XLSX)
    print(f"Reading {EXCEL_PATH.name}...")
    headers, data_rows = read_file(EXCEL_PATH)
    col = find_col_indices(headers, data_rows)
    print(f"  {len(data_rows)} data rows, {len(col)} columns matched")
    if "scan_id" not in col or "product_id" not in col or "timestamp" not in col:
        print(f"  ERROR: Missing required columns. Found: {col}")
        return

    # Load all client names for fuzzy matching
    print("Loading client names for auto-matching...")
    all_clients = []
    offset = 0
    while True:
        batch = sb.table("clients").select("id,name").range(offset, offset + 999).execute()
        all_clients.extend(batch.data)
        if len(batch.data) < 1000:
            break
        offset += 1000
    client_lookup = {c["name"]: c["id"] for c in all_clients}
    print(f"  {len(client_lookup)} clients available for matching")

    rows_to_insert = []
    auto_mapped = {}
    still_unmatched = {}
    skipped_existing = 0
    skipped_unmatched = 0
    skipped_invalid = 0

    for row in data_rows:
        scan_id = row[col["scan_id"]] if col["scan_id"] < len(row) else None
        if not scan_id:
            skipped_invalid += 1
            continue

        scan_id = str(scan_id).strip()

        # Skip if already imported
        if scan_id in existing_ids:
            skipped_existing += 1
            continue

        product_id = row[col["product_id"]] if col["product_id"] < len(row) else None
        if not product_id:
            skipped_invalid += 1
            continue
        product_id = str(product_id).strip()

        # Lookup client — try existing mapping first, then auto-match by campaign name
        client_id = pid_to_client.get(product_id)
        if not client_id:
            campaign_name = get_col("product_name") if 'get_col' in dir() else (
                str(row[col["product_name"]]).strip() if col.get("product_name") and col["product_name"] < len(row) and row[col["product_name"]] else None
            )
            matched_id, extracted = match_campaign_to_client(campaign_name, client_lookup)
            if matched_id:
                # Auto-create the Uniqode mapping
                pid_to_client[product_id] = matched_id
                client_id = matched_id
                if product_id not in auto_mapped:
                    auto_mapped[product_id] = {"campaign": campaign_name, "extracted": extracted, "client_id": matched_id}
            else:
                if product_id not in still_unmatched:
                    still_unmatched[product_id] = {"campaign": campaign_name, "extracted": extracted}
                skipped_unmatched += 1
                continue

        ts_val = row[col["timestamp"]] if col["timestamp"] < len(row) else None
        scan_time = parse_timestamp(ts_val)
        if not scan_time:
            skipped_invalid += 1
            continue

        def get_col(field):
            idx = col.get(field)
            if idx is not None and idx < len(row) and row[idx]:
                return str(row[idx]).strip()
            return None

        record = {
            "external_scan_id": scan_id,
            "qr_code_id": product_id,
            "client_id": client_id,
            "scan_time": scan_time,
            "campaign_name": get_col("product_name"),
            "source_url": get_col("product_url"),
            "scan_city": get_col("city"),
            "scan_state": get_col("state"),
            "device_type": parse_device_type(get_col("user_agent")),
        }
        rows_to_insert.append(record)

    # Deduplicate by scan ID (CSV sometimes has duplicate rows)
    seen_ids = set()
    deduped = []
    for r in rows_to_insert:
        if r["external_scan_id"] not in seen_ids:
            seen_ids.add(r["external_scan_id"])
            deduped.append(r)
    dupes_removed = len(rows_to_insert) - len(deduped)
    rows_to_insert = deduped

    print(f"\n{'='*60}")
    print(f"  IMPORT PLAN")
    print(f"{'='*60}")
    print(f"  Rows to insert:       {len(rows_to_insert)}")
    print(f"  Auto-matched QR codes:{len(auto_mapped)}")
    print(f"  Skipped (existing):   {skipped_existing}")
    print(f"  Skipped (unmatched):  {skipped_unmatched}")
    print(f"  Skipped (invalid):    {skipped_invalid}")
    print(f"{'='*60}")

    if auto_mapped:
        print(f"\n  Auto-matched {len(auto_mapped)} NEW QR codes to existing clients:")
        for pid, info in auto_mapped.items():
            client_name = next((n for n, cid in client_lookup.items() if cid == info['client_id']), '?')
            print(f"    {pid} \"{info['campaign']}\" -> {client_name}")

    if still_unmatched:
        print(f"\n  {len(still_unmatched)} QR codes could NOT be matched (manual review needed):")
        for pid, info in still_unmatched.items():
            print(f"    {pid} \"{info['campaign']}\" (extracted: \"{info['extracted']}\")")

    if dry_run:
        print("\n  DRY RUN — no data written.")
        if rows_to_insert:
            print(f"\n  Sample row:")
            for k, v in rows_to_insert[0].items():
                print(f"    {k}: {v}")
        return

    # Create Uniqode platform mappings for auto-matched QR codes
    if auto_mapped:
        print(f"\nCreating {len(auto_mapped)} new Uniqode platform mappings...")
        for pid, info in auto_mapped.items():
            try:
                sb.table("client_platform_ids").insert({
                    "client_id": info["client_id"],
                    "platform": "uniqode",
                    "external_id": f"UQ-XX-{pid}",
                    "external_name": info["campaign"],
                }).execute()
            except Exception:
                pass  # Already exists

    # Insert in batches
    print(f"\nInserting {len(rows_to_insert)} scans...")
    BATCH_SIZE = 200
    inserted = 0
    for i in range(0, len(rows_to_insert), BATCH_SIZE):
        batch = rows_to_insert[i:i + BATCH_SIZE]
        sb.table("qr_scans").upsert(batch, on_conflict="external_scan_id").execute()
        inserted += len(batch)
        if inserted % 1000 == 0:
            print(f"  ... {inserted} inserted")

    print(f"\n{'='*60}")
    print(f"  COMPLETE — {inserted} scans imported")
    print(f"{'='*60}")


def main():
    parser = argparse.ArgumentParser(description="Import Uniqode QR scan CSV/Excel into Supabase")
    parser.add_argument("--dry-run", action="store_true", help="Preview only")
    parser.add_argument("--file", type=str, help="Path to Uniqode Excel file")
    args = parser.parse_args()

    global EXCEL_PATH
    if args.file:
        EXCEL_PATH = Path(args.file)
    else:
        EXCEL_PATH = latest_uniqode_file()
        if not EXCEL_PATH:
            print("ERROR: No Uniqode file found in data/")
            sys.exit(1)

    if not SUPABASE_URL or not SUPABASE_KEY:
        print("ERROR: Missing env vars")
        sys.exit(1)

    if not EXCEL_PATH.exists():
        print(f"ERROR: File not found at {EXCEL_PATH}")
        sys.exit(1)

    run(dry_run=args.dry_run)


if __name__ == "__main__":
    main()
