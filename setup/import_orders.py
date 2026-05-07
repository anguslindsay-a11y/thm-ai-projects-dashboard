"""
Import Waterfall Order Data into Supabase orders table.

Reads the latest Waterfall spreadsheet and populates the orders table.
Matches companies to existing clients by name. Creates new clients with full
metadata (market, status, platform ID) for unmatched companies.
Uses mm_order_id + market_id as the unique key (upsert).

After upserting, detects orders that were REMOVED from the spreadsheet (cancelled/
paused) and deletes them from the DB. Only touches markets present in the file.

Usage:
  python setup/import_orders.py --dry-run          # Preview only
  python setup/import_orders.py                     # Run the import (latest file in data/)
  python setup/import_orders.py --file <path>       # Use a specific file
  python setup/import_orders.py --no-remove         # Skip the removed-orders pass
"""

import sys
import os
import re
import argparse
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
load_dotenv()

from supabase import create_client
from openpyxl import load_workbook
from helpers import parse_zone_from_product

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scripts"))
from import_report import generate_report

DATA_DIR = Path(__file__).resolve().parent.parent / "data"


def latest_waterfall():
    """Find the most recent Waterfall Order Data file in data/."""
    candidates = sorted(DATA_DIR.glob("Waterfall Order Data*Supabase*.xlsx"), reverse=True)
    if candidates:
        return candidates[0]
    candidates = sorted(DATA_DIR.glob("Waterfall Order Data*.xlsx"), reverse=True)
    return candidates[0] if candidates else None

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

# Mkt values in spreadsheet -> market codes in Supabase
MKT_TO_MARKET_CODE = {
    "CO": "CO",
    "UT": "UT",
    "AU": "AU",
    "SA": "SA",
}

# Columns to import (index -> field name)
# Skipping Excel helper columns: Issue Date Filter, Company Match, Mkt Filter,
# Start Issue, Category, Start Issue Filter, Priority, Priority Filter
COL_MAP = {
    0: "mkt",              # -> zone lookup
    1: "issue_date",       # text format "YY.MM.MonAbbrev" e.g. "24.05.May"
    2: "company",          # -> client lookup
    3: "product",
    4: "size",
    5: "position",
    6: "notes",
    7: "net",
    8: "gross",
    9: "sales_rep",
    10: "commission_rep",
    11: "contact_type",
    12: "ia_category",
    13: "opp_category",
    14: "mm_order_id",
    15: "special_section",
    16: "amount_due",
    17: "proposal_type",
    # col 18: "Match" (new helper column, skip)
    19: "space",
    # col 20-23: filters, skip
    # col 24: Priority, skip
    # col 25: Mkt Filter, skip
    26: "start_issue",     # datetime — client's original start issue
    27: "category",        # THM category (Glass & Doors, Windows, etc.)
    # col 28: Sales Attrib, skip
    29: "biz_category",
    30: "start_issue_filter",  # also datetime, duplicate of start_issue
}


def read_spreadsheet():
    wb = load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    headers = all_rows[0]
    data = []
    for row in all_rows[1:]:
        record = {}
        for idx, field in COL_MAP.items():
            record[field] = row[idx]
        data.append(record)
    wb.close()
    return data


def parse_date(val):
    """Handles: datetime objects, Waterfall's 'YY.MM.Mon' text, and standard date strings."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    # Try MM-style strings first
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Waterfall text format: "YY.MM.Mon" e.g. "24.05.May", "25.04.Spr"
    # Use the YY.MM portion; pin day to 15 (mid-month) for stable comparisons
    m = re.match(r"^(\d{2})\.(\d{2})\.", s)
    if m:
        yy, mm = int(m.group(1)), int(m.group(2))
        year = 2000 + yy
        if 1 <= mm <= 12:
            return f"{year:04d}-{mm:02d}-15"
    return None


def parse_year(val):
    """Extract year from datetime, int, or 'YY.MM.Mon' string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.year
    s = str(val).strip()
    if not s:
        return None
    # YY.MM.Mon format
    m = re.match(r"^(\d{2})\.", s)
    if m:
        return 2000 + int(m.group(1))
    try:
        return int(s)
    except (ValueError, TypeError):
        return None


def to_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def to_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def run_import(rows, dry_run=True, skip_remove=False):
    total = len(rows)
    unique_companies = set(to_str(r["company"]) for r in rows if r.get("company"))
    mkts = {}
    for r in rows:
        m = to_str(r.get("mkt"))
        if m:
            mkts[m] = mkts.get(m, 0) + 1

    print(f"\n{'='*50}")
    print(f"  WATERFALL ORDER DATA ANALYSIS")
    print(f"{'='*50}")
    print(f"  Total rows:          {total}")
    print(f"  Unique companies:    {len(unique_companies)}")
    print(f"  Markets:             {mkts}")
    print(f"{'='*50}\n")

    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Load markets and zones
    markets_result = sb.table("markets").select("*").execute()
    market_lookup = {m["code"]: m for m in markets_result.data}

    zones_result = sb.table("zones").select("*").execute()
    zone_by_abbrev = {z["abbreviation"]: z for z in zones_result.data}

    # Load all clients
    print("Loading existing clients...")
    all_clients = []
    offset = 0
    while True:
        batch = sb.table("clients").select("id,name").range(offset, offset + 999).execute()
        all_clients.extend(batch.data)
        if len(batch.data) < 1000:
            break
        offset += 1000
    client_lookup = {c["name"]: c["id"] for c in all_clients}
    print(f"  {len(client_lookup)} clients in DB")

    # ---- Enhancement A: Find unmatched companies with full metadata ----
    # Build per-company metadata from first occurrence in spreadsheet
    client_meta = {}
    for r in rows:
        company = to_str(r.get("company"))
        mkt = to_str(r.get("mkt"))
        if company and company not in client_lookup and company not in client_meta:
            # Prefer column 24 "Category" (THM category) over column 25 "biz_category" (InBook/OPP)
            cat = to_str(r.get("category")) or None
            client_meta[company] = {
                "mkt_code": MKT_TO_MARKET_CODE.get(mkt),
                "category": cat,
            }

    if client_meta:
        print(f"\n  {len(client_meta)} NEW companies found (not in DB):")
        for company in sorted(client_meta.keys()):
            meta = client_meta[company]
            print(f"    + {company} (market={meta['mkt_code'] or '?'}, category={meta.get('category') or '—'})")

    if dry_run:
        # ---- Dry-run: preview removed orders too ----
        print(f"\n  DRY RUN — no data will be written.\n")
        print(f"  Would upsert up to {total} order records")
        print(f"  Would create {len(client_meta)} new clients with market + platform ID")
        # (dry-run continues below and returns None)

        if not skip_remove:
            # Build spreadsheet keys and compare to DB
            spreadsheet_keys = {}  # {market_code: set of mm_order_ids}
            for r in rows:
                oid = r.get("mm_order_id")
                mkt = MKT_TO_MARKET_CODE.get(to_str(r.get("mkt")))
                if oid and mkt:
                    spreadsheet_keys.setdefault(mkt, set()).add(int(oid))

            total_removed = 0
            for mkt_code, file_ids in spreadsheet_keys.items():
                if mkt_code not in market_lookup:
                    continue
                mid = market_lookup[mkt_code]["id"]
                db_ids = set()
                off = 0
                while True:
                    batch = sb.table("orders").select("mm_order_id").eq("market_id", mid).range(off, off + 999).execute().data
                    db_ids.update(r["mm_order_id"] for r in batch)
                    if len(batch) < 1000:
                        break
                    off += 1000
                removed = db_ids - file_ids
                if removed:
                    print(f"  Would REMOVE {len(removed)} cancelled orders in {mkt_code}")
                    total_removed += len(removed)
            print(f"  Total removals: {total_removed}")
        print()
        return None

    # ---- Create new clients with full metadata ----
    if client_meta:
        print(f"\nCreating {len(client_meta)} new clients with metadata...")
        created = 0
        for company in sorted(client_meta.keys()):
            meta = client_meta[company]
            mkt_code = meta["mkt_code"]

            insert_data = {
                "name": company,
                "status": "active",
            }
            if mkt_code and mkt_code in market_lookup:
                insert_data["primary_market_id"] = market_lookup[mkt_code]["id"]
            if meta.get("category"):
                insert_data["category"] = meta["category"]

            result = sb.table("clients").insert(insert_data).execute()
            new_id = result.data[0]["id"]
            client_lookup[company] = new_id

            # Create Magazine Manager platform ID
            try:
                sb.table("client_platform_ids").insert({
                    "client_id": new_id,
                    "platform": "magazine_manager",
                    "external_id": f"MM-{mkt_code or 'XX'}-auto-{company[:80]}",
                    "external_name": company,
                }).execute()
            except Exception:
                pass  # Duplicate — already exists

            created += 1
            if created % 50 == 0:
                print(f"  ... {created} created")
        print(f"  Created {created} new clients")

    # ---- Upsert orders in batches ----
    print(f"\nUpserting {total} orders...")
    inserted = 0
    skipped = 0
    errors = 0
    batch = []
    BATCH_SIZE = 100

    for r in rows:
        company = to_str(r.get("company"))
        mkt = to_str(r.get("mkt"))
        order_id = r.get("mm_order_id")

        mkt_code = MKT_TO_MARKET_CODE.get(mkt)
        if not order_id or not mkt_code or mkt_code not in market_lookup:
            skipped += 1
            continue

        client_id = client_lookup.get(company) if company else None
        market_id = market_lookup[mkt_code]["id"]

        # Parse zone from product column
        product_str = to_str(r.get("product"))
        zone_abbrev = parse_zone_from_product(product_str)
        zone_id = zone_by_abbrev[zone_abbrev]["id"] if zone_abbrev and zone_abbrev in zone_by_abbrev else None

        record = {
            "mm_order_id": int(order_id),
            "client_id": client_id,
            "market_id": market_id,
            "zone_id": zone_id,
            "issue_date": to_str(r.get("issue_date")),
            "issue_date_parsed": parse_date(r.get("issue_date")),
            "product": to_str(r.get("product")),
            "size": to_str(r.get("size")),
            "position": to_str(r.get("position")),
            "notes": to_str(r.get("notes")),
            "net": to_float(r.get("net")),
            "gross": to_float(r.get("gross")),
            "amount_due": to_float(r.get("amount_due")),
            "sales_rep": to_str(r.get("sales_rep")),
            "commission_rep": to_str(r.get("commission_rep")),
            "contact_type": to_str(r.get("contact_type")),
            "ia_category": to_str(r.get("ia_category")),
            "opp_category": to_str(r.get("opp_category")),
            "biz_category": to_str(r.get("biz_category")),
            "special_section": to_str(r.get("special_section")),
            "proposal_type": to_str(r.get("proposal_type")),
            "space": to_str(r.get("space")),
            "year": parse_year(r.get("issue_date")),
        }
        batch.append(record)

        if len(batch) >= BATCH_SIZE:
            try:
                sb.table("orders").upsert(
                    batch, on_conflict="mm_order_id,market_id"
                ).execute()
                inserted += len(batch)
            except Exception as e:
                for rec in batch:
                    try:
                        sb.table("orders").upsert(
                            rec, on_conflict="mm_order_id,market_id"
                        ).execute()
                        inserted += 1
                    except Exception as e2:
                        errors += 1
                        if errors <= 5:
                            print(f"  WARNING: Order {rec['mm_order_id']}: {str(e2)[:100]}")
            batch = []
            if inserted % 5000 == 0 and inserted > 0:
                print(f"  ... {inserted} orders upserted")

    if batch:
        try:
            sb.table("orders").upsert(
                batch, on_conflict="mm_order_id,market_id"
            ).execute()
            inserted += len(batch)
        except Exception:
            for rec in batch:
                try:
                    sb.table("orders").upsert(
                        rec, on_conflict="mm_order_id,market_id"
                    ).execute()
                    inserted += 1
                except Exception:
                    errors += 1

    print(f"\n  Upserted: {inserted}  |  Skipped: {skipped}  |  Errors: {errors}")

    # ---- Enhancement B: Detect and remove cancelled orders ----
    removed_total = 0
    removed_by_client: dict = {}  # {client_id: {"orders_removed": N, "gross_lost": $, "market": mkt_code}}
    if not skip_remove:
        print("\nDetecting cancelled orders (in DB but NOT in spreadsheet)...")
        # Build spreadsheet order IDs per market
        spreadsheet_keys = {}
        for r in rows:
            oid = r.get("mm_order_id")
            mkt = MKT_TO_MARKET_CODE.get(to_str(r.get("mkt")))
            if oid and mkt:
                spreadsheet_keys.setdefault(mkt, set()).add(int(oid))

        for mkt_code, file_ids in spreadsheet_keys.items():
            if mkt_code not in market_lookup:
                continue
            mid = market_lookup[mkt_code]["id"]
            # Load all existing order IDs + client + gross + issue date for this market from DB
            db_orders = []
            off = 0
            while True:
                batch = sb.table("orders").select(
                    "id,mm_order_id,client_id,gross,size,issue_date_parsed"
                ).eq("market_id", mid).range(off, off + 999).execute().data
                db_orders.extend(batch)
                if len(batch) < 1000:
                    break
                off += 1000

            orders_to_remove = [o for o in db_orders if o["mm_order_id"] not in file_ids]
            if orders_to_remove:
                print(f"  {mkt_code}: removing {len(orders_to_remove)} cancelled orders")
                # Aggregate by client before deletion
                for o in orders_to_remove:
                    cid = o.get("client_id")
                    if not cid:
                        continue
                    entry = removed_by_client.setdefault(cid, {
                        "orders_removed": 0, "gross_lost": 0.0, "market": mkt_code,
                    })
                    entry["orders_removed"] += 1
                    entry["gross_lost"] += float(o.get("gross") or 0)
                # Delete in batches of 100
                to_remove_ids = [o["id"] for o in orders_to_remove]
                for i in range(0, len(to_remove_ids), 100):
                    chunk = to_remove_ids[i:i+100]
                    sb.table("orders").delete().in_("id", chunk).execute()
                removed_total += len(to_remove_ids)

    print(f"\n{'='*50}")
    print(f"  ORDER IMPORT COMPLETE")
    print(f"{'='*50}")
    print(f"  Orders upserted:     {inserted}")
    print(f"  Skipped:             {skipped}")
    print(f"  Errors:              {errors}")
    print(f"  New clients created: {len(client_meta)}")
    print(f"  Cancelled removed:   {removed_total}")
    print(f"{'='*50}\n")

    return {
        "new_clients": client_meta,
        "removed_count": removed_total,
        "removed_by_client": removed_by_client,
    }


def main():
    parser = argparse.ArgumentParser(description="Import Waterfall order data into Supabase")
    parser.add_argument("--dry-run", action="store_true", help="Preview only")
    parser.add_argument("--no-remove", action="store_true", help="Skip the removed-orders pass")
    parser.add_argument("--no-email", action="store_true", help="Skip the email report")
    parser.add_argument("--file", type=str, help="Path to Waterfall Excel file (auto-detects latest if omitted)")
    args = parser.parse_args()

    global EXCEL_PATH
    if args.file:
        EXCEL_PATH = Path(args.file)
    else:
        EXCEL_PATH = latest_waterfall()
        if not EXCEL_PATH:
            print("ERROR: No Waterfall Order Data file found in data/")
            sys.exit(1)

    if not SUPABASE_URL or not SUPABASE_KEY:
        print("ERROR: SUPABASE_URL and SUPABASE_KEY must be set in .env")
        sys.exit(1)
    if not EXCEL_PATH.exists():
        print(f"ERROR: Spreadsheet not found at {EXCEL_PATH}")
        sys.exit(1)

    print(f"Reading spreadsheet: {EXCEL_PATH.name}")
    rows = read_spreadsheet()
    result = run_import(rows, dry_run=args.dry_run, skip_remove=args.no_remove)

    # Post-import report (status sync + change log + email)
    if not args.dry_run and result:
        new_clients_info = [
            {"name": name, "market": meta.get("mkt_code", "?"), "category": meta.get("category", "")}
            for name, meta in result.get("new_clients", {}).items()
        ]
        generate_report(
            new_clients=new_clients_info,
            removed_count=result.get("removed_count", 0),
            removed_by_client=result.get("removed_by_client", {}),
            send=not args.no_email,
        )


if __name__ == "__main__":
    main()
