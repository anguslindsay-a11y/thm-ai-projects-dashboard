"""
Phase 4 — Import manual category approvals from a reviewed Excel file.

Reads the "1. Disagreements", "2. Low Confidence", "3. Medium Confidence" sheets
produced by scripts/build_category_review.py. For each row where the user has
filled `approved_primary`:

  - 'OK' -> upgrade existing LLM tags to source='manual' (lock them in)
  - <category name> -> replace LLM tags with the named primary; treat
                       comma-list in approved_secondaries as additional tags
  - <empty> -> skip the row

Once a client has source='manual' tags, future auto-classification skips them.

Usage:
  python setup/import_category_approvals.py "output/[C] Category Review 2026-04-29 4-29-2026.xlsx"
"""
from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from supabase import create_client

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

REVIEW_TABS = ["1. Disagreements", "2. Low Confidence", "3. Medium Confidence"]


def load_decisions(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    decisions = []
    for tab in REVIEW_TABS:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h: i for i, h in enumerate(headers) if h}
        if "client" not in col or "approved_primary" not in col:
            print(f"  [skip] {tab} missing 'client' or 'approved_primary' columns")
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            client = row[col["client"]] if col.get("client") is not None else None
            approved = row[col["approved_primary"]] if col.get("approved_primary") is not None else None
            secondaries = row[col["approved_secondaries"]] if col.get("approved_secondaries") is not None else None
            if not client or not approved:
                continue
            decisions.append({
                "client": str(client).strip(),
                "approved_primary": str(approved).strip(),
                "approved_secondaries": str(secondaries).strip() if secondaries else "",
                "tab": tab,
            })
    return decisions


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("path", help="Path to reviewed Excel file")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    sb = create_client(SUPABASE_URL, SUPABASE_KEY)
    decisions = load_decisions(Path(args.path))
    print(f"Loaded {len(decisions)} approved rows")

    if not decisions:
        return

    # Lookup categories by name (case-insensitive)
    cats = sb.table("categories").select("id,name,slug").execute().data
    name_to_cat = {c["name"].lower(): c for c in cats}
    slug_to_cat = {c["slug"]: c for c in cats}

    # Lookup clients by name (exact match)
    name_to_client = {}
    page = 0
    while True:
        chunk = sb.table("clients").select("id,name").range(page * 1000, page * 1000 + 999).execute().data
        if not chunk:
            break
        for c in chunk:
            name_to_client[c["name"]] = c["id"]
        if len(chunk) < 1000:
            break
        page += 1

    applied = ok_locked = replaced = skipped = unmatched = 0

    for d in decisions:
        client_id = name_to_client.get(d["client"])
        if not client_id:
            unmatched += 1
            print(f"  [unmatched] {d['client']}")
            continue

        approved_norm = d["approved_primary"].strip()
        if approved_norm.upper() == "OK":
            # Lock existing LLM tags as manual
            if not args.dry_run:
                sb.table("client_categories").update({"source": "manual"}).eq("client_id", client_id).eq("source", "llm_auto").execute()
            ok_locked += 1
            applied += 1
            continue

        # Resolve named primary
        cat = name_to_cat.get(approved_norm.lower()) or slug_to_cat.get(approved_norm.lower())
        if not cat:
            print(f"  [unknown category] {d['client']!r} -> {approved_norm!r}")
            skipped += 1
            continue

        # Resolve secondaries
        secondary_cats = []
        if d["approved_secondaries"]:
            for s in d["approved_secondaries"].split(","):
                s = s.strip()
                if not s:
                    continue
                sec = name_to_cat.get(s.lower()) or slug_to_cat.get(s.lower())
                if sec:
                    secondary_cats.append(sec)
                else:
                    print(f"  [unknown secondary] {d['client']!r} -> {s!r}")

        if args.dry_run:
            secs = ",".join(s["name"] for s in secondary_cats)
            print(f"  [DRY] {d['client']}: primary={cat['name']} secondaries=[{secs}]")
            applied += 1
            replaced += 1
            continue

        # Replace LLM tags with manual decision
        sb.table("client_categories").delete().eq("client_id", client_id).in_("source", ["llm_auto", "legacy_text"]).execute()
        rows = [{"client_id": client_id, "category_id": cat["id"], "is_primary": True, "source": "manual"}]
        for sec in secondary_cats:
            rows.append({"client_id": client_id, "category_id": sec["id"], "is_primary": False, "source": "manual"})
        sb.table("client_categories").upsert(rows, on_conflict="client_id,category_id").execute()
        applied += 1
        replaced += 1

    print(f"\n  applied: {applied}")
    print(f"    OK-locked (kept LLM): {ok_locked}")
    print(f"    replaced with manual: {replaced}")
    print(f"  unmatched clients: {unmatched}")
    print(f"  skipped (unknown category): {skipped}")


if __name__ == "__main__":
    main()
