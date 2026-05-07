"""
Import call tracking yes/no flags from THM Call Tracking.xlsx.

Source format:
  Mkt | Company Name | Rep Name | Call Tracking | Call Track Notes

CO/UT: 'Call Tracking' column has tokens like 'Call Track {ZONE} YES/NO',
       multiple tokens separated by '~'.
TX:    'Call Tracking' column is empty. Default to YES; flip to NO when notes
       contain markers like 'NCT', 'in ad', 'in book', 'cancelled', etc.

Writes:
  - client_zones.has_call_tracking  (per-zone)
  - clients.has_call_tracking       (rollup, TRUE if any zone TRUE)
  - clients.call_tracking_notes     (raw notes preserved)

Usage:
  python setup/import_call_tracking.py [--dry-run]
"""

import os
import re
import sys
import argparse
from pathlib import Path
from collections import defaultdict, Counter

import openpyxl
from dotenv import load_dotenv
from supabase import create_client

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

DATA_FILE = Path(__file__).resolve().parent.parent / "data" / "THM Call Tracking.xlsx"

# --- Zone token mapping ---
# Maps the 'Call Track X YES/NO' token to one or more zone abbreviations in DB.
ZONE_TOKEN_MAP = {
    # CO
    "SDenver": ["SD"],
    "NDenver": ["ND"],
    "NoCO": ["NOCO"],
    "EPC": ["EPC"],
    "N/S": ["ND", "SD"],
    # UT
    "NW": ["NW"],
    "CW": ["CW"],
    "SW": ["SW"],
    # Skip (cross-book / cross-market provided)
    "XBook": [],
}

ZONE_TOKEN_REGEX = re.compile(r"^\s*Call Track\s+(.+?)\s+(YES|NO)\s*$", re.IGNORECASE)

# --- TX no-tracking detection ---
# Case-insensitive substring match against the Call Track Notes field.
TX_NO_MARKERS = [
    "nct",
    "in ad",
    "in book",
    "inbook",
    "no call track",
    "not call track",
    "no ct",
    "cancelled",
    "canceled",
    "retired",
    "no longer",        # "no longer using", "no longer on CT"
    "not using",        # "not using CT"
    "off ct",
    "off cs",
    "stopped",
    "discontinued",
]


def is_not_tracked_by_notes(notes: str) -> bool:
    """Return True if free-text notes mark the client as NOT on call tracking.

    Used for TX (no structured CT column) and as a fallback for CO/UT rows
    where the structured Call Tracking column is blank.
    """
    if not notes or not notes.strip():
        return True
    n = notes.lower()
    return any(m in n for m in TX_NO_MARKERS)


def parse_co_ut_zones(call_tracking_str: str) -> dict:
    """Parse 'Call Track X YES~Call Track Y NO' -> {zone_abbr: bool}."""
    result = {}
    if not call_tracking_str:
        return result
    for token in str(call_tracking_str).split("~"):
        m = ZONE_TOKEN_REGEX.match(token.strip())
        if not m:
            continue
        zone_label = m.group(1).strip()
        is_yes = m.group(2).upper() == "YES"
        for abbr in ZONE_TOKEN_MAP.get(zone_label, []):
            # If already set by another token (e.g. N/S expanding to ND+SD,
            # then a separate explicit SDenver token), prefer the explicit YES.
            result[abbr] = result.get(abbr, False) or is_yes
            if not is_yes and abbr not in result:
                result[abbr] = False
    return result


def load_rows(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        mkt, name, rep, ct, notes = (row + (None,) * 5)[:5]
        if not name:
            continue
        rows.append({
            "market": (mkt or "").strip(),
            "name": str(name).strip(),
            "rep": (rep or "").strip() if rep else None,
            "call_tracking_raw": str(ct).strip() if ct else None,
            "notes": str(notes).strip() if notes else None,
        })
    return rows


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    print(f"Loading {DATA_FILE.name}...")
    rows = load_rows(DATA_FILE)
    by_mkt = Counter(r["market"] for r in rows)
    print(f"  {len(rows)} rows | by market: {dict(by_mkt)}")

    # --- Pull lookup data ---
    print("Loading clients + zones from Supabase...")
    clients = []
    page = 0
    while True:
        chunk = sb.table("clients").select("id,name").range(page * 1000, page * 1000 + 999).execute().data
        if not chunk:
            break
        clients.extend(chunk)
        if len(chunk) < 1000:
            break
        page += 1
    name_to_client = {c["name"]: c["id"] for c in clients}
    print(f"  {len(clients)} clients")

    zones = sb.table("zones").select("id,abbreviation,market_id,markets(code)").execute().data
    zone_abbr_to_id = {z["abbreviation"]: z["id"] for z in zones}

    # client_zones: client_id -> list of (zone_id, zone_abbr)
    cz_rows = []
    page = 0
    while True:
        chunk = sb.table("client_zones").select("id,client_id,zone_id").range(page * 1000, page * 1000 + 999).execute().data
        if not chunk:
            break
        cz_rows.extend(chunk)
        if len(chunk) < 1000:
            break
        page += 1
    zone_id_to_abbr = {z["id"]: z["abbreviation"] for z in zones}
    cz_by_client = defaultdict(dict)  # client_id -> {zone_abbr: cz_id}
    for r in cz_rows:
        abbr = zone_id_to_abbr.get(r["zone_id"])
        if abbr:
            cz_by_client[r["client_id"]][abbr] = r["id"]
    print(f"  {len(cz_rows)} client_zones rows")

    # --- Process rows ---
    unmatched = []
    cz_updates = {}      # cz_id -> bool
    client_notes = {}    # client_id -> notes (latest seen)
    client_rollup = defaultdict(lambda: False)  # client_id -> bool (any zone tracked)
    no_zone_records = [] # TX clients with no client_zones rows
    stats = Counter()

    for row in rows:
        mkt = row["market"]
        client_id = name_to_client.get(row["name"])
        if not client_id:
            unmatched.append(row)
            stats[f"unmatched_{mkt}"] += 1
            continue

        # Save raw notes for audit (last write wins if duplicate name)
        if row["notes"]:
            client_notes[client_id] = row["notes"]

        zone_flags = parse_co_ut_zones(row["call_tracking_raw"]) if mkt in ("CO", "UT") else {}

        if zone_flags:
            # Structured CO/UT path: apply per-zone flags
            applied = 0
            for abbr, is_tracked in zone_flags.items():
                cz_id = cz_by_client.get(client_id, {}).get(abbr)
                if not cz_id:
                    stats[f"{mkt}_zone_not_in_db"] += 1
                    continue
                cz_updates[cz_id] = is_tracked
                if is_tracked:
                    client_rollup[client_id] = True
                applied += 1
            if applied:
                stats[f"{mkt}_applied"] += 1
        else:
            # Notes-based fallback: TX always lands here, plus any CO/UT row
            # with a blank Call Tracking column.
            tracked = not is_not_tracked_by_notes(row["notes"])
            client_zones_for = cz_by_client.get(client_id, {})
            # For TX, propagate to AU/SA zones if any exist on the client.
            tx_zones = {"AN", "AS", "SAE", "SAW"}
            target_abbrs = (tx_zones & set(client_zones_for.keys())) if mkt == "TX" else set()
            if target_abbrs:
                for abbr in target_abbrs:
                    cz_updates[client_zones_for[abbr]] = tracked
            elif mkt == "TX":
                no_zone_records.append((row["name"], tracked))
            if tracked:
                client_rollup[client_id] = True
            label = "TX" if mkt == "TX" else f"{mkt}_notes_fallback"
            stats[f"{label}_{('YES' if tracked else 'NO')}"] += 1

    # Make sure every matched client gets an explicit rollup value
    for row in rows:
        cid = name_to_client.get(row["name"])
        if cid and cid not in client_rollup:
            client_rollup[cid] = False

    # ---------------------------------------------------------------
    # GUARD: trust CallRail evidence over the source file — TX ONLY.
    #
    # CO/UT have a structured 'Call Tracking' column in the MM export that
    # is authoritative — when CO/UT clients move from CT to in-ad, the file
    # is updated and we should respect that. TX has no structured column;
    # the entire signal lives in free-text notes that frequently miss
    # active CallRail tracking. So the guard only applies to TX clients.
    #
    # Rule: TX client + has CallRail platform_id + at least one call in
    # the last 365 days  =>  force has_call_tracking = true.
    # ---------------------------------------------------------------
    print("\nApplying CallRail-evidence guard for TX clients...")

    # Build the set of TX client_ids (AU/SA markets)
    tx_client_ids = set()
    page = 0
    while True:
        chunk = (sb.table("clients")
                 .select("id,markets!inner(code)")
                 .in_("markets.code", ["AU", "SA"])
                 .range(page * 1000, page * 1000 + 999).execute().data)
        if not chunk:
            break
        for r in chunk:
            tx_client_ids.add(r["id"])
        if len(chunk) < 1000:
            break
        page += 1
    print(f"  scoped to {len(tx_client_ids)} TX clients")

    # Pull every CallRail mapping (client_id -> [external_ids])
    cpids = []
    page = 0
    while True:
        chunk = (sb.table("client_platform_ids")
                 .select("client_id,external_id")
                 .eq("platform", "callrail")
                 .range(page * 1000, page * 1000 + 999).execute().data)
        if not chunk:
            break
        cpids.extend(chunk)
        if len(chunk) < 1000:
            break
        page += 1
    client_to_cr_ids = defaultdict(set)
    for r in cpids:
        client_to_cr_ids[r["client_id"]].add(r["external_id"])

    # Pull DISTINCT callrail_company_ids with a call in the last 365 days.
    # Supabase REST caps each page at 1000 rows, so we paginate fully.
    from datetime import datetime, timedelta, timezone
    cutoff = (datetime.now(timezone.utc) - timedelta(days=365)).isoformat()
    active_cr_ids = set()
    PAGE_SIZE = 1000
    page = 0
    while True:
        chunk = (sb.table("calls")
                 .select("callrail_company_id")
                 .gte("call_time", cutoff)
                 .order("call_time", desc=True)
                 .range(page * PAGE_SIZE, page * PAGE_SIZE + PAGE_SIZE - 1).execute().data)
        if not chunk:
            break
        for r in chunk:
            if r["callrail_company_id"]:
                active_cr_ids.add(r["callrail_company_id"])
        if len(chunk) < PAGE_SIZE:
            break
        page += 1
    print(f"  scanned calls (365d): {len(active_cr_ids)} distinct CallRail accounts active")

    # Find TX clients to override
    overridden_clients = []
    for cid, cr_ids in client_to_cr_ids.items():
        if cid not in tx_client_ids:
            continue  # CO/UT respect the structured source file — no override
        if cr_ids & active_cr_ids:
            if client_rollup.get(cid) is False:
                overridden_clients.append(cid)
                client_rollup[cid] = True
                # Cascade to TX zones
                for abbr, cz_id in cz_by_client.get(cid, {}).items():
                    if abbr in {"AN", "AS", "SAE", "SAW"}:
                        cz_updates[cz_id] = True

    print(f"  guard overrode {len(overridden_clients)} TX clients (CallRail evidence beat source file)")

    print()
    print("=== Stats ===")
    for k, v in sorted(stats.items()):
        print(f"  {k}: {v}")
    print(f"  unmatched total: {len(unmatched)}")
    print(f"  zone updates queued: {len(cz_updates)}")
    print(f"  client rollups queued: {len(client_rollup)}")
    print(f"  client notes queued: {len(client_notes)}")
    print(f"  TX clients with no client_zones: {len(no_zone_records)}")

    if args.dry_run:
        print("\n--- DRY RUN: no writes ---")
        if unmatched:
            print(f"\nFirst 30 unmatched:")
            for r in unmatched[:30]:
                print(f"  [{r['market']}] {r['name']}")
        return

    # --- Write client_zones updates in chunks ---
    print("\nWriting client_zones updates...")
    cz_items = list(cz_updates.items())
    for i in range(0, len(cz_items), 200):
        batch = cz_items[i:i + 200]
        # Group by value to halve roundtrips
        by_val = defaultdict(list)
        for cz_id, val in batch:
            by_val[val].append(cz_id)
        for val, ids in by_val.items():
            sb.table("client_zones").update({"has_call_tracking": val}).in_("id", ids).execute()
    print(f"  wrote {len(cz_updates)} client_zones rows")

    # --- Write clients rollup + notes ---
    print("Writing clients (rollup + notes)...")
    # Combine rollup + notes per client to one update each
    client_updates = {}
    for cid, val in client_rollup.items():
        client_updates[cid] = {"has_call_tracking": val}
    for cid, notes in client_notes.items():
        client_updates.setdefault(cid, {})["call_tracking_notes"] = notes
    written = 0
    for cid, payload in client_updates.items():
        sb.table("clients").update(payload).eq("id", cid).execute()
        written += 1
        if written % 200 == 0:
            print(f"  {written}/{len(client_updates)}...")
    print(f"  wrote {written} client rows")

    # Print unmatched summary
    if unmatched:
        print(f"\n=== {len(unmatched)} UNMATCHED clients ===")
        for r in unmatched[:50]:
            print(f"  [{r['market']}] {r['name']}")
        if len(unmatched) > 50:
            print(f"  ... and {len(unmatched) - 50} more")

    print("\nDone.")


if __name__ == "__main__":
    main()
