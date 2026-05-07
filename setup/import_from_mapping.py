"""
Import client mapping spreadsheet into Supabase.

Reads Client_Mapping_Name_Cleaning_Updated.xlsx and populates:
  1. sales_reps — unique rep names from RepName column
  2. clients — unique OfficialName entries
  3. client_zones — links each client to their zone(s)
  4. client_platform_ids — maps each _GUID_THM to a client

Usage:
  python setup/import_from_mapping.py --dry-run   # Preview only, no writes
  python setup/import_from_mapping.py              # Run the import
"""

import sys
import os
import argparse
from pathlib import Path

# Add project root to path so we can import config
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
load_dotenv()

from supabase import create_client
from openpyxl import load_workbook

# --- Configuration ---

EXCEL_PATH = Path(__file__).resolve().parent.parent / "data" / "Client Mapping Name Cleaning Updated for Supabase.xlsx"

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

# Spreadsheet DB column -> market code in Supabase
DB_TO_MARKET_CODE = {
    "CO": "CO",
    "UT": "UT",
    "AU": "AU",
    "SA": "SA",
    # "XX" (Cross-Market) has no market — these are Uniqode entries without a market key
}

# Spreadsheet platform names -> Supabase platform enum values
PLATFORM_MAP = {
    "Mag Manager": "magazine_manager",
    "CallRail": "callrail",
    "Uniqode": "uniqode",
    "Inbox Advantage": "inbox_advantage",
}


def read_spreadsheet():
    """Read the mapping spreadsheet and return list of row dicts."""
    wb = load_workbook(str(EXCEL_PATH), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data = []
    for row in rows[1:]:
        record = dict(zip(headers, row))
        data.append(record)

    wb.close()
    return data


def analyze(rows):
    """Analyze the spreadsheet and print a summary."""
    total = len(rows)
    with_official = [r for r in rows if r.get("OfficialName") and str(r["OfficialName"]).strip()]
    without_official = total - len(with_official)
    unique_clients = set(str(r["OfficialName"]).strip() for r in with_official)
    unique_reps = set(
        str(r["RepName"]).strip()
        for r in with_official
        if r.get("RepName") and str(r["RepName"]).strip()
    )
    platforms = {}
    for r in with_official:
        p = r.get("Native System", "Unknown")
        platforms[p] = platforms.get(p, 0) + 1

    zones = {}
    for r in with_official:
        z = r.get("DB", "Unknown")
        zones[z] = zones.get(z, 0) + 1

    print(f"\n{'='*50}")
    print(f"  SPREADSHEET ANALYSIS")
    print(f"{'='*50}")
    print(f"  Total rows:              {total}")
    print(f"  Rows with OfficialName:  {len(with_official)}")
    print(f"  Rows skipped (no name):  {without_official}")
    print(f"  Unique clients:          {len(unique_clients)}")
    print(f"  Unique sales reps:       {len(unique_reps)}")
    print()
    print(f"  Platform breakdown:")
    for p, c in sorted(platforms.items()):
        print(f"    {p}: {c}")
    print()
    print(f"  Zone breakdown:")
    for z, c in sorted(zones.items()):
        print(f"    {z}: {c}")
    print(f"{'='*50}\n")

    return with_official, unique_clients, unique_reps


def run_import(rows, dry_run=True):
    """Import rows into Supabase. If dry_run=True, only analyze and preview."""
    valid_rows, unique_clients, unique_reps = analyze(rows)

    if dry_run:
        print("  DRY RUN — no data will be written to Supabase.\n")
        print(f"  Would create:")
        print(f"    {len(unique_reps)} sales reps")
        print(f"    {len(unique_clients)} clients")
        print(f"    Up to {len(valid_rows)} platform ID mappings")
        print(f"    Client-zone links for each client\n")

        # Show sample of what would be imported
        print("  Sample records (first 5 with OfficialName):")
        for r in valid_rows[:5]:
            print(f"    {r['OfficialName']!s:40s} | {r['DB']} | {r['Native System']:15s} | {r['_GUID_THM']}")
        print()

        # Show skipped rows summary
        skipped = [r for r in rows if not r.get("OfficialName") or not str(r.get("OfficialName", "")).strip()]
        if skipped:
            platforms_skipped = {}
            for r in skipped:
                p = r.get("Native System", "Unknown")
                platforms_skipped[p] = platforms_skipped.get(p, 0) + 1
            print(f"  Skipped rows by platform (no OfficialName):")
            for p, c in sorted(platforms_skipped.items()):
                print(f"    {p}: {c}")
        print()
        return

    # --- Live import ---
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Step 1: Load existing markets
    print("Step 1/4: Loading markets...")
    markets_result = sb.table("markets").select("*").execute()
    market_lookup = {m["code"]: m for m in markets_result.data}
    print(f"  Loaded {len(market_lookup)} markets: {list(market_lookup.keys())}")

    # Step 2: Create sales reps
    print("\nStep 2/4: Creating sales reps...")
    rep_lookup = {}  # rep name -> rep row
    # Load existing reps first
    existing_reps = sb.table("sales_reps").select("*").execute()
    for rep in existing_reps.data:
        rep_lookup[rep["name"]] = rep

    reps_created = 0
    reps_skipped = 0
    for rep_name in sorted(unique_reps):
        clean_name = rep_name.strip()
        if not clean_name or clean_name.startswith("*"):
            reps_skipped += 1
            continue
        if clean_name in rep_lookup:
            continue
        result = sb.table("sales_reps").insert({"name": clean_name}).execute()
        rep_lookup[clean_name] = result.data[0]
        reps_created += 1
    print(f"  Created {reps_created} reps, {len(rep_lookup)} total in DB")

    # Step 3: Create clients (unique by OfficialName)
    print("\nStep 3/4: Creating clients...")
    client_lookup = {}  # official name -> client row
    # Load existing clients first
    existing_clients = sb.table("clients").select("id,name,sales_rep_id,primary_market_id").execute()
    for c in existing_clients.data:
        client_lookup[c["name"]] = c

    # Build a map of client -> first market and first rep from spreadsheet
    client_meta = {}  # official_name -> {market_code, rep_name}
    for r in valid_rows:
        name = str(r["OfficialName"]).strip()
        if name not in client_meta:
            client_meta[name] = {
                "market_code": DB_TO_MARKET_CODE.get(r.get("DB")),
                "rep_name": str(r.get("RepName", "")).strip() if r.get("RepName") else None,
            }

    clients_created = 0
    for name in sorted(unique_clients):
        clean_name = name.strip()
        if not clean_name:
            continue
        if clean_name in client_lookup:
            continue

        meta = client_meta.get(clean_name, {})
        insert_data = {"name": clean_name}

        # Set primary market if we can resolve it
        market_code = meta.get("market_code")
        if market_code and market_code in market_lookup:
            insert_data["primary_market_id"] = market_lookup[market_code]["id"]

        # Set sales rep if we can resolve it
        rep_name = meta.get("rep_name")
        if rep_name and not rep_name.startswith("*") and rep_name in rep_lookup:
            insert_data["sales_rep_id"] = rep_lookup[rep_name]["id"]

        result = sb.table("clients").insert(insert_data).execute()
        client_lookup[clean_name] = result.data[0]
        clients_created += 1

        if clients_created % 100 == 0:
            print(f"  ... {clients_created} clients created")

    print(f"  Created {clients_created} clients, {len(client_lookup)} total in DB")

    # Step 4: Map platform IDs
    # Note: client_zones are populated from order data (import_orders.py), not from this mapping spreadsheet
    print("\nStep 4/4: Mapping platform IDs...")
    # Load existing platform mappings
    existing_pids = sb.table("client_platform_ids").select("external_id").execute()
    existing_ext_ids = set(r["external_id"] for r in existing_pids.data)

    pid_created = 0
    pid_skipped = 0
    for r in valid_rows:
        name = str(r["OfficialName"]).strip()
        guid = r.get("_GUID_THM")
        native_system = r.get("Native System")
        native_name = r.get("Native System Name")

        if name not in client_lookup:
            pid_skipped += 1
            continue
        if not guid or str(guid).strip() in existing_ext_ids:
            continue

        platform = PLATFORM_MAP.get(native_system)
        if not platform:
            pid_skipped += 1
            continue

        client_id = client_lookup[name]["id"]
        insert_data = {
            "client_id": client_id,
            "platform": platform,
            "external_id": str(guid).strip(),
        }
        if native_name:
            insert_data["external_name"] = str(native_name).strip()

        try:
            sb.table("client_platform_ids").insert(insert_data).execute()
            existing_ext_ids.add(str(guid).strip())
            pid_created += 1
        except Exception as e:
            err = str(e)
            if "duplicate" in err.lower() or "23505" in err:
                continue
            print(f"  WARNING: Failed to insert {guid}: {err[:100]}")
            pid_skipped += 1

        if pid_created % 200 == 0 and pid_created > 0:
            print(f"  ... {pid_created} platform IDs created")

    print(f"  Created {pid_created} platform ID mappings ({pid_skipped} skipped)")

    # Summary
    print(f"\n{'='*50}")
    print(f"  IMPORT COMPLETE")
    print(f"{'='*50}")
    print(f"  Sales reps created:      {reps_created}")
    print(f"  Clients created:         {clients_created}")
    print(f"  Platform ID mappings:    {pid_created}")
    print(f"{'='*50}\n")


def main():
    parser = argparse.ArgumentParser(description="Import client mapping spreadsheet into Supabase")
    parser.add_argument("--dry-run", action="store_true", help="Preview only, no writes to Supabase")
    args = parser.parse_args()

    if not SUPABASE_URL or not SUPABASE_KEY:
        print("ERROR: SUPABASE_URL and SUPABASE_KEY must be set in .env")
        sys.exit(1)

    if not EXCEL_PATH.exists():
        print(f"ERROR: Spreadsheet not found at {EXCEL_PATH}")
        sys.exit(1)

    print(f"Reading spreadsheet: {EXCEL_PATH.name}")
    rows = read_spreadsheet()
    run_import(rows, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
