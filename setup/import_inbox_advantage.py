"""
Import Inbox Advantage email campaign data from the weekly spreadsheet.

Reads the 'All 2' sheet from the IA Data spreadsheet and upserts into
the email_campaigns table. Idempotent — re-running with an updated file
will refresh metrics on existing campaigns and add new ones.

Usage:
  python setup/import_inbox_advantage.py --dry-run
  python setup/import_inbox_advantage.py
"""

import sys
import os
import argparse
from pathlib import Path
from datetime import datetime
from difflib import SequenceMatcher

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
load_dotenv()

from supabase import create_client
from openpyxl import load_workbook

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
DEFAULT_EXCEL = DATA_DIR / "IA Data 2024.12.5 CO,UT,TX Supabase Copy.xlsx"


def latest_ia_file():
    candidates = sorted(DATA_DIR.glob("IA Data*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None
SHEET_NAME = "All 2"

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

# Spreadsheet zone names -> our zone abbreviations
ZONE_MAP = {
    "denver s": "SD",
    "denver n": "ND",
    "northern co": "NOCO",
    "co springs": "EPC",
    "wasatch n": "NW",
    "wasatch c": "CW",
    "wasatch s": "SW",
    "austin n": "AN",
    "austin s": "AS",
    "san antonio e": "SAE",
    "san antonio w": "SAW",
}

# Column indexes (All 2 sheet)
COL = {
    "order_id": 0,
    "campaign_type": 1,
    "zone": 2,
    "state": 3,
    "client_name": 4,
    "drop_date": 5,
    "d1_date": 6,
    "audience": 7,
    "d1_views": 8,
    "d1_clicks": 9,
    "d10_date": 13,
    "d10_views": 14,
    "d10_clicks": 15,
    "d30_date": 19,
    "d30_views": 20,
    "d30_clicks": 21,
    "d30_view_pct": 22,
    "d30_click_pct": 23,
    "d30_ctv_pct": 24,
    "rate": 25,
}


def normalize(name):
    if not name:
        return ""
    n = str(name).lower().strip()
    for s in [", llc", ", inc", " llc", " inc", " co.", " co"]:
        n = n.replace(s, "")
    return n.strip()


def similarity(a, b):
    return SequenceMatcher(None, normalize(a), normalize(b)).ratio()


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s or s.startswith("="):
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def to_int(val):
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def to_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def to_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def extract_campaign_type(campaign_str):
    """Extract the campaign type (Exclusive/Sponsored/National) from '1-25 Exclusive'."""
    if not campaign_str:
        return None
    s = str(campaign_str).strip()
    for t in ["Exclusive", "Sponsored", "National"]:
        if t.lower() in s.lower():
            return t
    return None


def run(dry_run=True, workbook=None, source_label=None):
    """Parse the IA spreadsheet and upsert campaigns + client links.

    workbook: optional pre-loaded openpyxl Workbook (e.g. from a SharePoint
              download). If None, the file at module-global EXCEL_PATH is opened.
    source_label: human-readable string for log output (e.g. 'SharePoint:.../IA.xlsx').
    """
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Load lookups
    print("Loading lookups...")

    # Zones by abbreviation
    zones_result = sb.table("zones").select("id,abbreviation,market_id").execute()
    zone_by_abbrev = {z["abbreviation"]: z for z in zones_result.data if z.get("abbreviation")}
    print(f"  {len(zone_by_abbrev)} zones loaded")

    # Markets
    markets_result = sb.table("markets").select("id,code").execute()
    market_by_code = {m["code"]: m for m in markets_result.data}

    # All clients (paginated)
    print("  Loading clients...")
    all_clients = []
    offset = 0
    while True:
        batch = sb.table("clients").select("id,name,primary_market_id").range(offset, offset + 999).execute()
        all_clients.extend(batch.data)
        if len(batch.data) < 1000:
            break
        offset += 1000
    print(f"  {len(all_clients)} clients loaded")

    # Build name lookup, keyed by normalized name
    # Note: there can be duplicates (e.g. Apex Clean Air CO + Apex Clean Air UT after splits)
    clients_by_norm_name = {}
    for c in all_clients:
        key = normalize(c["name"])
        clients_by_norm_name.setdefault(key, []).append(c)

    # Read the spreadsheet
    if workbook is None:
        print(f"\nReading {EXCEL_PATH.name} -> {SHEET_NAME}...")
        wb = load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    else:
        print(f"\nReading {source_label or '(pre-loaded workbook)'} -> {SHEET_NAME}...")
        wb = workbook
    ws = wb[SHEET_NAME]

    # Group rows by order_id - one campaign can have multiple client rows
    campaigns_by_order = {}  # order_id -> {campaign_record, client_links: [(client_id, source_name), ...]}
    skipped_no_order_id = 0
    unmatched_zones = {}
    unmatched_clients = {}

    # Known aliases the fuzzy matcher can't reach (spreadsheet name -> DB client name)
    CLIENT_ALIASES = {
        "ABD": "ABD (Associates in Building + Design, Ltd.)",
        "Renovation By Burbach": "Burbach Exteriors (Renovation By Burbach)",
    }

    def lookup_client(client_name, market):
        """Returns client_id or None."""
        if not client_name:
            return None
        # Explicit alias first
        alias_target = CLIENT_ALIASES.get(client_name.strip())
        if alias_target:
            for c in all_clients:
                if c["name"] == alias_target:
                    return c["id"]
        norm_name = normalize(client_name)
        candidates = clients_by_norm_name.get(norm_name, [])

        # Prefix match
        if not candidates:
            for c in all_clients:
                db_norm = normalize(c["name"])
                if db_norm.startswith(norm_name + " ") and len(norm_name) >= 8:
                    candidates = [c]
                    clients_by_norm_name[norm_name] = candidates
                    break

        # Fuzzy fallback
        if not candidates:
            best_score = 0
            best_match = None
            for c in all_clients:
                score = similarity(client_name, c["name"])
                if score > best_score:
                    best_score = score
                    best_match = c
            if best_score >= 0.80 and best_match:
                candidates = [best_match]
                clients_by_norm_name[norm_name] = candidates

        if len(candidates) == 1:
            return candidates[0]["id"]
        if len(candidates) > 1:
            if market:
                matched = [c for c in candidates if c["primary_market_id"] == market["id"]]
                if matched:
                    return matched[0]["id"]
            return candidates[0]["id"]
        unmatched_clients[client_name] = unmatched_clients.get(client_name, 0) + 1
        return None

    for row in ws.iter_rows(min_row=2, values_only=True):
        order_id = row[COL["order_id"]]
        if not order_id:
            skipped_no_order_id += 1
            continue
        order_id = to_int(order_id)
        if not order_id:
            skipped_no_order_id += 1
            continue

        # Zone lookup
        zone_raw = to_str(row[COL["zone"]])
        zone_key = (zone_raw or "").lower().strip()
        zone = zone_by_abbrev.get(ZONE_MAP.get(zone_key, ""))
        if not zone:
            unmatched_zones[zone_raw] = unmatched_zones.get(zone_raw, 0) + 1

        # Market lookup
        state = to_str(row[COL["state"]])
        market_code = state if state in market_by_code else None
        if state == "TX" and zone:
            if zone["abbreviation"] in ("AN", "AS"):
                market_code = "AU"
            elif zone["abbreviation"] in ("SAE", "SAW"):
                market_code = "SA"
        market = market_by_code.get(market_code) if market_code else None
        if not market and zone:
            market = {"id": zone.get("market_id")}

        # Client lookup
        client_name = to_str(row[COL["client_name"]])
        client_id = lookup_client(client_name, market)

        # Build campaign record (only on first sighting of order_id)
        if order_id not in campaigns_by_order:
            campaigns_by_order[order_id] = {
                "campaign": {
                    "ia_order_id": order_id,
                    "zone_id": zone["id"] if zone else None,
                    "market_id": market["id"] if market else None,
                    "campaign_type": extract_campaign_type(row[COL["campaign_type"]]),
                    "campaign_name": to_str(row[COL["campaign_type"]]),
                    "drop_date": parse_date(row[COL["drop_date"]]),
                    "audience_size": to_int(row[COL["audience"]]),
                    "d1_date": parse_date(row[COL["d1_date"]]),
                    "d1_views": to_int(row[COL["d1_views"]]),
                    "d1_clicks": to_int(row[COL["d1_clicks"]]),
                    "d10_date": parse_date(row[COL["d10_date"]]),
                    "d10_views": to_int(row[COL["d10_views"]]),
                    "d10_clicks": to_int(row[COL["d10_clicks"]]),
                    "d30_date": parse_date(row[COL["d30_date"]]),
                    "d30_views": to_int(row[COL["d30_views"]]),
                    "d30_clicks": to_int(row[COL["d30_clicks"]]),
                    "d30_view_pct": to_float(row[COL["d30_view_pct"]]),
                    "d30_click_pct": to_float(row[COL["d30_click_pct"]]),
                    "d30_ctv_pct": to_float(row[COL["d30_ctv_pct"]]),
                    "rate": to_float(row[COL["rate"]]),
                    "source_zone": zone_raw,
                },
                "client_links": [],
            }
        if client_id:
            campaigns_by_order[order_id]["client_links"].append((client_id, client_name))

    wb.close()

    # Reporting
    total_campaigns = len(campaigns_by_order)
    total_links = sum(len(c["client_links"]) for c in campaigns_by_order.values())
    campaigns_with_clients = sum(1 for c in campaigns_by_order.values() if c["client_links"])
    sponsored_no_clients = sum(1 for c in campaigns_by_order.values() if not c["client_links"])

    print(f"\n{'='*60}")
    print(f"  IMPORT PLAN")
    print(f"{'='*60}")
    print(f"  Unique campaigns:           {total_campaigns}")
    print(f"  Campaigns with clients:     {campaigns_with_clients}")
    print(f"  Campaigns w/o clients:      {sponsored_no_clients}")
    print(f"  Total client links:         {total_links}")
    print(f"  Skipped (no order id):      {skipped_no_order_id}")
    print(f"{'='*60}")

    if unmatched_zones:
        print(f"\n  Unmatched zones ({len(unmatched_zones)}):")
        for z, c in sorted(unmatched_zones.items(), key=lambda x: -x[1]):
            print(f"    '{z}': {c} rows")

    if unmatched_clients:
        print(f"\n  Unmatched clients ({len(unmatched_clients)}):")
        for n, c in sorted(unmatched_clients.items(), key=lambda x: -x[1]):
            print(f"    '{n}': {c} rows")

    if dry_run:
        print("\n  DRY RUN - no data written")
        return {
            "campaigns_planned": total_campaigns,
            "campaigns_with_clients": campaigns_with_clients,
            "sponsored_no_clients": sponsored_no_clients,
            "links_planned": total_links,
            "skipped_no_order_id": skipped_no_order_id,
            "unmatched_zones_count": len(unmatched_zones),
            "unmatched_clients_count": len(unmatched_clients),
            "campaigns_upserted": 0,
            "links_upserted": 0,
        }

    # Upsert campaigns
    print(f"\nUpserting {total_campaigns} campaigns...")
    BATCH = 200
    campaign_records = [c["campaign"] for c in campaigns_by_order.values()]
    for i in range(0, len(campaign_records), BATCH):
        sb.table("email_campaigns").upsert(campaign_records[i:i + BATCH], on_conflict="ia_order_id").execute()
        if (i + BATCH) % 500 == 0 or i + BATCH >= len(campaign_records):
            print(f"  ... {min(i + BATCH, len(campaign_records))} campaigns upserted")

    # Re-fetch campaign IDs by ia_order_id (need them for the junction)
    print("Loading campaign IDs...")
    order_id_to_campaign_id = {}
    order_ids = list(campaigns_by_order.keys())
    for i in range(0, len(order_ids), 100):
        batch = order_ids[i:i + 100]
        result = sb.table("email_campaigns").select("id,ia_order_id").in_("ia_order_id", batch).execute()
        for r in result.data:
            order_id_to_campaign_id[r["ia_order_id"]] = r["id"]

    # Build junction rows, dedupe by (campaign_id, client_id)
    print("Building client links...")
    junction_rows = []
    seen = set()
    for order_id, info in campaigns_by_order.items():
        campaign_id = order_id_to_campaign_id.get(order_id)
        if not campaign_id:
            continue
        for client_id, client_name in info["client_links"]:
            key = (campaign_id, client_id)
            if key in seen:
                continue
            seen.add(key)
            junction_rows.append({
                "campaign_id": campaign_id,
                "client_id": client_id,
                "source_client_name": client_name,
            })

    print(f"Upserting {len(junction_rows)} client links...")
    for i in range(0, len(junction_rows), BATCH):
        sb.table("email_campaign_clients").upsert(
            junction_rows[i:i + BATCH], on_conflict="campaign_id,client_id"
        ).execute()

    print(f"\n{'='*60}")
    print(f"  COMPLETE")
    print(f"  {total_campaigns} campaigns, {len(junction_rows)} client links")
    print(f"{'='*60}")

    return {
        "campaigns_planned": total_campaigns,
        "campaigns_with_clients": campaigns_with_clients,
        "sponsored_no_clients": sponsored_no_clients,
        "links_planned": total_links,
        "skipped_no_order_id": skipped_no_order_id,
        "unmatched_zones_count": len(unmatched_zones),
        "unmatched_clients_count": len(unmatched_clients),
        "campaigns_upserted": total_campaigns,
        "links_upserted": len(junction_rows),
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--file", type=str, help="Path to IA Data Excel file")
    args = parser.parse_args()

    global EXCEL_PATH
    if args.file:
        EXCEL_PATH = Path(args.file)
    else:
        EXCEL_PATH = latest_ia_file() or DEFAULT_EXCEL

    if not all([SUPABASE_URL, SUPABASE_KEY]):
        print("ERROR: Missing env vars")
        sys.exit(1)
    if not EXCEL_PATH.exists():
        print(f"ERROR: File not found at {EXCEL_PATH}")
        sys.exit(1)

    run(dry_run=args.dry_run)


if __name__ == "__main__":
    main()
