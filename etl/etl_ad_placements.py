"""Ad Placements ETL — pull CO + UT placement spreadsheets from SharePoint,
detect "no updates since last run" per market, parse, and upsert.

Wraps setup.import_ad_placements with two additions:
  * --source switch (local file path vs SharePoint download)
  * Per-market no-op detection — skip parsing files unchanged since the
    most recent successful run. Common case is "everything's the same",
    in which case we exit cleanly with rows_upserted=0 and a notes blob
    that explains the skip.

The script never raises on no-op — that's the expected weekly outcome.

Usage:
  python etl/etl_ad_placements.py --source local       --dry-run
  python etl/etl_ad_placements.py --source sharepoint  --dry-run
  python etl/etl_ad_placements.py --source sharepoint
"""

import os
import sys
import socket
import argparse
import traceback
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
load_dotenv()

from openpyxl import load_workbook
from supabase import create_client

from etl.sharepoint_client import (
    SharePointClient,
    SharePointAuthError,
    SharePointAPIError,
)
import setup.import_ad_placements as legacy

ETL_NAME = "ad_placements"
SP_SITE_URL = "https://thehomemagwest.sharepoint.com/sites/Sales"
SP_LIBRARY_NAME = "Sales (Corner)"

# (market_code, file_path_inside_library). Add TX here when its
# spreadsheet is created.
SHAREPOINT_TARGETS = [
    ("CO", "Design CO/THM Colorado Ad Placement.xlsx"),
    ("UT", "Design UT/THM Utah Ad Placement.xlsx"),
]

# Header column expectations. Both files use these 4 columns; the only
# difference is whether there's a title row above the headers, which we
# tolerate by accepting the header on either row 1 or row 2.
EXPECTED_HEADERS = ["Issue", "Zone", "Name", "Page"]


# ---------- header validation ----------

def find_and_validate_headers(ws) -> tuple:
    """Return (header_row_idx, mismatches). header_row_idx=None means not found.

    Looks at rows 1-3 for a row whose first 4 cells (case-insensitive trimmed)
    match EXPECTED_HEADERS exactly. Tolerates either the CO layout
    (title row 1, headers row 2) or the UT layout (headers row 1, no title).
    """
    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    for idx, row in enumerate(rows, start=1):
        if not row or len(row) < 4:
            continue
        cells = [(str(c) if c is not None else "").strip().lower() for c in row[:4]]
        expected = [h.lower() for h in EXPECTED_HEADERS]
        if cells == expected:
            return idx, []
    actual_top_rows = [list(r[:6]) if r else [] for r in rows]
    return None, [
        f"Expected first 4 columns to be {EXPECTED_HEADERS} on row 1 or 2; "
        f"saw rows: {actual_top_rows}"
    ]


# ---------- audit logging ----------

def _audit_client():
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_KEY")
    if not url or not key:
        return None
    return create_client(url, key)


def audit_start(sb, source: str, dry_run: bool) -> str | None:
    if sb is None:
        return None
    try:
        row = {
            "etl_name": ETL_NAME,
            "source": source,
            "dry_run": dry_run,
            "host": socket.gethostname(),
            "github_run_id": os.getenv("GITHUB_RUN_ID"),
            "github_run_url": (
                f"{os.getenv('GITHUB_SERVER_URL', '')}/"
                f"{os.getenv('GITHUB_REPOSITORY', '')}/actions/runs/"
                f"{os.getenv('GITHUB_RUN_ID', '')}"
                if os.getenv("GITHUB_RUN_ID") else None
            ),
        }
        result = sb.table("etl_runs").insert(row).execute()
        return result.data[0]["id"] if result.data else None
    except Exception as e:
        print(f"  WARNING: audit start failed: {type(e).__name__}: {e}")
        return None


def audit_finish(sb, run_id: str | None, *, success: bool,
                 counts: dict | None = None,
                 source_meta: dict | None = None,
                 notes: dict | None = None,
                 error_stage: str | None = None,
                 error_message: str | None = None) -> None:
    if sb is None or run_id is None:
        return
    try:
        update = {
            "finished_at": datetime.now(timezone.utc).isoformat(),
            "success": success,
            "error_stage": error_stage,
            "error_message": error_message,
        }
        if source_meta:
            update.update(source_meta)
        if counts:
            update["rows_read"] = counts.get("rows_planned")
            update["rows_upserted_campaigns"] = counts.get("rows_upserted")  # placements use campaigns column
            update["unmatched_clients"] = counts.get("unique_unmatched_client_names")
        if notes is not None:
            update["notes"] = notes
        sb.table("etl_runs").update(update).eq("id", run_id).execute()
    except Exception as e:
        print(f"  WARNING: audit finish failed: {type(e).__name__}: {e}")


def get_last_per_market_modified(sb) -> dict:
    """Return {market_code: lastModifiedDateTime ISO string} from the most
    recent successful, non-dry-run, non-skipped placement run for each market.

    Used to detect 'no updates since last run' and short-circuit.
    """
    if sb is None:
        return {}
    out: dict = {}
    try:
        # Pull recent successful real runs; walk backwards until we have a
        # real (not skipped) modified-at timestamp for each market.
        result = (
            sb.table("etl_runs")
            .select("notes")
            .eq("etl_name", ETL_NAME)
            .eq("success", True)
            .eq("dry_run", False)
            .order("started_at", desc=True)
            .limit(20)
            .execute()
        )
        for row in result.data or []:
            notes = row.get("notes") or {}
            files = notes.get("files") or {}
            for mc, info in files.items():
                if mc in out:
                    continue
                modified = info.get("modified_at")
                # Skip-only entries are tracked but don't represent a fresh
                # successful parse, so still use their modified_at as the
                # baseline (it was unchanged at the time of skip).
                if modified:
                    out[mc] = modified
            if all(mc in out for mc, _ in SHAREPOINT_TARGETS):
                break
    except Exception as e:
        print(f"  WARNING: per-market history lookup failed: {type(e).__name__}: {e}")
    return out


# ---------- email failure ----------

def email_failure(subject: str, html_body: str) -> None:
    try:
        from scripts.import_report import send_email
        send_email(subject, html_body)
    except Exception as e:
        print(f"  WARNING: failed to send failure email: {type(e).__name__}: {e}")


# ---------- workbook fetchers ----------

def fetch_workbooks_sharepoint() -> list:
    """Returns a list of dicts, one per market, with keys:
    market_code, workbook (or None if skipped), label, source_meta,
    skip_reason (str or None).
    """
    sp = SharePointClient()
    site_id = sp.resolve_site_id(SP_SITE_URL)
    drive_id = sp.resolve_drive_id(site_id, SP_LIBRARY_NAME)

    sb_audit = _audit_client()
    last_modified_by_market = get_last_per_market_modified(sb_audit)
    print("  Last source_modified_at by market (from etl_runs):")
    for mc, _ in SHAREPOINT_TARGETS:
        print(f"    {mc}: {last_modified_by_market.get(mc) or '(no prior run)'}")

    out = []
    for market_code, file_path in SHAREPOINT_TARGETS:
        print(f"\n  Fetching metadata: {market_code} -> {file_path}")
        meta = sp.get_file_metadata(site_id, file_path, drive_id=drive_id)
        modified_at = meta.get("lastModifiedDateTime")
        modified_by = (meta.get("lastModifiedBy") or {}).get("user", {}).get("displayName")
        size = meta.get("size")
        print(f"    Modified: {modified_at} by {modified_by or '(unknown)'} ({size:,} bytes)")

        prior_modified = last_modified_by_market.get(market_code)
        if prior_modified and modified_at and prior_modified == modified_at:
            print(f"    UNCHANGED since last run — skipping parse + upsert.")
            out.append({
                "market_code": market_code,
                "workbook": None,
                "label": file_path,
                "source_meta": {
                    "path": file_path,
                    "modified_at": modified_at,
                    "modified_by": modified_by,
                    "size": size,
                },
                "skip_reason": "unchanged_since_last_run",
            })
            continue

        # Modified or new — download
        print(f"    Downloading...")
        data = sp.download_file_bytes(site_id, file_path, drive_id=drive_id)
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        out.append({
            "market_code": market_code,
            "workbook": wb,
            "label": f"SharePoint:{file_path}",
            "source_meta": {
                "path": file_path,
                "modified_at": modified_at,
                "modified_by": modified_by,
                "size": size,
            },
            "skip_reason": None,
        })
    return out


def fetch_workbooks_local() -> list:
    """For local --source mode. No no-op detection (always re-parses)."""
    out = []
    for path, market_code in legacy.detect_placement_files() or legacy.DEFAULT_FILES:
        if not Path(path).exists():
            print(f"  WARNING: {path} not found, skipping")
            continue
        print(f"  Reading local: {path}")
        wb = load_workbook(str(path), read_only=True, data_only=True)
        stat = Path(path).stat()
        out.append({
            "market_code": market_code,
            "workbook": wb,
            "label": f"local:{Path(path).name}",
            "source_meta": {
                "path": str(path),
                "modified_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
                "modified_by": None,
                "size": stat.st_size,
            },
            "skip_reason": None,
        })
    return out


# ---------- main ----------

def main():
    parser = argparse.ArgumentParser(description=__doc__.split("\n", 1)[0])
    parser.add_argument("--source", choices=["local", "sharepoint"], default="local")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--no-email", action="store_true")
    parser.add_argument("--no-audit", action="store_true")
    parser.add_argument("--force", action="store_true",
                        help="Force re-process all files even if unchanged.")
    args = parser.parse_args()

    print("=" * 70)
    print(f"Ad Placements ETL — source={args.source}  dry_run={args.dry_run}  force={args.force}")
    print("=" * 70)

    sb_audit = None if args.no_audit else _audit_client()
    run_id = audit_start(sb_audit, source=args.source, dry_run=args.dry_run)
    if run_id:
        print(f"  audit run_id: {run_id}")

    files_notes: dict = {}
    counts = None
    error_stage = None
    error_message = None

    try:
        # 1) Fetch workbooks
        print(f"\n[1/3] Fetching workbooks from {args.source}...")
        try:
            if args.source == "sharepoint":
                fetched = fetch_workbooks_sharepoint()
            else:
                fetched = fetch_workbooks_local()
        except (SharePointAuthError, SharePointAPIError, RuntimeError) as e:
            error_stage = "fetch"
            error_message = f"{type(e).__name__}: {e}"
            print(f"\n  ABORT: {error_message}")
            if not args.no_email and not args.dry_run:
                email_failure(
                    "[ALERT] Ad Placements ETL aborted — fetch failed",
                    f"<p>Fetch failed during placements ETL.</p><pre>{error_message}</pre>"
                )
            sys.exit(1)

        # If --force, override skip_reason for all sources
        if args.force:
            for src in fetched:
                if src["skip_reason"] == "unchanged_since_last_run" and src["workbook"] is None:
                    # Need to re-download — but in --force mode we already had a workbook?
                    # In SharePoint mode, no-op skip means we didn't download. Force re-download.
                    print(f"  --force: re-downloading {src['label']}")
                    sp = SharePointClient()
                    site_id = sp.resolve_site_id(SP_SITE_URL)
                    drive_id = sp.resolve_drive_id(site_id, SP_LIBRARY_NAME)
                    data = sp.download_file_bytes(
                        site_id, src["source_meta"]["path"], drive_id=drive_id
                    )
                    src["workbook"] = load_workbook(BytesIO(data), read_only=True, data_only=True)
                    src["label"] = f"SharePoint:{src['source_meta']['path']} (--force)"
                    src["skip_reason"] = None

        # Build per-file notes blob
        for src in fetched:
            files_notes[src["market_code"]] = {
                **src["source_meta"],
                "skip_reason": src["skip_reason"],
            }

        # 2) Validate headers on the workbooks we did download
        print(f"\n[2/3] Validating headers...")
        sources_for_legacy = []
        for src in fetched:
            if src["workbook"] is None:
                continue  # skipped
            wb = src["workbook"]
            ws = wb[wb.sheetnames[0]]
            header_row, mismatches = find_and_validate_headers(ws)
            if mismatches:
                error_stage = "validate"
                error_message = (
                    f"header mismatch in {src['label']}: " + "; ".join(mismatches)
                )
                print(f"  ABORT: {error_message}")
                if not args.no_email and not args.dry_run:
                    email_failure(
                        "[ALERT] Ad Placements ETL aborted — header drift",
                        f"<p>Header validation failed for {src['label']}.</p>"
                        f"<pre>{error_message}</pre>"
                    )
                sys.exit(1)
            print(f"  OK [{src['market_code']}]: header on row {header_row}, "
                  f"sheet={ws.title!r}")
            sources_for_legacy.append({
                "workbook": wb,
                "market_code": src["market_code"],
                "label": src["label"],
            })

        # 3) Run parse + upsert (only on changed files)
        if not sources_for_legacy:
            print(f"\n[3/3] All files unchanged since last successful run. "
                  f"Nothing to do.")
            counts = {"rows_planned": 0, "rows_upserted": 0, "unique_unmatched_client_names": 0}
        else:
            print(f"\n[3/3] Parsing {len(sources_for_legacy)} file(s)...")
            try:
                counts = legacy.run(dry_run=args.dry_run, sources=sources_for_legacy)
            except Exception as e:
                error_stage = "upsert" if not args.dry_run else "parse"
                error_message = f"{type(e).__name__}: {e}"
                tb = traceback.format_exc(limit=4)
                print(f"\n  ABORT: {error_message}\n{tb}")
                if not args.no_email and not args.dry_run:
                    email_failure(
                        "[ALERT] Ad Placements ETL failed during parse/upsert",
                        f"<p>Unexpected error in placements ETL.</p>"
                        f"<pre>{error_message}\n\n{tb}</pre>"
                    )
                sys.exit(1)

        print("\nAd Placements ETL completed successfully.")
    finally:
        # Pick the most-recent modified_at across files for the top-level audit column
        max_modified = max(
            (n.get("modified_at") for n in files_notes.values() if n.get("modified_at")),
            default=None,
        )
        modified_by_for_max = next(
            (n.get("modified_by") for n in files_notes.values()
             if n.get("modified_at") == max_modified),
            None,
        )
        total_size = sum(n.get("size") or 0 for n in files_notes.values())

        success = error_stage is None
        audit_finish(
            sb_audit, run_id,
            success=success,
            counts=counts,
            source_meta={
                "source_path": ", ".join(n.get("path") or "" for n in files_notes.values()),
                "source_modified_at": max_modified,
                "source_modified_by": modified_by_for_max,
                "source_size_bytes": total_size or None,
            },
            notes={"files": files_notes},
            error_stage=error_stage,
            error_message=error_message,
        )


if __name__ == "__main__":
    main()
