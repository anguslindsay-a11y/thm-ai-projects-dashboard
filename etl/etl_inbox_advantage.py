"""Inbox Advantage ETL — pull from SharePoint (or local file), parse, upsert.

This is a thin wrapper around setup.import_inbox_advantage that adds:
  * a --source switch for choosing local file vs SharePoint download,
  * a header-row validator that aborts loudly on column drift, and
  * an email failure alert when a real (non-dry-run) run aborts pre-write.

Parsing/upsert logic is unchanged — same idempotent behavior on
email_campaigns.ia_order_id and email_campaign_clients.(campaign_id, client_id).

Usage:
  python etl/etl_inbox_advantage.py --source local      --dry-run
  python etl/etl_inbox_advantage.py --source sharepoint --dry-run
  python etl/etl_inbox_advantage.py --source sharepoint
"""

import os
import socket
import sys
import argparse
import traceback
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
load_dotenv()

from openpyxl import load_workbook
from supabase import create_client

from etl.sharepoint_client import (
    SharePointClient,
    SharePointAuthError,
    SharePointAPIError,
)
import setup.import_inbox_advantage as legacy

ETL_NAME = "inbox_advantage"

SHEET_NAME = "All 2"

# Column-by-column expectations for row 1 of the 'All 2' sheet. The legacy
# importer indexes columns positionally — if Inside Sales reorders the sheet,
# every value would map to the wrong field silently. We compare row 1 against
# this map and abort hard on any mismatch.
#
# Match is case-insensitive startswith, which tolerates the existing quirks
# ('Campaign Type ' has a trailing space, 'Audience x' has a trailing token).
EXPECTED_HEADERS = {
    0:  "Order ID",
    1:  "Campaign Type",
    2:  "Zone",
    3:  "State",
    4:  "Client Name",
    5:  "Drop Date",
    6:  "D1 Date",
    7:  "Audience",
    8:  "D1 Views",
    9:  "D1 Clicks",
    13: "D10 Date",
    14: "D10 Views",
    15: "D10 Clicks",
    19: "D30",
    20: "D30 Views",
    21: "D30 Clicks",
    22: "D30 View %",
    23: "D30 Click %",
    24: "D30 C/V %",
    25: "Rate",
}


def validate_headers(ws) -> list[str]:
    """Return a list of column-mismatch messages. Empty list = all good."""
    actual = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    mismatches = []
    for idx, expected in EXPECTED_HEADERS.items():
        if idx >= len(actual):
            mismatches.append(
                f"col {idx}: missing (sheet has only {len(actual)} cols)"
            )
            continue
        got = (str(actual[idx]) if actual[idx] is not None else "").strip().lower()
        want = expected.strip().lower()
        if not got.startswith(want):
            mismatches.append(
                f"col {idx}: expected '{expected}' (startswith), got {actual[idx]!r}"
            )
    return mismatches


def email_failure(subject: str, html_body: str) -> None:
    """Best-effort failure alert. Never raises — failing to email a failure
    must not mask the original failure."""
    try:
        from scripts.import_report import send_email
        send_email(subject, html_body)
    except Exception as e:
        print(f"  WARNING: failed to send failure email: {type(e).__name__}: {e}")


def fetch_workbook_sharepoint() -> tuple:
    """Download IA xlsx from SharePoint. Returns (workbook, source_label, meta_dict)."""
    site_url = os.getenv("SHAREPOINT_IA_SITE_URL")
    file_path = os.getenv("SHAREPOINT_IA_FILE_PATH")
    if not site_url or not file_path:
        raise RuntimeError(
            "SHAREPOINT_IA_SITE_URL and/or SHAREPOINT_IA_FILE_PATH not set in .env"
        )

    sp = SharePointClient()
    site_id = sp.resolve_site_id(site_url)
    print(f"  Site URL:  {site_url}")
    print(f"  File path: {file_path}")
    meta = sp.get_file_metadata(site_id, file_path)
    size = meta.get("size")
    modified = meta.get("lastModifiedDateTime")
    modified_by = (meta.get("lastModifiedBy") or {}).get("user", {}).get("displayName")
    print(f"  Size:      {size:,} bytes" if size else "  Size:      (unknown)")
    print(f"  Modified:  {modified} by {modified_by or '(unknown)'}")

    data = sp.download_file_bytes(site_id, file_path)
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    return wb, f"SharePoint:{file_path}", {
        "source_path": file_path,
        "source_modified_at": modified,
        "source_modified_by": modified_by,
        "source_size_bytes": size,
    }


def fetch_workbook_local() -> tuple:
    """Open latest IA xlsx under data/. Returns (workbook, source_label, meta_dict)."""
    path = legacy.latest_ia_file() or legacy.DEFAULT_EXCEL
    if not Path(path).exists():
        raise RuntimeError(f"No local IA file found in {legacy.DATA_DIR}")
    print(f"  Local file: {path}")
    wb = load_workbook(str(path), read_only=True, data_only=True)
    stat = Path(path).stat()
    return wb, f"local:{Path(path).name}", {
        "source_path": str(path),
        "source_modified_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
        "source_modified_by": None,
        "source_size_bytes": stat.st_size,
    }


# ---------- audit logging ----------

def _audit_client():
    """Create a Supabase client just for etl_runs writes. Separate from the
    one inside legacy.run() so a failure during legacy run can still write
    a finalize row."""
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_KEY")
    if not url or not key:
        return None
    return create_client(url, key)


def audit_start(sb, source: str, dry_run: bool) -> str | None:
    """Insert a starting row in etl_runs and return its id (or None on failure).

    A failure to log MUST NOT block the actual ETL from running — return None
    and the wrapper continues. Better to do the work without an audit trail
    than to skip the work because audit logging itself broke.
    """
    if sb is None:
        return None
    try:
        row = {
            "etl_name": ETL_NAME,
            "source": source,
            "dry_run": dry_run,
            "host": socket.gethostname(),
            "github_run_id": os.getenv("GITHUB_RUN_ID"),
            "github_run_url": (
                f"{os.getenv('GITHUB_SERVER_URL', '')}/"
                f"{os.getenv('GITHUB_REPOSITORY', '')}/actions/runs/"
                f"{os.getenv('GITHUB_RUN_ID', '')}"
                if os.getenv("GITHUB_RUN_ID") else None
            ),
        }
        result = sb.table("etl_runs").insert(row).execute()
        return result.data[0]["id"] if result.data else None
    except Exception as e:
        print(f"  WARNING: audit start failed: {type(e).__name__}: {e}")
        return None


def audit_finish(sb, run_id: str | None, *, success: bool,
                 counts: dict | None = None,
                 source_meta: dict | None = None,
                 error_stage: str | None = None,
                 error_message: str | None = None) -> None:
    """Update the audit row with finish state. Never raises."""
    if sb is None or run_id is None:
        return
    try:
        update = {
            "finished_at": datetime.now(timezone.utc).isoformat(),
            "success": success,
            "error_stage": error_stage,
            "error_message": error_message,
        }
        if source_meta:
            update.update(source_meta)
        if counts:
            update["rows_upserted_campaigns"] = counts.get("campaigns_upserted")
            update["rows_upserted_links"] = counts.get("links_upserted")
            update["unmatched_clients"] = counts.get("unmatched_clients_count")
            update["unmatched_zones"] = counts.get("unmatched_zones_count")
            update["rows_read"] = counts.get("campaigns_planned")
            update["notes"] = {
                "campaigns_with_clients": counts.get("campaigns_with_clients"),
                "sponsored_no_clients": counts.get("sponsored_no_clients"),
                "links_planned": counts.get("links_planned"),
                "skipped_no_order_id": counts.get("skipped_no_order_id"),
            }
        sb.table("etl_runs").update(update).eq("id", run_id).execute()
    except Exception as e:
        print(f"  WARNING: audit finish failed: {type(e).__name__}: {e}")


def main():
    parser = argparse.ArgumentParser(description=__doc__.split("\n", 1)[0])
    parser.add_argument(
        "--source", choices=["local", "sharepoint"], default="local",
        help="Where to read the IA spreadsheet from. Default: local."
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="Parse and report only — no DB writes.")
    parser.add_argument("--no-email", action="store_true",
                        help="Suppress failure emails (useful for local debugging).")
    parser.add_argument("--no-audit", action="store_true",
                        help="Skip writing to etl_runs (useful for local debugging).")
    args = parser.parse_args()

    print("=" * 70)
    print(f"IA ETL — source={args.source}  dry_run={args.dry_run}")
    print("=" * 70)

    # Open audit row immediately so even early failures are recorded
    sb_audit = None if args.no_audit else _audit_client()
    run_id = audit_start(sb_audit, source=args.source, dry_run=args.dry_run)
    if run_id:
        print(f"  audit run_id: {run_id}")

    source_meta: dict = {}
    counts: dict | None = None
    error_stage: str | None = None
    error_message: str | None = None

    try:
        # 1) Fetch the workbook
        print(f"\n[1/3] Fetching workbook from {args.source}...")
        try:
            if args.source == "sharepoint":
                wb, source_label, source_meta = fetch_workbook_sharepoint()
            else:
                wb, source_label, source_meta = fetch_workbook_local()
        except (SharePointAuthError, SharePointAPIError, RuntimeError) as e:
            error_stage = "fetch"
            error_message = f"{type(e).__name__}: {e}"
            print(f"\n  ABORT: {error_message}")
            if not args.no_email and not args.dry_run:
                email_failure(
                    "[ALERT] IA ETL aborted — fetch failed",
                    f"<p>The IA ETL was unable to fetch the source spreadsheet.</p>"
                    f"<pre>{error_message}</pre>"
                    f"<p><b>No data written.</b> Check SharePoint access and credentials.</p>"
                )
            sys.exit(1)

        # 2) Validate header row
        print(f"\n[2/3] Validating header row on '{SHEET_NAME}'...")
        if SHEET_NAME not in wb.sheetnames:
            error_stage = "validate"
            error_message = (
                f"Sheet '{SHEET_NAME}' not found in workbook. "
                f"Sheets present: {wb.sheetnames}"
            )
            print(f"  ABORT: {error_message}")
            if not args.no_email and not args.dry_run:
                email_failure(
                    "[ALERT] IA ETL aborted — sheet missing",
                    f"<p>The expected '{SHEET_NAME}' sheet was not found in the IA workbook.</p>"
                    f"<pre>{error_message}</pre>"
                    f"<p><b>No data written.</b></p>"
                )
            sys.exit(1)

        ws = wb[SHEET_NAME]
        mismatches = validate_headers(ws)
        if mismatches:
            error_stage = "validate"
            error_message = "header drift: " + "; ".join(mismatches)
            print(f"  ABORT: header drift detected ({len(mismatches)} column(s)):")
            for m in mismatches:
                print(f"    {m}")
            if not args.no_email and not args.dry_run:
                list_html = "".join(f"<li>{m}</li>" for m in mismatches)
                email_failure(
                    "[ALERT] IA ETL aborted — column header drift",
                    f"<p>Column headers in the IA spreadsheet do not match the expected layout. "
                    f"Hardcoded column indexes would map to the wrong fields. "
                    f"<b>Run aborted, no data written.</b></p>"
                    f"<p>Source: <code>{source_label}</code></p>"
                    f"<ul>{list_html}</ul>"
                    f"<p>Either fix the spreadsheet header row, or update <code>EXPECTED_HEADERS</code> "
                    f"in <code>etl/etl_inbox_advantage.py</code> and <code>COL</code> in "
                    f"<code>setup/import_inbox_advantage.py</code> together.</p>"
                )
            sys.exit(1)
        print(f"  OK — all {len(EXPECTED_HEADERS)} expected columns present in correct positions")

        # 3) Hand off to the existing parse + upsert pipeline
        print(f"\n[3/3] Running parse + upsert (source: {source_label})...")
        try:
            counts = legacy.run(dry_run=args.dry_run, workbook=wb, source_label=source_label)
        except Exception as e:
            error_stage = "upsert" if not args.dry_run else "parse"
            error_message = f"{type(e).__name__}: {e}"
            tb = traceback.format_exc(limit=4)
            print(f"\n  ABORT: {error_message}\n{tb}")
            if not args.no_email and not args.dry_run:
                email_failure(
                    "[ALERT] IA ETL failed during parse/upsert",
                    f"<p>The IA ETL hit an unexpected error after passing pre-flight checks.</p>"
                    f"<pre>{error_message}\n\n{tb}</pre>"
                    f"<p>Source: <code>{source_label}</code></p>"
                    f"<p>Check etl_runs for run_id <code>{run_id}</code> and the audit log for context.</p>"
                )
            sys.exit(1)

        print("\nIA ETL completed successfully.")
    finally:
        success = error_stage is None
        audit_finish(
            sb_audit, run_id,
            success=success,
            counts=counts,
            source_meta=source_meta,
            error_stage=error_stage,
            error_message=error_message,
        )


if __name__ == "__main__":
    main()
